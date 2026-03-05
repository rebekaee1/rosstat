"""
Калибровка матрицы B на РЕАЛЬНЫХ данных Росстата.

Данные:
  1. Симметричная ТЗВ 2021 (baz_tzv-2021.xlsx) → матрица A, выпуск X, ВДС
  2. ИЦП по ВЭД (Proizvoditeli_Ind_VED.xlsx)  → P_t квартальные
  3. Индексы физического объёма ВДС (VDS_kvartal_OKVED2.xlsx, лист 6) → Q_t квартальные

Метод:
  - Первые разности: ΔQ_t = B·ΔP_t + ε
  - OLS, Ridge (с приоритетом из формулы 4.2), проверка отрицательной определённости
"""

import os
import re
import numpy as np
import openpyxl
from numpy.linalg import eigvalsh

np.set_printoptions(precision=5, suppress=True, linewidth=140)

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
LABELS = ['Primary', 'Secondary', 'Tertiary']

# ═══════════════════════════════════════════════════════
#  Утилиты
# ═══════════════════════════════════════════════════════

def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(',', '.').replace('\xa0', '').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _leading_code(raw_code):
    if not raw_code:
        return None
    m = re.match(r'(\d+)', raw_code.strip())
    return int(m.group(1)) if m else None


def classify_sector(code_str):
    lc = _leading_code(code_str)
    if lc is None:
        return -1
    if lc <= 9:
        return 0
    if lc <= 43:
        return 1
    return 2


# ═══════════════════════════════════════════════════════
#  1. Таблица «затраты-выпуск» 2021 → 3 сектора
# ═══════════════════════════════════════════════════════

def parse_and_aggregate_io():
    path = os.path.join(DATA_DIR, 'baz_tzv-2021.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Симм ТЗВ']

    N = 113
    COL0 = 4
    ROW_GVA = 128
    ROW_OUTPUT = 129

    all_rows = {}
    for row in ws.iter_rows(min_row=1, max_row=ROW_OUTPUT, max_col=COL0 + N, values_only=False):
        r = row[0].row
        all_rows[r] = [cell.value for cell in row]
    wb.close()

    codes = [str(all_rows[5 + i][1] or '') for i in range(N)]
    sectors = np.array([classify_sector(c) for c in codes])

    Z = np.zeros((N, N))
    for i in range(N):
        for j in range(N):
            Z[i, j] = safe_float(all_rows[5 + i][COL0 - 1 + j])

    X_total = np.array([safe_float(all_rows[ROW_OUTPUT][COL0 - 1 + j]) for j in range(N)])
    VA = np.array([safe_float(all_rows[ROW_GVA][COL0 - 1 + j]) for j in range(N)])

    ns = 3
    Z_agg = np.zeros((ns, ns))
    X_agg = np.zeros(ns)
    VA_agg = np.zeros(ns)
    for s in range(ns):
        ms = sectors == s
        X_agg[s] = X_total[ms].sum()
        VA_agg[s] = VA[ms].sum()
        for t in range(ns):
            mt = sectors == t
            Z_agg[s, t] = Z[np.ix_(ms, mt)].sum()

    A_agg = Z_agg / X_agg[np.newaxis, :]

    print("═" * 70)
    print(" IO-ТАБЛИЦА 2021 (агрегация → 3 сектора)")
    print("═" * 70)
    for s in range(ns):
        print(f"  {LABELS[s]:12s}: X={X_agg[s]:>14,.0f}  VA={VA_agg[s]:>14,.0f}")
    print(f"\n  Матрица A:")
    for i in range(ns):
        print(f"    {LABELS[i]:12s}: {' '.join(f'{A_agg[i,j]:>8.4f}' for j in range(ns))}")

    return dict(A=A_agg, X=X_agg, VA=VA_agg, Z=Z_agg)


# ═══════════════════════════════════════════════════════
#  2. Ценовые индексы (ИЦП) → квартальные P_t
# ═══════════════════════════════════════════════════════

def parse_ppi_sheet(ws, yr_row, m_start, yr_col):
    data = {}
    for row in ws.iter_rows(min_row=yr_row, max_row=m_start + 11, values_only=False):
        r = row[0].row
        data[r] = [cell.value for cell in row]

    yr_cells = data[yr_row]
    years = {}
    ci = yr_col - 1
    while ci < len(yr_cells):
        raw = yr_cells[ci]
        if raw is None:
            break
        m_yr = re.match(r'(\d{4})', str(raw).strip())
        if not m_yr:
            break
        yr = int(m_yr.group(1))
        months = []
        for m in range(12):
            rd = data.get(m_start + m, [])
            v = safe_float(rd[ci] if ci < len(rd) else None)
            months.append(v)
        years[yr] = months
        ci += 1
    return years


def build_cumulative(monthly_pct, base_year=2019):
    all_years = sorted(monthly_pct.keys())
    cum = {}
    prev = 1.0
    for yr in all_years:
        vals = monthly_pct[yr]
        yr_idx = np.zeros(12)
        for m in range(12):
            pct = vals[m]
            if pct == 0:
                yr_idx[m] = prev
            else:
                prev = prev * (pct / 100.0)
                yr_idx[m] = prev
        cum[yr] = yr_idx

    if base_year in cum:
        base_val = cum[base_year][:3].mean()
    else:
        base_val = 1.0
    for yr in cum:
        cum[yr] = cum[yr] / base_val

    quarterly = {}
    for yr in cum:
        quarterly[yr] = np.array([cum[yr][i*3:(i+1)*3].mean() for i in range(4)])
    return quarterly


def get_sector_prices():
    path_ppi = os.path.join(DATA_DIR, 'Proizvoditeli_Ind_VED.xlsx')
    wb = openpyxl.load_workbook(path_ppi, read_only=True, data_only=True)
    series = {}
    for name, sheet, yr_row, m_start, yr_col in [
        ('PPI_Industry',     '2.1', 3, 5, 2),
        ('PPI_Mining',       '2.2', 3, 5, 2),
        ('PPI_Manufacturing','2.3', 3, 5, 2),
        ('PPI_Energy',       '2.4', 3, 5, 2),
    ]:
        series[name] = build_cumulative(parse_ppi_sheet(wb[sheet], yr_row, m_start, yr_col))
    wb.close()

    path_ag = os.path.join(DATA_DIR, 'ind_sx.xlsx')
    wb2 = openpyxl.load_workbook(path_ag, read_only=True, data_only=True)
    series['PPI_Agriculture'] = build_cumulative(parse_ppi_sheet(wb2['1.1'], 4, 6, 2))
    wb2.close()

    mapping = {
        0: ['PPI_Agriculture', 'PPI_Mining'],
        1: ['PPI_Manufacturing', 'PPI_Energy'],
        2: ['PPI_Industry'],
    }

    all_years = set()
    for d in series.values():
        all_years.update(d.keys())
    all_years = sorted(all_years)

    result = {}
    for yr in all_years:
        p_yr = np.ones((4, 3))
        for s in range(3):
            vals = [series[nm][yr] for nm in mapping[s] if yr in series.get(nm, {})]
            if vals:
                p_yr[:, s] = np.mean(vals, axis=0)
        result[yr] = p_yr

    print("\n" + "═" * 70)
    print(" СЕКТОРАЛЬНЫЕ ЦЕНЫ (ИЦП), база Q1 2019 = 1.0")
    print("═" * 70)
    for yr in [2019, 2020, 2021, 2022, 2023, 2024]:
        if yr in result:
            print(f"  {yr}:")
            for q in range(4):
                print(f"    Q{q+1}: {' '.join(f'{LABELS[s]}={result[yr][q,s]:.4f}' for s in range(3))}")
    return result


# ═══════════════════════════════════════════════════════
#  3. Индексы физического объёма ВДС → квартальные Q_t
# ═══════════════════════════════════════════════════════

def get_sector_volumes():
    """
    Лист '6' в VDS_kvartal_OKVED2.xlsx: ИФО ВДС, % к соотв. кварталу предыдущего года.
    Секторы: A (row 9), B (10), C (11), D (12), E (13), F (14), G (15), ...T (28).
    Агрегация: Primary=A+B, Secondary=C+D+E+F, Tertiary=G+...+T.
    """
    path = os.path.join(DATA_DIR, 'VDS_kvartal_OKVED2.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['6']

    rows_data = {}
    for r in range(3, 29):
        rows_data[r] = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    wb.close()

    yr_row = rows_data[3]
    years_cols = {}
    current_year = None
    for c in range(3, len(yr_row)):
        val = yr_row[c]
        if val is not None:
            m = re.match(r'(\d{4})', str(val))
            if m:
                current_year = int(m.group(1))
        if current_year:
            q_offset = (c - 3) % 4
            years_cols.setdefault(current_year, [None]*4)
            years_cols[current_year][q_offset] = c

    sector_rows = {
        'A': 9, 'B': 10, 'C': 11, 'D': 12, 'E': 13, 'F': 14,
        'G': 15, 'H': 16, 'I': 17, 'J': 18, 'K': 19, 'L': 20,
        'M': 21, 'N': 22, 'O': 23, 'P': 24, 'Q': 25, 'R': 26,
        'S': 27, 'T': 28,
    }

    primary_letters = ['A', 'B']
    secondary_letters = ['C', 'D', 'E', 'F']
    tertiary_letters = ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']

    raw = {}
    for letter, r in sector_rows.items():
        raw[letter] = {}
        for yr, cols in years_cols.items():
            vals = []
            for q_idx, c in enumerate(cols):
                if c is not None and c < len(rows_data.get(r, [])):
                    vals.append(safe_float(rows_data[r][c]))
                else:
                    vals.append(0.0)
            raw[letter][yr] = vals

    wb2 = openpyxl.load_workbook(path, data_only=True)
    ws1 = wb2['1']
    gva_rows = {}
    for r in range(9, 29):
        gva_rows[r] = [ws1.cell(r, c).value for c in range(1, ws1.max_column + 1)]

    r3 = [ws1.cell(3, c).value for c in range(1, ws1.max_column + 1)]
    wb2.close()

    yr_2021_cols = []
    cur_yr = None
    for c in range(3, len(r3)):
        val = r3[c]
        if val is not None:
            m = re.match(r'(\d{4})', str(val))
            if m:
                cur_yr = int(m.group(1))
        if cur_yr == 2021:
            yr_2021_cols.append(c)
            if len(yr_2021_cols) == 4:
                break

    gva_2021 = {}
    for letter, r in sector_rows.items():
        if r in gva_rows:
            total = 0
            for c in yr_2021_cols:
                if c < len(gva_rows[r]):
                    total += safe_float(gva_rows[r][c])
            gva_2021[letter] = total

    w_primary = {l: gva_2021.get(l, 1) for l in primary_letters}
    w_secondary = {l: gva_2021.get(l, 1) for l in secondary_letters}
    w_tertiary = {l: gva_2021.get(l, 1) for l in tertiary_letters}

    wp_sum = sum(w_primary.values())
    ws_sum = sum(w_secondary.values())
    wt_sum = sum(w_tertiary.values())

    print("\n" + "═" * 70)
    print(" ИНДЕКСЫ ФИЗИЧЕСКОГО ОБЪЁМА ВДС (год-к-году)")
    print("═" * 70)
    print("  Веса агрегации (ВДС 2021, млрд руб.):")
    for l in primary_letters:
        print(f"    {l}: {gva_2021.get(l,0):>10,.1f}")
    for l in secondary_letters:
        print(f"    {l}: {gva_2021.get(l,0):>10,.1f}")
    for l in tertiary_letters[:5]:
        print(f"    {l}: {gva_2021.get(l,0):>10,.1f}")
    print(f"    ...")

    def weighted_avg(letters, weights_dict, wsum, year, q):
        s = 0.0
        for l in letters:
            v = raw[l].get(year, [0]*4)[q]
            if v > 0:
                s += (weights_dict[l] / wsum) * v
        return s

    # Построение индексов уровня (цепная линковка год-к-году)
    yrs = sorted(years_cols.keys())
    yr_min = min(y for y in yrs if y >= 2013)
    yr_max = max(y for y in yrs if all(raw['A'].get(y, [0]*4)[q] > 0 for q in range(4)))

    level = {}
    base_yr = 2019
    level[base_yr] = np.ones((4, 3))

    for yr in range(base_yr + 1, yr_max + 1):
        lev = np.ones((4, 3))
        for q in range(4):
            lev[q, 0] = level[yr-1][q, 0] * weighted_avg(primary_letters, w_primary, wp_sum, yr, q) / 100
            lev[q, 1] = level[yr-1][q, 1] * weighted_avg(secondary_letters, w_secondary, ws_sum, yr, q) / 100
            lev[q, 2] = level[yr-1][q, 2] * weighted_avg(tertiary_letters, w_tertiary, wt_sum, yr, q) / 100
        level[yr] = lev

    for yr in range(base_yr - 1, yr_min - 1, -1):
        lev = np.ones((4, 3))
        for q in range(4):
            yoy_0 = weighted_avg(primary_letters, w_primary, wp_sum, yr+1, q) / 100
            yoy_1 = weighted_avg(secondary_letters, w_secondary, ws_sum, yr+1, q) / 100
            yoy_2 = weighted_avg(tertiary_letters, w_tertiary, wt_sum, yr+1, q) / 100
            lev[q, 0] = level[yr+1][q, 0] / yoy_0 if yoy_0 > 0 else 1.0
            lev[q, 1] = level[yr+1][q, 1] / yoy_1 if yoy_1 > 0 else 1.0
            lev[q, 2] = level[yr+1][q, 2] / yoy_2 if yoy_2 > 0 else 1.0
        level[yr] = lev

    print(f"\n  Индексы уровня (база: 2019 = 1.0), {yr_min}–{yr_max}:")
    for yr in [2019, 2020, 2021, 2022, 2023, 2024]:
        if yr in level:
            print(f"  {yr}:")
            for q in range(4):
                print(f"    Q{q+1}: {' '.join(f'{LABELS[s]}={level[yr][q,s]:.4f}' for s in range(3))}")

    return level


# ═══════════════════════════════════════════════════════
#  4. Формула (4.2) — приоритет для off-diagonal
# ═══════════════════════════════════════════════════════

def formula_42(A, X, n=3):
    """
    Формула (4.2) из статьи: b_ki = -m_i * φ_k / Σ_{j≠i} φ_j
    где m_i = diag(B)_i (из калибровки), φ_k = x_k = X_k / X.sum()
    Работает с нормированным выпуском x (не абсолютным X).
    """
    x = X / X.sum()
    B_offdiag = np.zeros((n, n))

    m_diag = np.zeros(n)
    for i in range(n):
        ai = A[:, i].sum()
        m_diag[i] = -x[i] / (2 * (1 - ai))

    for i in range(n):
        denom = sum(x[j] for j in range(n) if j != i)
        for k in range(n):
            if k != i:
                B_offdiag[k, i] = -m_diag[i] * x[k] / denom

    B_42 = np.diag(m_diag) + B_offdiag
    return B_42, m_diag, x


# ═══════════════════════════════════════════════════════
#  5. Калибровка: OLS и Ridge на первых разностях
# ═══════════════════════════════════════════════════════

def build_time_series(prices, volumes, yr_start=2019, yr_end=2024):
    """
    Строит непрерывные квартальные ряды P_t (n_obs × 3) и Q_t (n_obs × 3).
    Индексирует: t=0 → yr_start Q1, t=1 → yr_start Q2, ...
    """
    P_list, Q_list = [], []
    for yr in range(yr_start, yr_end + 1):
        if yr in prices and yr in volumes:
            for q in range(4):
                P_list.append(prices[yr][q, :])
                Q_list.append(volumes[yr][q, :])
    return np.array(P_list), np.array(Q_list)


def estimate_B_first_diff(P, Q, method='ols', B_prior=None, alpha=1.0):
    """
    ΔQ_t = B · ΔP_t + ε_t
    Для каждой строки i: ΔQ_{i,t} = B[i,:] · ΔP_t + ε
    """
    n = P.shape[1]
    T = P.shape[0]
    dP = np.diff(P, axis=0)  # (T-1) × n
    dQ = np.diff(Q, axis=0)

    B_est = np.zeros((n, n))
    residuals = np.zeros((T - 1, n))

    for i in range(n):
        y = dQ[:, i]
        X_reg = dP

        if method == 'ols':
            beta, res, _, _ = np.linalg.lstsq(X_reg, y, rcond=None)
        elif method == 'ridge':
            prior = B_prior[i, :] if B_prior is not None else np.zeros(n)
            XtX = X_reg.T @ X_reg
            Xty = X_reg.T @ y
            beta = np.linalg.solve(XtX + alpha * np.eye(n), Xty + alpha * prior)
        else:
            raise ValueError(f"Unknown method: {method}")

        B_est[i, :] = beta
        residuals[:, i] = y - X_reg @ beta

    return B_est, residuals


def estimate_B_yoy_diff(P, Q, method='ols', B_prior=None, alpha=1.0):
    """
    Год-к-году разности: ΔQ_{y,q} = Q_{y,q} - Q_{y-1,q}, аналогично для P.
    Снимает сезонность, что важно для квартальных данных.
    """
    n = P.shape[1]
    T = P.shape[0]
    if T < 8:
        return estimate_B_first_diff(P, Q, method, B_prior, alpha)

    dP_list, dQ_list = [], []
    for t in range(4, T):
        dP_list.append(P[t] - P[t-4])
        dQ_list.append(Q[t] - Q[t-4])

    dP = np.array(dP_list)
    dQ = np.array(dQ_list)

    B_est = np.zeros((n, n))
    for i in range(n):
        y = dQ[:, i]
        X_reg = dP
        if method == 'ols':
            beta, _, _, _ = np.linalg.lstsq(X_reg, y, rcond=None)
        elif method == 'ridge':
            prior = B_prior[i, :] if B_prior is not None else np.zeros(n)
            XtX = X_reg.T @ X_reg
            Xty = X_reg.T @ y
            beta = np.linalg.solve(XtX + alpha * np.eye(n), Xty + alpha * prior)
        else:
            raise ValueError(method)
        B_est[i, :] = beta

    return B_est, dQ - dP @ B_est.T


# ═══════════════════════════════════════════════════════
#  6. Аналитическая рекалибровка
# ═══════════════════════════════════════════════════════

def recalibrate(B_est, A, X):
    """
    Аналитическая подстройка diag(B) и θ так, чтобы при P*=1 модель
    воспроизводила базовые X (P* = 1 → Q(1) = X_base_normalized).
    """
    n = A.shape[0]
    x = X / X.sum()
    ones = np.ones(n)

    B_new = B_est.copy()
    for i in range(n):
        offdiag_sum = sum(B_new[j, i] for j in range(n) if j != i)
        ai = A[:, i].sum()
        B_new[i, i] = -x[i] / (2 * (1 - ai)) - offdiag_sum / 2

    theta = x - B_new @ ones
    return B_new, theta


# ═══════════════════════════════════════════════════════
#  7. Проверка отрицательной определённости
# ═══════════════════════════════════════════════════════

def check_neg_def(B, label="B"):
    sym = (B + B.T) / 2
    eigs = eigvalsh(sym)
    is_nd = all(e < 0 for e in eigs)
    print(f"\n  {label}:")
    print(f"    Собственные значения (B+Bᵀ)/2: {eigs}")
    print(f"    Отрицательно определена: {'ДА ✓' if is_nd else 'НЕТ ✗'}")
    return is_nd


# ═══════════════════════════════════════════════════════
#  8. Проверка равновесия
# ═══════════════════════════════════════════════════════

def solve_equilibrium(A, B, theta, n=3):
    """
    Находит P* из условия равновесия (Нэш):
      P_i = (A[:,i]ᵀ P) + VA_i(P) / Q_i(P)
    Итеративно.
    """
    P = np.ones(n)
    for _ in range(5000):
        Q = B @ P + theta
        Q = np.maximum(Q, 1e-10)
        VA = Q * P - A.T @ (Q * np.ones((n, 1)).flatten()[:n] * P)
        P_new = np.zeros(n)
        for i in range(n):
            cost_i = A[:, i] @ P
            qi = Q[i]
            va_i = qi * P[i] - cost_i * qi
            P_new[i] = cost_i + va_i / qi if qi > 0 else 1.0
        if np.max(np.abs(P_new - P)) < 1e-12:
            return P_new, Q
        P = P_new
    return P, B @ P + theta


# ═══════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════

def main():
    io = parse_and_aggregate_io()
    A = io['A']
    X = io['X']
    VA = io['VA']
    x = X / X.sum()

    prices = get_sector_prices()
    volumes = get_sector_volumes()

    P_ts, Q_ts = build_time_series(prices, volumes, yr_start=2019, yr_end=2024)
    print(f"\n  Временные ряды: P shape={P_ts.shape}, Q shape={Q_ts.shape}")
    print(f"  P[0] (2019 Q1): {P_ts[0]}")
    print(f"  Q[0] (2019 Q1): {Q_ts[0]}")

    # Масштабирование Q на нормированный выпуск
    Q_scaled = Q_ts * x[np.newaxis, :]

    # ─── Формула (4.2) ───
    B_42, m_diag, phi = formula_42(A, X)
    B_42r, theta_42 = recalibrate(B_42, A, X)

    print("\n" + "▓" * 70)
    print(" РЕЗУЛЬТАТЫ КАЛИБРОВКИ")
    print("▓" * 70)

    print("\n  ═══ Формула (4.2) + рекалибровка ═══")
    print(f"  B_42:\n{B_42r}")
    print(f"  θ_42: {theta_42}")
    check_neg_def(B_42r, "B(4.2)")

    # ─── OLS на первых разностях ───
    B_ols, res_ols = estimate_B_first_diff(P_ts, Q_scaled, method='ols')
    B_olsr, theta_ols = recalibrate(B_ols, A, X)

    print("\n  ═══ OLS (первые разности) + рекалибровка ═══")
    print(f"  B_ols:\n{B_olsr}")
    print(f"  θ_ols: {theta_ols}")
    check_neg_def(B_olsr, "B(OLS)")

    # ─── Ridge с приоритетом из (4.2) ───
    for alpha in [0.01, 0.1, 1.0, 5.0]:
        B_ridge, _ = estimate_B_first_diff(P_ts, Q_scaled, method='ridge',
                                            B_prior=B_42r, alpha=alpha)
        B_rr, theta_rr = recalibrate(B_ridge, A, X)
        nd = check_neg_def(B_rr, f"B(Ridge α={alpha})")
        if nd:
            print(f"    ← найден отрицательно определённый вариант!")

    # ─── OLS на год-к-году разностях (снимает сезонность) ───
    B_yoy, res_yoy = estimate_B_yoy_diff(P_ts, Q_scaled, method='ols')
    B_yoyr, theta_yoy = recalibrate(B_yoy, A, X)

    print("\n  ═══ OLS (год-к-году разности) + рекалибровка ═══")
    print(f"  B_yoy:\n{B_yoyr}")
    print(f"  θ_yoy: {theta_yoy}")
    check_neg_def(B_yoyr, "B(YoY-OLS)")

    # ─── Ridge на год-к-году разностях ───
    best_B = None
    best_alpha = None
    for alpha in [0.01, 0.1, 0.5, 1.0, 3.0, 5.0, 10.0]:
        B_yr, _ = estimate_B_yoy_diff(P_ts, Q_scaled, method='ridge',
                                       B_prior=B_42r, alpha=alpha)
        B_yrr, theta_yr = recalibrate(B_yr, A, X)
        nd = check_neg_def(B_yrr, f"B(YoY-Ridge α={alpha})")
        if nd and best_B is None:
            best_B = B_yrr
            best_alpha = alpha

    # ─── Итоговый лучший вариант ───
    print("\n" + "▓" * 70)
    print(" ИТОГОВЫЙ РЕЗУЛЬТАТ")
    print("▓" * 70)

    if best_B is not None:
        B_final = best_B
        theta_final = x - B_final @ np.ones(3)
        print(f"\n  Лучший метод: YoY-Ridge (α={best_alpha})")
    else:
        B_final = B_42r
        theta_final = theta_42
        print(f"\n  Лучший метод: Формула (4.2) (ни один Ridge не дал neg-def)")

    print(f"\n  B_final:\n{B_final}")
    print(f"  θ_final: {theta_final}")

    P_eq, Q_eq = solve_equilibrium(A, B_final, theta_final)
    print(f"\n  Равновесие: P* = {P_eq}")
    print(f"              Q* = {Q_eq}")
    print(f"              x  = {x}")
    print(f"              Q*/sum = {Q_eq / Q_eq.sum()}")

    # ─── Вычисление Фробениус норм ───
    print(f"\n  ||B_OLS - B_42||_F    = {np.linalg.norm(B_olsr - B_42r):.4f}")
    print(f"  ||B_YoY - B_42||_F    = {np.linalg.norm(B_yoyr - B_42r):.4f}")
    if best_B is not None:
        print(f"  ||B_Ridge - B_42||_F  = {np.linalg.norm(best_B - B_42r):.4f}")

    # ─── Сохранение ───
    np.savez(
        os.path.join(os.path.dirname(__file__), 'calibration_results.npz'),
        A=A, X=X, VA=VA, x=x,
        B_42=B_42r, theta_42=theta_42,
        B_ols=B_olsr, theta_ols=theta_ols,
        B_final=B_final, theta_final=theta_final,
    )
    print("\n  → calibration_results.npz сохранён")


if __name__ == '__main__':
    main()
