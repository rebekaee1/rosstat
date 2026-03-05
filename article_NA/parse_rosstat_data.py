"""
Парсинг данных Росстата:
  1. Симметричная таблица «затраты-выпуск» 2021 (baz_tzv-2021.xlsx)
  2. ИЦП по ВЭД (Proizvoditeli_Ind_VED.xlsx, листы 2.1–2.4)
  3. ИЦП с/х продукции (ind_sx.xlsx, лист 1.1)

Агрегация 113 отраслей → 3 сектора:
  Primary   : Сельское хозяйство (01–03) + Добыча (05–09)
  Secondary : Обработка (10–33) + Энергетика (35) + Водоснабжение (36–39) + Строительство (41–43)
  Tertiary  : Услуги (45+)
"""

import sys
import os
import re
import numpy as np
import openpyxl

np.set_printoptions(precision=2, suppress=True, linewidth=140)

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')


# ══════════════════════════════════════════════════════
#  Вспомогательные функции
# ══════════════════════════════════════════════════════

def _leading_code(raw_code: str) -> int | None:
    """Извлекает ведущее целое число из кода ОКПД2 (напр. '10.1' → 10, '01.1+01.2+...' → 1)."""
    if not raw_code:
        return None
    m = re.match(r'(\d+)', raw_code.strip())
    return int(m.group(1)) if m else None


def classify_sector(code_str: str) -> int:
    """0=Primary, 1=Secondary, 2=Tertiary.  Определяется по ведущему коду ОКПД2."""
    lc = _leading_code(code_str)
    if lc is None:
        return -1
    if lc <= 9:
        return 0
    if lc <= 43:
        return 1
    return 2


def safe_float(v) -> float:
    """Превращает строковое/числовое значение в float; None → 0."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(',', '.').replace('\xa0', '').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


# ══════════════════════════════════════════════════════
#  1. Таблица «затраты-выпуск»
# ══════════════════════════════════════════════════════

def parse_io_table():
    """
    Возвращает dict с ключами:
      codes       – список 113 кодов ОКПД2
      names       – список 113 названий отраслей
      sectors     – np.array (113,) с метками 0/1/2
      Z           – матрица промежуточного потребления 113×113 (млн руб.)
      Y_final     – вектор конечного использования 113 (сумма: потребление + накопление + экспорт)
      X_total     – вектор валового выпуска 113
      VA          – вектор ВДС 113
      imp         – вектор импорта 113
    """
    path = os.path.join(DATA_DIR, 'baz_tzv-2021.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['Симм ТЗВ']

    N = 113
    COL0 = 4  # первый столбец данных (D), 1-indexed

    # row indices (1-indexed)
    COL_FINAL_USE = 128
    ROW_TOTAL_IC = 122
    ROW_GVA = 128
    ROW_OUTPUT = 129
    ROW_IMPORT = 130

    all_rows = {}
    for row in ws.iter_rows(min_row=1, max_row=ROW_IMPORT, max_col=COL_FINAL_USE, values_only=False):
        r = row[0].row
        all_rows[r] = [cell.value for cell in row]

    wb.close()

    codes = [str(all_rows[5 + i][1] or '') for i in range(N)]
    names = [str(all_rows[5 + i][2] or '') for i in range(N)]
    sectors = np.array([classify_sector(c) for c in codes])

    Z = np.zeros((N, N))
    for i in range(N):
        row_data = all_rows[5 + i]
        for j in range(N):
            Z[i, j] = safe_float(row_data[COL0 - 1 + j])

    Y_final = np.array([safe_float(all_rows[5 + j][COL_FINAL_USE - 1]) for j in range(N)])

    row_gva = all_rows[ROW_GVA]
    row_out = all_rows[ROW_OUTPUT]
    row_imp = all_rows[ROW_IMPORT]
    row_tic = all_rows[ROW_TOTAL_IC]

    X_total = np.array([safe_float(row_out[COL0 - 1 + j]) for j in range(N)])
    VA = np.array([safe_float(row_gva[COL0 - 1 + j]) for j in range(N)])
    imp = np.array([safe_float(row_imp[COL0 - 1 + j]) for j in range(N)])
    total_ic_row = np.array([safe_float(row_tic[COL0 - 1 + j]) for j in range(N)])

    print("═" * 70)
    print(" ТАБЛИЦА «ЗАТРАТЫ-ВЫПУСК» 2021 (113 отраслей)")
    print("═" * 70)
    print(f"  Размер Z: {Z.shape}")
    print(f"  Секторы: Primary={np.sum(sectors==0)}, Secondary={np.sum(sectors==1)}, Tertiary={np.sum(sectors==2)}")
    print(f"\n  Промежуточное потребление Z.sum() = {Z.sum():,.0f} млн руб.")
    print(f"  Конечное использование Y.sum()     = {Y_final.sum():,.0f}")
    print(f"  Выпуск X.sum()                     = {X_total.sum():,.0f}")
    print(f"  ВДС VA.sum()                        = {VA.sum():,.0f}")
    print(f"  Импорт imp.sum()                    = {imp.sum():,.0f}")

    print(f"\n  Проверка: Z_col_sums + VA ≈ X_total (для каждой отрасли)")
    z_col = Z.sum(axis=0)
    discrepancy = X_total - (total_ic_row + VA)
    print(f"  Макс. расхождение (X - IC_row - VA): {np.max(np.abs(discrepancy)):,.0f}")

    print(f"\n  Первые 5 кодов: {codes[:5]}")
    print(f"  Первые 5 названий: {[n[:40] for n in names[:5]]}")

    return dict(codes=codes, names=names, sectors=sectors, Z=Z,
                Y_final=Y_final, X_total=X_total, VA=VA, imp=imp)


def aggregate_io(io_data, n_sectors=3):
    """
    Агрегирует 113-отраслевую таблицу до n_sectors=3.
    Возвращает: Z_agg, A_agg, X_agg, VA_agg, Y_agg, imp_agg
    """
    sectors = io_data['sectors']
    Z = io_data['Z']
    X_total = io_data['X_total']
    VA = io_data['VA']
    Y_final = io_data['Y_final']
    imp = io_data['imp']

    Z_agg = np.zeros((n_sectors, n_sectors))
    X_agg = np.zeros(n_sectors)
    VA_agg = np.zeros(n_sectors)
    Y_agg = np.zeros(n_sectors)
    imp_agg = np.zeros(n_sectors)

    for s in range(n_sectors):
        mask = sectors == s
        X_agg[s] = X_total[mask].sum()
        VA_agg[s] = VA[mask].sum()
        Y_agg[s] = Y_final[mask].sum()
        imp_agg[s] = imp[mask].sum()
        for t in range(n_sectors):
            mask_t = sectors == t
            Z_agg[s, t] = Z[np.ix_(mask, mask_t)].sum()

    A_agg = Z_agg / X_agg[np.newaxis, :]

    print("\n" + "═" * 70)
    print(" АГРЕГИРОВАННАЯ ТАБЛИЦА (3 сектора)")
    print("═" * 70)
    labels = ['Primary', 'Secondary', 'Tertiary']
    for s in range(n_sectors):
        print(f"  {labels[s]:12s}: X={X_agg[s]:>12,.0f}  VA={VA_agg[s]:>12,.0f}  Y={Y_agg[s]:>12,.0f}  Imp={imp_agg[s]:>10,.0f}")

    print(f"\n  Матрица Z_agg (промежуточное потребление, млн руб.):")
    for i in range(n_sectors):
        row_str = '  '.join(f'{Z_agg[i,j]:>12,.0f}' for j in range(n_sectors))
        print(f"    {labels[i]:12s}: {row_str}")

    print(f"\n  Матрица A (коэффициенты прямых затрат):")
    for i in range(n_sectors):
        row_str = '  '.join(f'{A_agg[i,j]:>8.4f}' for j in range(n_sectors))
        print(f"    {labels[i]:12s}: {row_str}")

    print(f"\n  Доли ВДС: {VA_agg / VA_agg.sum()}")
    print(f"  Доли выпуска: {X_agg / X_agg.sum()}")

    return dict(Z=Z_agg, A=A_agg, X=X_agg, VA=VA_agg, Y=Y_agg, imp=imp_agg)


# ══════════════════════════════════════════════════════
#  2. Индексы цен производителей (ИЦП)
# ══════════════════════════════════════════════════════

def _parse_ppi_sheet(ws, year_row: int, month_row_start: int, year_col_start: int):
    """
    Универсальный парсер листа PPI.
    Возвращает dict {year: [jan, feb, ..., dec]} где значения — проценты к предыдущему месяцу.
    """
    rows_needed = list(range(year_row, year_row + 1)) + list(range(month_row_start, month_row_start + 12))
    max_row_needed = month_row_start + 11
    data = {}
    for row in ws.iter_rows(min_row=year_row, max_row=max_row_needed, values_only=False):
        r = row[0].row
        data[r] = [cell.value for cell in row]

    yr_cells = data[year_row]
    years = {}
    col_idx = year_col_start - 1  # 0-indexed
    while col_idx < len(yr_cells):
        raw = yr_cells[col_idx]
        if raw is None:
            break
        m_yr = re.match(r'(\d{4})', str(raw).strip())
        if not m_yr:
            break
        yr = int(m_yr.group(1))
        months = []
        for m in range(12):
            row_data = data.get(month_row_start + m, [])
            v = safe_float(row_data[col_idx] if col_idx < len(row_data) else None)
            months.append(v)
        years[yr] = months
        col_idx += 1
    return years


def parse_ppi():
    """
    Читает ИЦП из Proizvoditeli_Ind_VED.xlsx (листы 2.1–2.4) и ind_sx.xlsx (лист 1.1).
    Возвращает dict: {name: {year: [12 monthly values]}}
    """
    result = {}

    path_ppi = os.path.join(DATA_DIR, 'Proizvoditeli_Ind_VED.xlsx')
    wb = openpyxl.load_workbook(path_ppi, read_only=True, data_only=True)

    sheets_info = {
        'PPI_Industry':      ('2.1', 3, 5, 2),
        'PPI_Mining':         ('2.2', 3, 5, 2),
        'PPI_Manufacturing':  ('2.3', 3, 5, 2),
        'PPI_Energy':         ('2.4', 3, 5, 2),
    }

    for name, (sheet, yr_row, m_start, yr_col) in sheets_info.items():
        ws = wb[sheet]
        result[name] = _parse_ppi_sheet(ws, yr_row, m_start, yr_col)

    wb.close()

    path_ag = os.path.join(DATA_DIR, 'ind_sx.xlsx')
    wb2 = openpyxl.load_workbook(path_ag, read_only=True, data_only=True)
    ws_ag = wb2['1.1']
    result['PPI_Agriculture'] = _parse_ppi_sheet(ws_ag, year_row=4, month_row_start=6, year_col_start=2)
    wb2.close()

    print("\n" + "═" * 70)
    print(" ИНДЕКСЫ ЦЕН ПРОИЗВОДИТЕЛЕЙ (% к предыдущему месяцу)")
    print("═" * 70)
    for name, ydata in result.items():
        yr_range = sorted(ydata.keys())
        print(f"\n  {name}: {yr_range[0]}–{yr_range[-1]} ({len(yr_range)} лет)")
        for yr in [2019, 2020, 2021]:
            if yr in ydata:
                vals = ydata[yr]
                print(f"    {yr}: {['%.1f' % v for v in vals]}")

    return result


def build_cumulative_index(monthly_pct: dict, base_year=2019, base_quarter=1):
    """
    Из dict {year: [12 monthly %_to_prev]} строит кумулятивный индекс.
    Базирует на среднее значение указанного квартала = 1.0.

    Возвращает dict {year: np.array(12)}, а также quarterly dict {year: np.array(4)}.
    """
    all_years = sorted(monthly_pct.keys())
    if not all_years:
        return {}, {}

    cum = {}
    prev = 1.0
    for yr in all_years:
        vals = monthly_pct[yr]
        yr_idx = np.zeros(12)
        for m in range(12):
            pct = vals[m]
            if pct == 0:
                yr_idx[m] = prev
            else:
                prev = prev * (pct / 100.0)
                yr_idx[m] = prev
        cum[yr] = yr_idx

    q_start = (base_quarter - 1) * 3
    q_end = q_start + 3
    if base_year in cum:
        base_val = cum[base_year][q_start:q_end].mean()
    else:
        base_val = 1.0

    for yr in cum:
        cum[yr] = cum[yr] / base_val

    quarterly = {}
    for yr in cum:
        q = np.array([cum[yr][i*3:(i+1)*3].mean() for i in range(4)])
        quarterly[yr] = q

    return cum, quarterly


def process_price_indices(ppi_raw):
    """
    Обрабатывает все серии ИЦП: строит кумулятивные индексы с базой Q1 2019 = 1.0.
    Возвращает dict серий, каждая содержит monthly и quarterly.
    """
    result = {}

    print("\n" + "═" * 70)
    print(" КУМУЛЯТИВНЫЕ ИНДЕКСЫ ЦЕН (база: Q1 2019 = 1.0)")
    print("═" * 70)

    for name, ydata in ppi_raw.items():
        monthly, quarterly = build_cumulative_index(ydata, base_year=2019, base_quarter=1)
        result[name] = dict(monthly=monthly, quarterly=quarterly)

        print(f"\n  {name}:")
        for yr in [2019, 2020, 2021, 2022, 2023, 2024]:
            if yr in quarterly:
                q = quarterly[yr]
                print(f"    {yr} Q: {q}")

    return result


# ══════════════════════════════════════════════════════
#  3. Построение секторальных ценовых индексов
# ══════════════════════════════════════════════════════

def build_sector_prices(price_data):
    """
    Сопоставляет секторальные цены:
      Primary   = среднее(PPI_Agriculture, PPI_Mining)  — взвешенное можно потом
      Secondary = среднее(PPI_Manufacturing, PPI_Energy)
      Tertiary  = PPI_Industry (как прокси)

    Возвращает quarterly dict: {year: np.array(3)} — цены трёх секторов.
    """
    mapping = {
        0: ['PPI_Agriculture', 'PPI_Mining'],
        1: ['PPI_Manufacturing', 'PPI_Energy'],
        2: ['PPI_Industry'],
    }
    labels = ['Primary', 'Secondary', 'Tertiary']

    all_years = set()
    for name, data in price_data.items():
        all_years.update(data['quarterly'].keys())
    all_years = sorted(all_years)

    sector_quarterly = {}
    for yr in all_years:
        p = np.ones(3)
        for s in range(3):
            vals = []
            for series_name in mapping[s]:
                if series_name in price_data and yr in price_data[series_name]['quarterly']:
                    vals.append(price_data[series_name]['quarterly'][yr])
            if vals:
                p_q = np.mean(vals, axis=0)
                sector_quarterly.setdefault(yr, []).append(p_q)
            else:
                sector_quarterly.setdefault(yr, []).append(np.ones(4))

    result = {}
    for yr in all_years:
        result[yr] = np.column_stack(sector_quarterly[yr])  # shape (4, 3)

    print("\n" + "═" * 70)
    print(" СЕКТОРАЛЬНЫЕ ЦЕНОВЫЕ ИНДЕКСЫ (квартальные)")
    print("═" * 70)
    for yr in [2019, 2020, 2021, 2022, 2023, 2024]:
        if yr in result:
            arr = result[yr]
            print(f"\n  {yr}:")
            for q in range(4):
                print(f"    Q{q+1}: {' '.join(f'{labels[s]}={arr[q,s]:.4f}' for s in range(3))}")

    return result


# ══════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════

def main():
    io_data = parse_io_table()
    agg = aggregate_io(io_data)

    ppi_raw = parse_ppi()
    price_data = process_price_indices(ppi_raw)
    sector_prices = build_sector_prices(price_data)

    print("\n" + "═" * 70)
    print(" ИТОГОВЫЕ ДАННЫЕ ДЛЯ МОДЕЛИ")
    print("═" * 70)

    A = agg['A']
    X = agg['X']
    VA = agg['VA']
    labels = ['Primary', 'Secondary', 'Tertiary']

    print(f"\n  Матрица A (коэф. прямых затрат):\n{A}")
    print(f"\n  Выпуск X (млн руб.): {X}")
    print(f"  Нормированный выпуск x = X/X.sum(): {X / X.sum()}")
    print(f"  ВДС VA (млн руб.): {VA}")
    print(f"  Нормированная ВДС: {VA / VA.sum()}")

    if 2021 in sector_prices:
        P_2021 = sector_prices[2021]
        print(f"\n  Средние цены 2021 (Q1–Q4):\n{P_2021}")

    print("\n  === Данные сохранены как numpy-массивы ===")
    np.savez(
        os.path.join(os.path.dirname(__file__), 'parsed_data.npz'),
        A=A,
        X=X,
        VA=VA,
        Y=agg['Y'],
        imp=agg['imp'],
        Z=agg['Z'],
    )
    print("  → parsed_data.npz")

    return agg, price_data, sector_prices


if __name__ == '__main__':
    main()
