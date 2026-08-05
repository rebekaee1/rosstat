"""Аудит экспорта CSV/Excel для world-рядов.

Путь: API data → те же билдеры, что /api/v1/export/table (_build_csv/_build_xlsx).
Гостевой лимит на HTTP часто исчерпан — проверяем и прямой вызов билдеров,
и HTTP при возможности.

Проверяет: совпадение чисел с API, заголовки, единицу в подписи, даты,
десятичный разделитель, BOM/UTF-8, наличие указания источника.

Запуск:
  python3 scripts/audit-world-export.py
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

import requests
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from audit_world_lib import (  # noqa: E402
    API_BASE,
    api_get,
    connect,
    run_async,
    values_close,
    write_json,
    eprint,
)

# Импорт боевых билдеров (не копия) — аудит ловит регресс provenance/запятой.
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "backend"))
from app.api.export import _build_csv, _build_xlsx, _resolve_meta  # noqa: E402


def _parse_ru_number(raw: str) -> float | None:
    if raw is None or raw == "":
        return None
    s = raw.replace("\u202f", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


SAMPLE_CODES_SQL = """
WITH ranked AS (
  SELECT i.code, c.slug, c.code AS geo, i.name_ru, i.unit_ru, i.frequency,
         i.dataset_id, i.source, i.source_url, c.name_ru AS country_name,
         row_number() OVER (
           PARTITION BY c.code
           ORDER BY i.points_count DESC
         ) AS rn
  FROM world_indicators i
  JOIN world_countries c ON c.id = i.country_id
  WHERE i.is_listed AND c.is_active
    AND c.code IN ('DE','FR','PL','IT','TR','MT')
)
SELECT * FROM ranked WHERE rn = 1
UNION ALL
SELECT i.code, c.slug, c.code AS geo, i.name_ru, i.unit_ru, i.frequency,
       i.dataset_id, i.source, i.source_url, c.name_ru AS country_name, 0
FROM world_indicators i
JOIN world_countries c ON c.id = i.country_id
WHERE i.is_listed AND c.is_active
  AND (
    i.name_ru ILIKE '%безработ%'
    OR i.name_ru ILIKE '%инфляц%'
    OR i.name_ru ILIKE '%ВВП%'
    OR i.name_ru ILIKE '%населен%'
  )
  AND c.code IN ('DE','FR','ES','PL')
LIMIT 12
"""


async def amain() -> int:
    conn = await connect()
    try:
        rows = await conn.fetch(SAMPLE_CODES_SQL)
        seen = set()
        sample = []
        for r in rows:
            if r["code"] in seen:
                continue
            seen.add(r["code"])
            sample.append(dict(r))
        sample = sample[:8]
        eprint(f"Export sample: {len(sample)}")

        results = []
        for ind in sample:
            slug, code = ind["slug"], ind["code"]
            eprint(f"  {slug}/{code}")
            api = api_get(f"/world/indicators/{slug}/{code}/data", params={"mode": "level"})
            points = api.get("points") or []
            unit = api.get("unit") or ind["unit_ru"] or ""
            value_label = f"{ind['name_ru']} ({unit})" if unit else ind["name_ru"]
            meta = _resolve_meta(
                None,
                indicator_name=ind["name_ru"],
                unit=unit,
                frequency=ind.get("frequency") or "",
                country=ind.get("country_name") or "",
                source=ind.get("source") or "Евростат",
                source_url=ind.get("source_url") or "",
            )

            facts = [
                (p["date"], None if p["value"] is None else round(float(p["value"]), 4))
                for p in points
            ]
            forecasts = []
            csv_bytes = _build_csv(facts, forecasts, value_label, meta)
            xlsx_bytes = _build_xlsx(facts, forecasts, value_label, meta)

            csv_text = csv_bytes.decode("utf-8-sig")
            csv_lines = csv_text.splitlines()
            meta_lines = [ln for ln in csv_lines if ln.startswith("# ")]
            data_lines = [ln for ln in csv_lines if not ln.startswith("#")]
            header = data_lines[0] if data_lines else ""
            body = data_lines[1:] if len(data_lines) > 1 else []

            has_bom = csv_bytes.startswith("\ufeff".encode("utf-8"))
            uses_semicolon = ";" in header
            decimal_dots = 0
            decimal_commas = 0
            value_mismatches = []
            for i, line in enumerate(body[: min(50, len(body))]):
                parts = line.split(";")
                if len(parts) < 3:
                    continue
                d, raw_v, typ = parts[0], parts[1], parts[2]
                if "," in raw_v:
                    decimal_commas += 1
                if "." in raw_v and "," not in raw_v:
                    decimal_dots += 1
                if typ != "факт":
                    continue
                api_v = facts[i][1] if i < len(facts) else None
                csv_v = _parse_ru_number(raw_v)
                if api_v is not None and csv_v is not None and not values_close(api_v, csv_v, abs_tol=1e-3):
                    value_mismatches.append({"date": d, "csv": csv_v, "api": api_v})

            wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
            assert "Описание" in wb.sheetnames
            desc = list(wb["Описание"].iter_rows(values_only=True))
            desc_blob = " ".join(f"{a} {b}" for a, b in desc if a).lower()
            ws = wb["Факт"]
            xrows = list(ws.iter_rows(values_only=True))
            x_header = xrows[0] if xrows else ()
            x_mism = []
            for i, row in enumerate(xrows[1: min(51, len(xrows))]):
                if not row or row[0] is None:
                    continue
                d, v = row[0], row[1]
                api_v = facts[i][1] if i < len(facts) else None
                if api_v is not None and v is not None and not values_close(float(v), api_v, abs_tol=1e-3):
                    x_mism.append({"date": str(d), "xlsx": v, "api": api_v})

            blob_lower = (csv_text + desc_blob).lower()
            has_source = any(
                s in blob_lower for s in ("eurostat", "евростат", "источник", "source")
            )
            has_required_meta = (
                "показатель" in desc_blob
                and "источник" in desc_blob
                and "дата выгрузки" in desc_blob
            )

            http_status = None
            http_note = None
            try:
                payload = {
                    "format": "csv",
                    "filename": f"{code}_level_all.csv",
                    "value_label": value_label,
                    "indicator_name": ind["name_ru"],
                    "unit": unit,
                    "frequency": ind.get("frequency") or "",
                    "country": ind.get("country_name") or "",
                    "source": ind.get("source") or "Евростат",
                    "source_url": ind.get("source_url") or "",
                    "points": [
                        {"date": p["date"], "actual": p["value"]}
                        for p in points[:20]
                    ],
                }
                resp = requests.post(
                    f"{API_BASE}/export/table",
                    json=payload,
                    timeout=30,
                )
                http_status = resp.status_code
                if resp.status_code == 200:
                    http_note = "ok"
                    http_csv = resp.content.decode("utf-8-sig")
                    if "евростат" in http_csv.lower() or "источник" in http_csv.lower():
                        http_note = "ok_with_source"
                elif resp.status_code == 403:
                    http_note = "guest_download_limit"
                else:
                    http_note = resp.text[:200]
            except Exception as exc:  # noqa: BLE001
                http_note = str(exc)

            results.append({
                "code": code,
                "slug": slug,
                "geo": ind["geo"],
                "name_ru": ind["name_ru"],
                "unit_ru": unit,
                "points_n": len(points),
                "csv": {
                    "has_utf8_bom": has_bom,
                    "semicolon_sep": uses_semicolon,
                    "header": header,
                    "meta_lines": meta_lines[:8],
                    "decimal_dot_count": decimal_dots,
                    "decimal_comma_count": decimal_commas,
                    "uses_russian_comma": decimal_commas > 0 and decimal_dots == 0,
                    "value_mismatches_n": len(value_mismatches),
                    "value_mismatches_sample": value_mismatches[:5],
                    "row_count": len(body),
                },
                "xlsx": {
                    "header": list(x_header),
                    "sheet": "Факт",
                    "description_rows": [list(r) for r in desc[:8]],
                    "value_mismatches_n": len(x_mism),
                    "value_mismatches_sample": x_mism[:5],
                    "row_count": max(0, len(xrows) - 1),
                },
                "source_in_file": has_source,
                "required_meta_in_file": has_required_meta,
                "indicator_source_field": ind.get("source"),
                "indicator_source_url": ind.get("source_url"),
                "http_export": {"status": http_status, "note": http_note},
            })

        all_csv_ok = all(r["csv"]["value_mismatches_n"] == 0 for r in results) if results else False
        all_xlsx_ok = all(r["xlsx"]["value_mismatches_n"] == 0 for r in results) if results else False
        any_ru_comma = any(r["csv"]["uses_russian_comma"] for r in results)
        any_source = any(r["source_in_file"] for r in results)
        any_dot = any(r["csv"]["decimal_dot_count"] > 0 for r in results)
        all_meta = all(r["required_meta_in_file"] for r in results) if results else False

        summary = {
            "series_checked": len(results),
            "csv_values_match_api": all_csv_ok,
            "xlsx_values_match_api": all_xlsx_ok,
            "csv_uses_russian_decimal_comma": any_ru_comma,
            "csv_uses_dot_decimal": any_dot,
            "source_attribution_in_file": any_source,
            "required_meta_in_file": all_meta,
            "notes": [
                "CSV через display.format_number_ru (русская запятая).",
                "Шапка: показатель, единица, частота, страна, источник, дата выгрузки.",
            ],
        }
        out = {"summary": summary, "results": results}
        path = write_json("export-audit.json", out)
        eprint(f"Wrote {path}")
        print(
            f"EXPORT: n={len(results)} csv_ok={all_csv_ok} xlsx_ok={all_xlsx_ok} "
            f"ru_comma={any_ru_comma} source_in_file={any_source} meta={all_meta}"
        )
        for r in results:
            print(
                f"  {r['geo']} {r['code']}: csv_mism={r['csv']['value_mismatches_n']} "
                f"xlsx_mism={r['xlsx']['value_mismatches_n']} http={r['http_export']}"
            )
        if results and (not all_csv_ok or not all_xlsx_ok or not any_source or not any_ru_comma):
            return 1
        return 0
    finally:
        await conn.close()


if __name__ == "__main__":
    raise SystemExit(run_async(amain()))
