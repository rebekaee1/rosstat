"""Дособор показателей, исключённых из Excel-приложения 2025, из редакций 2023/2022.

Росстат прекращает публиковать часть таблиц; последние доступные ряды лежат
в Excel-приложениях прошлых лет (Pril_2023, Pril_2022 — та же структура листов,
что и Pril_Region_Pokaz_2025). Здесь — только по-настоящему исчезнувшие таблицы
(переименованные в 2025 пропущены, у них ряды уже есть):

  Pril_2023: внешняя торговля 21.1.x/21.2.x (экспорт/импорт, дальнее зарубежье/СНГ,
  2000-2021, данные ФТС), госслужащие 2.15-2.18.x, зарплата муниципальных
  служащих 3.5.x, прожиточный минимум пенсионера 3.14, прямые иностранные
  инвестиции 10.4.x, лён-долгунец 13.8/13.17/13.18, ж/д грузы 16.1.
  Pril_2022: изменение численности населения 1.8, организации с интернетом 17.2,
  ПК на 100 работников 17.6.

Скрипт дописывает показатели и точки в артефакт backend/app/data/regional/
(запускать ПОСЛЕ parse_pril_2025.py). Идемпотентен: пере-
запуск перезаписывает свои же строки.

Запуск: python3 scripts/regional/backfill_pril_2022_2023.py /tmp/reg/2023/Pril_2023 /tmp/reg/2022/Pril_2022
"""

import csv
import glob
import gzip
import json
import os
import re
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from parse_pril_2025 import (  # noqa: E402
    OUT_DIR,
    clean_title,
    extract_unit,
    find_header,
    norm_text,
    parse_toc,
    parse_value,
    translit_slug,
)
from regions_registry import resolve_region  # noqa: E402

# (файл-паттерн, лист, section_num в нашей нумерации, section_name)
WHITELIST_2023 = [
    ("Труд", "2.15.", 2, "Труд"),
    ("Труд", "2.16.", 2, "Труд"),
    ("Труд", "2.17.", 2, "Труд"),
    ("Труд", "2.18.1", 2, "Труд"),
    ("Труд", "2.18.2", 2, "Труд"),
    ("Труд", "2.18.3", 2, "Труд"),
    ("Уровень жизни", "3.5.1", 3, "Уровень жизни населения"),
    ("Уровень жизни", "3.5.2", 3, "Уровень жизни населения"),
    ("Уровень жизни", "3.5.3", 3, "Уровень жизни населения"),
    ("Уровень жизни", "3.14.", 3, "Уровень жизни населения"),
    ("Инвестиции", "10.4.1", 10, "Инвестиции"),
    ("Инвестиции", "10.4.2", 10, "Инвестиции"),
    ("Инвестиции", "10.4.3", 10, "Инвестиции"),
    # glob-паттерн без «й»: имена файлов на macOS в NFD, «й» декомпозирована
    ("Сельское", "13.8.", 13, "Сельское хозяйство"),
    ("Сельское", "13.17.", 13, "Сельское хозяйство"),
    ("Сельское", "13.18.", 13, "Сельское хозяйство"),
    ("Транспорт", "16.1.", 16, "Транспорт"),
    ("Внешняя торговля", "21.1.1", 21, "Внешняя торговля"),
    ("Внешняя торговля", "21.1.2", 21, "Внешняя торговля"),
    ("Внешняя торговля", "21.2.1", 21, "Внешняя торговля"),
    ("Внешняя торговля", "21.2.2", 21, "Внешняя торговля"),
]
WHITELIST_2022 = [
    ("Население", "1.8.", 1, "Население"),
    ("Информационные", "17.2.", 17, "Информационные и коммуникационные технологии"),
    ("Информационные", "17.6.1", 17, "Информационные и коммуникационные технологии"),
    ("Информационные", "17.6.2", 17, "Информационные и коммуникационные технологии"),
]

# косметика для склеенных «родитель — потомок» имён
NAME_OVERRIDES = {
    "17.6.2": "Число персональных компьютеров на 100 работников — с доступом к сети Интернет",
}

DISCONTINUED_NOTE = (
    "Показатель исключён из актуальных выпусков сборника Росстата; "
    "приведены данные последней редакции, в которой он публиковался ({ed})."
)


def parse_sheet(path: str, sheet: str, sec_num: int, sec_name: str,
                edition: str, existing_slugs: set):
    wb = load_workbook(path, read_only=True, data_only=True)
    toc = parse_toc(wb)
    target = None
    for sh in wb.sheetnames:
        if sh.strip().rstrip(".") == sheet.strip().rstrip("."):
            target = sh
            break
    if target is None:
        wb.close()
        return None, [], f"нет листа {sheet}"

    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    title_idx, note, year_idx, years_by_col = find_header(rows)
    if year_idx is None:
        wb.close()
        return None, [], f"{sheet}: нет строки лет"

    table_code = re.sub(r"\.$", "", target.strip())
    title_sheet = clean_title(str(rows[title_idx][0])) if title_idx is not None else ""
    name = toc.get(table_code) or toc.get(table_code + ".") or title_sheet
    parent = ".".join(table_code.split(".")[:2])
    if table_code.count(".") == 2 and parent in toc:
        child = title_sheet or name
        if child and child.lower() != toc[parent].lower():
            if len(child) > 1 and child[1].islower():
                child = child[0].lower() + child[1:]
            name = f"{toc[parent]} — {child}"
        else:
            name = toc[parent]
    name = NAME_OVERRIDES.get(table_code) or norm_text(name)

    seen_years = {}
    for col, y in sorted(years_by_col.items()):
        seen_years.setdefault(y, col)
    unit = extract_unit(note, str(rows[title_idx][0]) if title_idx is not None else "")
    is_pct = bool(re.search(r"процент|%", (unit + " " + note + " " + name).lower()))

    base_slug = translit_slug(name) or f"tab-{table_code.replace('.', '-')}"
    slug = base_slug
    if slug in existing_slugs:
        slug = f"{base_slug}-{table_code.replace('.', '-')}"
    if slug in existing_slugs:
        slug = f"{slug}-{edition}"

    points, years_present = [], set()
    any_region = False
    for row in rows[year_idx + 1:]:
        c0 = row[0]
        if not isinstance(c0, str) or not c0.strip():
            continue
        t = c0.strip()
        if t.startswith(("_", "¾")):
            break
        if t.startswith("(") or re.match(r"^\d\)", t):
            continue
        rslug, scale = resolve_region(t)
        if rslug is None:
            n_vals = sum(1 for c in row[1:] if isinstance(c, (int, float)))
            if not any_region and n_vals >= 3:
                rslug, scale = "russia", 1.0
                if re.search(r",\s*(млрд|млн)\s", t):
                    scale = 1000.0
            else:
                continue
        any_region = True
        for y, col in seen_years.items():
            v = parse_value(row[col] if col < len(row) else None, is_pct)
            if v is None or v == "UNPARSED":
                continue
            points.append((slug, rslug, y, v * scale))
            years_present.add(y)
    wb.close()
    if not points:
        return None, [], f"{sheet}: 0 точек"

    ind = {
        "code": slug,
        "table_code": table_code,
        "section_num": sec_num,
        "section_name": sec_name,
        "name": name,
        "unit": unit,
        "note": DISCONTINUED_NOTE.format(ed=edition),
        "source_sheet": f"{os.path.basename(path)} ({edition}) / {target}",
        "year_min": min(years_present),
        "year_max": max(years_present),
        "n_points": len(points),
    }
    return ind, points, None


def main(root_2023: str, root_2022: str):
    indicators = json.load(open(os.path.join(OUT_DIR, "indicators.json")))
    existing_slugs = {i["code"] for i in indicators}
    # свои прежние строки выкидываем (идемпотентность по source_sheet-метке)
    my_mark = re.compile(r"\((2022|2023)\) /")
    keep = [i for i in indicators if not my_mark.search(i.get("source_sheet", ""))]
    dropped_codes = {i["code"] for i in indicators} - {i["code"] for i in keep}
    indicators = keep
    existing_slugs = {i["code"] for i in indicators}

    points = []
    with gzip.open(os.path.join(OUT_DIR, "data.csv.gz"), "rt", encoding="utf-8") as fh:
        r = csv.reader(fh, delimiter=";")
        next(r)
        for code, rslug, year, value in r:
            if code in dropped_codes:
                continue
            points.append((code, rslug, int(year), float(value)))

    new_inds, new_points, errors = [], [], []
    jobs = [(root_2023, "2023", WHITELIST_2023), (root_2022, "2022", WHITELIST_2022)]
    for root, edition, wl in jobs:
        for fpat, sheet, sec_num, sec_name in wl:
            matches = glob.glob(os.path.join(root, f"*{fpat}*.xlsx"))
            if not matches:
                errors.append(f"{edition}: нет файла {fpat}")
                continue
            ind, pts, err = parse_sheet(
                matches[0], sheet, sec_num, sec_name, edition, existing_slugs
            )
            if err:
                errors.append(f"{edition}/{fpat}: {err}")
                continue
            existing_slugs.add(ind["code"])
            new_inds.append(ind)
            new_points.extend(pts)
            print(f"  + [{edition}] {ind['table_code']:8s} {ind['name'][:70]} "
                  f"({ind['year_min']}-{ind['year_max']}, {ind['n_points']} тчк)")

    indicators.extend(new_inds)
    points.extend(new_points)

    with open(os.path.join(OUT_DIR, "indicators.json"), "w", encoding="utf-8") as fh:
        json.dump(indicators, fh, ensure_ascii=False, indent=1)
    with gzip.open(os.path.join(OUT_DIR, "data.csv.gz"), "wt", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(["indicator_code", "region_slug", "year", "value"])
        for p in points:
            w.writerow(p)

    print(f"\nдобавлено показателей: {len(new_inds)}, точек: {len(new_points)}")
    print(f"итог артефакта: {len(indicators)} показателей, {len(points)} точек")
    for e in errors:
        print("  ERR:", e)


if __name__ == "__main__":
    main(
        sys.argv[1] if len(sys.argv) > 1 else "/tmp/reg/2023/Pril_2023",
        sys.argv[2] if len(sys.argv) > 2 else "/tmp/reg/2022/Pril_2022",
    )
