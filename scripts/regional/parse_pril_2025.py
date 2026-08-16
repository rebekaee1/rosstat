"""Парсер Excel-приложения «Регионы России. Социально-экономические показатели. 2025».

Вход:  распакованный архив Pril_Region_Pokaz_2025 (20 файлов «Раздел N - ....xlsx»).
Выход: артефакт в backend/app/data/regional/:
  - indicators.json  — каталог показателей (code, table_code, section, name, unit, note, годы)
  - data.csv.gz      — точки: indicator_code;region_slug;year;value
  - regions.json     — реестр регионов (копия regions_registry)
  - parse_report.json — контроль качества (счётчики, аномалии, пропуски)

Структура листа (везде одинакова, проверено аудитом 450 листов):
  [пустые строки]* -> строка-колонтитул -> «N. РАЗДЕЛ» -> «N.M. ЗАГОЛОВОК»
  -> опционально «(примечание; единица)» -> строка лет -> строки регионов
  -> «______» -> сноски.
Годы: int/str, встречается формат учебного года «2000/2001» (берём первый год).
Значения: float/int; строки «-» (явление отсутствует) и «…» (нет данных) пропускаем;
«0,0»/«41.0» парсим; «в N,N р.» конвертируем в N*100 только для процентных таблиц.

Запуск: python3 scripts/regional/parse_pril_2025.py "<путь к распакованной папке>"
"""

import csv
import glob
import gzip
import json
import os
import re
import sys
import unicodedata
from collections import defaultdict

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from regions_registry import REGIONS, resolve_region  # noqa: E402

OUT_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "..", "backend", "app", "data", "regional"
)

YEAR_STR_RE = re.compile(r"^(19[89]\d|20[0-2]\d)(/\d{2,4})?\s*(?:\d\))?$")
TABLE_CODE_RE = re.compile(r"^(\d+)\.(\d+)(?:\.(\d+))?\.?$")
MULT_RE = re.compile(r"^в\s+(\d+[.,]\d+)\s*р\.?$", re.I)

TRANSLIT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh",
    "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o",
    "п": "p", "р": "r", "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "ts",
    "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu",
    "я": "ya",
}


def translit_slug(text: str, max_words: int = 7) -> str:
    text = unicodedata.normalize("NFKC", text).lower()
    out = []
    for ch in text:
        if ch in TRANSLIT:
            out.append(TRANSLIT[ch])
        elif ch.isalnum() and ch.isascii():
            out.append(ch)
        else:
            out.append(" ")
    words = "".join(out).split()
    return "-".join(words[:max_words])


def norm_text(s: str) -> str:
    s = s.replace("\xa0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def clean_title(s: str) -> str:
    """Убрать номер таблицы и сноски из заголовка, привести к sentence case."""
    s = norm_text(s)
    s = re.sub(r"^\s*\d+\.\d+(\.\d+)?\.?\s*", "", s)
    s = re.sub(r"\d\)", "", s)
    s = s.strip(" .;:")
    if s.isupper():
        s = s.capitalize()
    return s


def parse_toc(wb) -> dict:
    """Оглавление раздела (первый лист): код таблицы -> каноническое имя."""
    toc = {}
    first = wb[wb.sheetnames[0]]
    for row in first.iter_rows(values_only=True):
        cells = [c for c in row if c is not None and str(c).strip()]
        if len(cells) < 2:
            continue
        m = TABLE_CODE_RE.match(str(cells[0]).strip())
        if not m:
            continue
        code = ".".join(g for g in m.groups() if g)
        name = norm_text(str(cells[1]))
        name = re.sub(r"\d\)", "", name).strip(" .")
        if name:
            toc[code] = name
    return toc


def find_header(rows):
    """Найти (title_row_idx, note, year_row_idx, years_by_col)."""
    year_idx = None
    years_by_col = {}
    for i, row in enumerate(rows[:130]):
        cols = {}
        for j, c in enumerate(row):
            if j == 0:
                continue
            y = None
            if isinstance(c, (int, float)) and 1985 <= c <= 2026:
                y = int(c)
            elif isinstance(c, str):
                m = YEAR_STR_RE.match(c.strip())
                if m:
                    y = int(m.group(1))
            if y:
                cols[j] = y
        if len(cols) >= 2:
            year_idx, years_by_col = i, cols
            break
    if year_idx is None:
        return None, "", None, {}
    title_idx, note = None, ""
    for i in range(max(0, year_idx - 8), year_idx):
        c0 = rows[i][0]
        if not isinstance(c0, str):
            continue
        t = norm_text(c0)
        if re.match(r"^\s*\d+\.\d+(\.\d+)?\.?\s+\S", t):
            title_idx = i
        elif title_idx is not None and t.startswith("("):
            note = t.strip("() ")
    return title_idx, note, year_idx, years_by_col


def extract_unit(note: str, title: str) -> str:
    if note:
        parts = [p.strip() for p in note.split(";") if p.strip()]
        if parts:
            cand = parts[-1]
            cand = re.sub(r"\d\)", "", cand).strip(" .")
            if len(cand) < 60 and not re.search(r"\d{4}", cand):
                return cand
    m = re.search(r";\s*([^;()]{3,50})\)?\s*$", title)
    if m:
        return m.group(1).strip()
    return ""


def parse_value(c, is_percentish: bool):
    """Значение ячейки -> float | None. None = точки нет (законно)."""
    if c is None:
        return None
    if isinstance(c, (int, float)):
        return float(c)
    s = norm_text(str(c))
    if not s or s in {"-", "–", "—", "‐", "−"} or s.startswith(("…", "...", "….", "..")):
        return None
    if re.fullmatch(r"[.,\s]+", s):
        return None
    # значение со сноской: '17,83)' -> 17,8 + сноска 3); '1202)' -> 120 + 2)
    if s.endswith(")"):
        s = re.sub(r"\s*\d\)$", "", s)
    s2 = s.replace(",", ".").replace(" ", "")
    try:
        return float(s2)
    except ValueError:
        pass
    m = MULT_RE.match(s)
    if m and is_percentish:
        return round(float(m.group(1).replace(",", ".")) * 100, 1)
    return "UNPARSED"


def main(src_dir: str):
    os.makedirs(OUT_DIR, exist_ok=True)
    files = sorted(
        glob.glob(os.path.join(src_dir, "**", "Раздел*.xlsx"), recursive=True),
        key=lambda p: int(re.search(r"Раздел\s*(\d+)", os.path.basename(p)).group(1)),
    )
    assert files, f"Не найдено файлов 'Раздел*.xlsx' в {src_dir}"

    indicators = []
    points = []  # (code, region_slug, year, value)
    report = {
        "files": len(files), "sheets_total": 0, "sheets_parsed": 0,
        "points": 0, "unresolved_rows": defaultdict(int),
        "unparsed_values": [], "skipped_sheets": [], "duplicate_year_cols": [],
    }
    slug_seen = {}

    for f in files:
        base = os.path.basename(f)
        sec_num = int(re.search(r"Раздел\s*(\d+)", base).group(1))
        sec_name = norm_text(re.sub(r"^Раздел\s*\d+\s*-\s*|\.xlsx$", "", base))
        # имена файлов сокращены — публичное имя раздела расшифровываем
        sec_name = {
            "Зем ресурсы и охр окр прир среды":
                "Земельные ресурсы и охрана окружающей среды",
        }.get(sec_name, sec_name)
        wb = load_workbook(f, read_only=True, data_only=True)
        toc = parse_toc(wb)

        for sh in wb.sheetnames:
            m = TABLE_CODE_RE.match(sh.strip())
            if not m:
                continue
            report["sheets_total"] += 1
            table_code = ".".join(g for g in m.groups() if g)
            ws = wb[sh]
            rows = list(ws.iter_rows(values_only=True))
            title_idx, note, year_idx, years_by_col = find_header(rows)
            if year_idx is None:
                report["skipped_sheets"].append((base, sh, "no_year_row"))
                continue

            # имя: TOC приоритетен; для подтаблиц 1.6.1 склеиваем с родителем
            title_sheet = clean_title(str(rows[title_idx][0])) if title_idx is not None else ""
            name = toc.get(table_code) or title_sheet
            parent_code = ".".join(table_code.split(".")[:2])
            if table_code.count(".") == 2 and parent_code in toc:
                child = name if name != toc.get(parent_code) else title_sheet
                if child and child.lower() != toc[parent_code].lower():
                    # первую букву вниз только для обычного слова (не аббревиатуры)
                    if len(child) > 1 and child[1].islower():
                        child = child[0].lower() + child[1:]
                    name = f"{toc[parent_code]} — {child}"
                else:
                    name = toc[parent_code]
            name = norm_text(name)
            if not name:
                report["skipped_sheets"].append((base, sh, "no_name"))
                continue

            # дубли лет в шапке (два блока колонок) — берём ПЕРВУЮ колонку года
            seen_years = {}
            for col, y in sorted(years_by_col.items()):
                if y in seen_years:
                    report["duplicate_year_cols"].append((base, sh, y))
                    continue
                seen_years[y] = col
            col_by_year = {y: c for y, c in seen_years.items()}

            unit = extract_unit(note, str(rows[title_idx][0]) if title_idx is not None else "")
            is_pct = bool(re.search(r"процент|%", (unit + " " + note + " " + name).lower()))

            # slug
            base_slug = translit_slug(name) or f"tab-{table_code.replace('.', '-')}"
            slug = base_slug
            if slug in slug_seen:
                slug = f"{base_slug}-{table_code.replace('.', '-')}"
            slug_seen[slug] = table_code

            # В-8: 79 таблиц публикуются без единицы в шапке — кураторский
            # фолбэк, чтобы пересборка артефакта не потеряла заполненные unit.
            from unit_fallbacks import fill_unit
            from unit_normalize import normalize_unit
            unit = fill_unit(slug, unit)
            unit, note = normalize_unit(slug, unit, note, name)

            n_points_before = len(points)
            years_present = set()
            any_region_seen = False
            for row in rows[year_idx + 1:]:
                c0 = row[0]
                if not isinstance(c0, str) or not c0.strip():
                    continue
                t = c0.strip()
                if t.startswith(("_", "¾")):
                    break  # разделитель сносок — дальше только примечания
                if t.startswith("(") or re.match(r"^\d\)", t):
                    continue
                rslug, scale = resolve_region(t)
                if rslug is None:
                    n_vals = sum(
                        1 for c in row[1:]
                        if isinstance(c, (int, float))
                        or (isinstance(c, str) and re.match(r"^\s*\d", c))
                    )
                    # Первая строка с данными до первого опознанного региона —
                    # это РФ, подписанная именем показателя (разделы 8, 13, 15, 19)
                    if not any_region_seen and n_vals >= 3:
                        rslug, scale = "russia", 1.0
                        if re.search(r",\s*(млрд|млн)\s", t):
                            scale = 1000.0
                    else:
                        key = norm_text(t)[:60]
                        if n_vals >= 3 and "в том числе" not in key.lower():
                            report["unresolved_rows"][key] += 1
                        continue
                any_region_seen = True
                for y, col in col_by_year.items():
                    v = parse_value(row[col] if col < len(row) else None, is_pct)
                    if v is None:
                        continue
                    if v == "UNPARSED":
                        report["unparsed_values"].append(
                            (base, sh, t[:30], y, str(row[col])[:20])
                        )
                        continue
                    points.append((slug, rslug, y, v * scale))
                    years_present.add(y)

            n_new = len(points) - n_points_before
            if n_new == 0:
                report["skipped_sheets"].append((base, sh, "0_points"))
                continue
            report["sheets_parsed"] += 1
            indicators.append({
                "code": slug,
                "table_code": table_code,
                "section_num": sec_num,
                "section_name": sec_name,
                "name": name,
                "unit": unit,
                "note": note,
                "source_sheet": f"{base} / {sh}",
                "year_min": min(years_present),
                "year_max": max(years_present),
                "n_points": n_new,
            })
        wb.close()

    report["points"] = len(points)
    report["indicators"] = len(indicators)
    report["unresolved_rows"] = dict(report["unresolved_rows"])

    with open(os.path.join(OUT_DIR, "indicators.json"), "w", encoding="utf-8") as fh:
        json.dump(indicators, fh, ensure_ascii=False, indent=1)
    with open(os.path.join(OUT_DIR, "regions.json"), "w", encoding="utf-8") as fh:
        json.dump(REGIONS, fh, ensure_ascii=False, indent=1)
    with gzip.open(os.path.join(OUT_DIR, "data.csv.gz"), "wt", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter=";")
        w.writerow(["indicator_code", "region_slug", "year", "value"])
        for p in points:
            w.writerow(p)
    with open(os.path.join(OUT_DIR, "parse_report.json"), "w", encoding="utf-8") as fh:
        json.dump(report, fh, ensure_ascii=False, indent=1, default=str)

    print(f"файлов={report['files']} листов={report['sheets_total']} "
          f"распарсено={report['sheets_parsed']} показателей={report['indicators']} "
          f"точек={report['points']}")
    print(f"пропущено листов: {len(report['skipped_sheets'])}")
    for s in report["skipped_sheets"]:
        print("  SKIP:", s)
    print(f"неопознанных строк-имён: {len(report['unresolved_rows'])}")
    for k, v in sorted(report["unresolved_rows"].items(), key=lambda x: -x[1])[:15]:
        print(f"  {v:4d} {k}")
    print(f"неразобранных значений: {len(report['unparsed_values'])}")
    for u in report["unparsed_values"][:15]:
        print("  ", u)
    print(f"дублей лет: {len(report['duplicate_year_cols'])}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "/tmp/cur")
