#!/usr/bin/env python3
"""Харвестер каталогов зарубежных статведомств (разведка «что можно достать»).

Зачем: ручная инвентаризация в Excel не масштабируется — у одного Eurostat 8957
наборов. Скрипт забирает каталог источника ЦЕЛИКОМ через его же API и нормализует
в `research.source_catalog`. Excel строится как срез поверх таблицы, а не как
первичный носитель.

Схема `research` намеренно отдельная от `public`: это исследовательский артефакт,
он не входит в `Base.metadata` и не участвует в alembic-миграциях, поэтому
`backend/scripts/check-migration-drift.py` (сравнивает только default-схему) его
не видит и CI не краснеет.

Метрика «обогащённости» (ответ на вопрос «так ли глубоко, как у Росстата») —
не число наборов, а размерность: `n_dims` (оси декомпозиции) и `n_geo_levels`
(территориальные уровни). Обе берутся из метаданных источника, не на глаз.

Источники прототипа — те, что отдают каталог без ключа:
  eurostat  TOC (Германия, Франция + весь ЕС), границы истории в одном запросе
  ibge      Бразилия, дерево агрегатов + метаданные по таблице
  ons       Великобритания, /v1/datasets + версии

Примеры:
    python3 scripts/catalog-harvest.py --sources eurostat --no-db --excel /tmp/x.xlsx
    python3 scripts/catalog-harvest.py --sources ibge --enrich --workers 12
    python3 scripts/catalog-harvest.py --sources eurostat,ibge,ons --enrich
"""

from __future__ import annotations

import argparse
import csv
import gzip
import io
import json
import os
import re
import ssl
import sys
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field, asdict
from typing import Callable, Iterable, Iterator

_UA = "Mozilla/5.0 (compatible; forecasteconomy-catalog-harvest/1.0)"
_SSL = ssl.create_default_context()
_SSL.check_hostname = False
_SSL.verify_mode = ssl.CERT_NONE


# ---------------------------------------------------------------- транспорт

def fetch(url: str, timeout: int = 120, retries: int = 3) -> bytes:
    last: Exception | None = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(
                url, headers={"User-Agent": _UA, "Accept-Encoding": "gzip"}
            )
            with urllib.request.urlopen(req, timeout=timeout, context=_SSL) as r:
                body = r.read()
            return gzip.decompress(body) if body[:2] == b"\x1f\x8b" else body
        except Exception as exc:  # noqa: BLE001 — сеть, интересен только последний
            last = exc
            time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"{url}: {type(last).__name__}: {last}")


def fetch_json(url: str, **kw):
    return json.loads(fetch(url, **kw))


# ---------------------------------------------------------------- модель

@dataclass
class Entry:
    """Одна запись каталога, нормализованная под все источники."""

    source: str
    country: str
    agency: str
    dataset_id: str
    title: str
    frequency: str | None = None      # D/W/M/Q/A/mixed
    period_start: str | None = None
    period_end: str | None = None
    n_dims: int | None = None         # осей декомпозиции
    n_geo_levels: int | None = None   # территориальных уровней
    dims: list[str] = field(default_factory=list)
    url: str = ""
    extra: dict = field(default_factory=dict)


# Формат периода → частота. Порядок важен: D перед M, иначе '1980-01-01' → M.
_FREQ_BY_PERIOD = [
    (re.compile(r"^\d{4}-\d{2}-\d{2}$"), "D"),
    (re.compile(r"^\d{4}-W\d{1,2}$", re.I), "W"),
    (re.compile(r"^\d{4}-Q\d$", re.I), "Q"),
    (re.compile(r"^\d{4}-S\d$", re.I), "S"),
    (re.compile(r"^\d{4}-\d{2}$"), "M"),
    (re.compile(r"^\d{4}$"), "A"),
]


def freq_from_period(period: str | None) -> str | None:
    if not period:
        return None
    p = period.strip()
    for pat, code in _FREQ_BY_PERIOD:
        if pat.match(p):
            return code
    return None


# ---------------------------------------------------------------- Eurostat

def harvest_eurostat(enrich: bool, limit: int | None, workers: int) -> Iterator[Entry]:
    """TOC отдаёт код, имя, тип и ГОТОВЫЕ границы истории — один запрос на всё."""
    raw = fetch(
        "https://ec.europa.eu/eurostat/api/dissemination/catalogue/toc/txt?lang=en"
    ).decode("utf-8", "ignore")
    rows = list(csv.reader(io.StringIO(raw), delimiter="\t", quotechar='"'))
    n = 0
    for r in rows[1:]:
        if len(r) < 7 or r[2].strip() not in ("dataset", "table"):
            continue
        code = r[1].strip()
        start, end = r[5].strip(), r[6].strip()
        yield Entry(
            source="eurostat",
            country="EU (вкл. DE, FR)",
            agency="Eurostat",
            dataset_id=code,
            title=r[0].strip(),
            frequency=freq_from_period(start),
            period_start=start or None,
            period_end=end or None,
            url=f"https://ec.europa.eu/eurostat/databrowser/view/{code}/default/table",
            extra={"kind": r[2].strip(), "last_update": r[3].strip(),
                   "last_structure_change": r[4].strip(),
                   "values": (r[7].strip() if len(r) > 7 else "")},
        )
        n += 1
        if limit and n >= limit:
            return


# ---------------------------------------------------------------- IBGE (Бразилия)

_IBGE_FREQ = {"mensal": "M", "trimestral": "Q", "anual": "A", "semestral": "S",
              "decenal": "A", "quinquenal": "A", "bienal": "A", "diaria": "D",
              "semanal": "W", "irregular": "irregular"}


def _ibge_meta(agg_id: str) -> dict:
    try:
        return fetch_json(
            f"https://servicodados.ibge.gov.br/api/v3/agregados/{agg_id}/metadados",
            timeout=60, retries=2,
        )
    except Exception:
        return {}


def harvest_ibge(enrich: bool, limit: int | None, workers: int) -> Iterator[Entry]:
    tree = fetch_json("https://servicodados.ibge.gov.br/api/v3/agregados")
    flat: list[tuple[str, str, str]] = []  # (id, nome, pesquisa)
    for research in tree:
        for agg in research.get("agregados", []):
            flat.append((str(agg["id"]), agg.get("nome", ""), research.get("nome", "")))
    if limit:
        flat = flat[:limit]

    metas: dict[str, dict] = {}
    if enrich:
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for (aid, _, _), meta in zip(flat, pool.map(_ibge_meta, [f[0] for f in flat])):
                metas[aid] = meta

    for aid, nome, pesquisa in flat:
        m = metas.get(aid, {})
        per = m.get("periodicidade") or {}
        geo = m.get("nivelTerritorial") or {}
        classif = m.get("classificacoes") or []
        variaveis = m.get("variaveis") or []
        # Оси декомпозиции = классификаторы (продукт/отрасль/пол/…) + число
        # измеряемых переменных в таблице.
        n_dims = (len(classif) + len(variaveis)) if m else None
        n_geo = sum(len(v) for v in geo.values() if isinstance(v, list)) if geo else None
        yield Entry(
            source="ibge",
            country="Бразилия",
            agency="IBGE (SIDRA)",
            dataset_id=aid,
            title=nome,
            frequency=_IBGE_FREQ.get(str(per.get("frequencia", "")).lower()),
            period_start=str(per["inicio"]) if per.get("inicio") else None,
            period_end=str(per["fim"]) if per.get("fim") else None,
            n_dims=n_dims,
            n_geo_levels=n_geo,
            dims=[c.get("nome", "") for c in classif][:20],
            url=f"https://sidra.ibge.gov.br/tabela/{aid}",
            extra={"pesquisa": pesquisa,
                   "variaveis": [v.get("nome", "") for v in variaveis][:12]},
        )


# ---------------------------------------------------------------- ONS (Великобритания)

_ONS_FREQ = {"monthly": "M", "quarterly": "Q", "annual": "A", "annually": "A",
             "weekly": "W", "daily": "D", "yearly": "A"}


def _ons_version(url: str) -> dict:
    try:
        return fetch_json(url, timeout=60, retries=2)
    except Exception:
        return {}


def harvest_ons(enrich: bool, limit: int | None, workers: int) -> Iterator[Entry]:
    items: list[dict] = []
    offset = 0
    while True:
        page = fetch_json(
            f"https://api.beta.ons.gov.uk/v1/datasets?limit=100&offset={offset}"
        )
        batch = page.get("items") or []
        items.extend(batch)
        offset += len(batch)
        if not batch or offset >= (page.get("total_count") or 0):
            break
    if limit:
        items = items[:limit]

    versions: dict[str, dict] = {}
    if enrich:
        urls = {}
        for it in items:
            href = ((it.get("links") or {}).get("latest_version") or {}).get("href")
            if href:
                urls[it["id"]] = href
        with ThreadPoolExecutor(max_workers=workers) as pool:
            for did, ver in zip(urls, pool.map(_ons_version, urls.values())):
                versions[did] = ver

    for it in items:
        ver = versions.get(it["id"], {})
        dims = [d.get("name", "") for d in (ver.get("dimensions") or [])]
        # Время — не ось декомпозиции: вычитаем, чтобы метрика была сравнима с IBGE.
        n_dims = len([d for d in dims if d.lower() not in ("time", "geography")]) or None
        yield Entry(
            source="ons",
            country="Великобритания",
            agency="ONS",
            dataset_id=it["id"],
            title=it.get("title", ""),
            frequency=_ONS_FREQ.get(str(it.get("release_frequency", "")).strip().lower()),
            period_start=(ver.get("release_date") or "")[:10] or None,
            period_end=(it.get("last_updated") or "")[:10] or None,
            n_dims=n_dims,
            n_geo_levels=1 if any(d.lower() == "geography" for d in dims) else None,
            dims=dims[:20],
            url=f"https://www.ons.gov.uk/datasets/{it['id']}",
            extra={"keywords": (it.get("keywords") or [])[:8],
                   "unit": it.get("unit_of_measure") or "",
                   "national_statistic": it.get("national_statistic")},
        )


HARVESTERS: dict[str, Callable[[bool, int | None, int], Iterable[Entry]]] = {
    "eurostat": harvest_eurostat,
    "ibge": harvest_ibge,
    "ons": harvest_ons,
}


# ---------------------------------------------------------------- Postgres

_DDL = """
create schema if not exists research;
create table if not exists research.source_catalog (
    source        text not null,
    country       text not null,
    agency        text not null,
    dataset_id    text not null,
    title         text,
    frequency     text,
    period_start  text,
    period_end    text,
    n_dims        integer,
    n_geo_levels  integer,
    dims          jsonb,
    url           text,
    extra         jsonb,
    harvested_at  timestamptz not null default now(),
    primary key (source, dataset_id)
);
create index if not exists ix_source_catalog_freq on research.source_catalog (source, frequency);
create index if not exists ix_source_catalog_start on research.source_catalog (period_start);
"""

_UPSERT = """
insert into research.source_catalog
    (source, country, agency, dataset_id, title, frequency, period_start,
     period_end, n_dims, n_geo_levels, dims, url, extra, harvested_at)
values (%(source)s, %(country)s, %(agency)s, %(dataset_id)s, %(title)s,
        %(frequency)s, %(period_start)s, %(period_end)s, %(n_dims)s,
        %(n_geo_levels)s, %(dims)s, %(url)s, %(extra)s, now())
on conflict (source, dataset_id) do update set
    country = excluded.country, agency = excluded.agency, title = excluded.title,
    frequency = excluded.frequency, period_start = excluded.period_start,
    period_end = excluded.period_end, n_dims = excluded.n_dims,
    n_geo_levels = excluded.n_geo_levels, dims = excluded.dims,
    url = excluded.url, extra = excluded.extra, harvested_at = now()
"""


def dsn_from_env() -> str:
    raw = os.environ.get("RUSTATS_DATABASE_URL") or ""
    if not raw:
        env = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env")
        if os.path.exists(env):
            for line in open(env, encoding="utf-8"):
                if line.startswith("RUSTATS_DATABASE_URL="):
                    raw = line.split("=", 1)[1].strip()
                    break
    if not raw:
        raw = "postgresql://rustats:rustats_dev@localhost:5434/rustats"
    return raw.replace("postgresql+asyncpg://", "postgresql://")


def read_db(sources: list[str]) -> list[Entry]:
    """Срез строится из таблицы, а не из свежего прогона: Excel всегда отражает БД."""
    import psycopg2  # noqa: PLC0415

    conn = psycopg2.connect(dsn_from_env())
    try:
        with conn.cursor() as cur:
            cur.execute(
                "select source, country, agency, dataset_id, title, frequency,"
                " period_start, period_end, n_dims, n_geo_levels, dims, url, extra"
                " from research.source_catalog where source = any(%s)",
                (sources,),
            )
            return [
                Entry(source=r[0], country=r[1], agency=r[2], dataset_id=r[3],
                      title=r[4] or "", frequency=r[5], period_start=r[6],
                      period_end=r[7], n_dims=r[8], n_geo_levels=r[9],
                      dims=r[10] or [], url=r[11] or "", extra=r[12] or {})
                for r in cur.fetchall()
            ]
    finally:
        conn.close()


def write_db(entries: list[Entry]) -> int:
    import psycopg2  # noqa: PLC0415 — опциональная зависимость для --no-db прогонов
    from psycopg2.extras import execute_batch

    conn = psycopg2.connect(dsn_from_env())
    try:
        with conn, conn.cursor() as cur:
            cur.execute(_DDL)
            payload = []
            for e in entries:
                d = asdict(e)
                d["dims"] = json.dumps(d["dims"], ensure_ascii=False)
                d["extra"] = json.dumps(d["extra"], ensure_ascii=False)
                payload.append(d)
            execute_batch(cur, _UPSERT, payload, page_size=500)
        return len(entries)
    finally:
        conn.close()


# ---------------------------------------------------------------- Excel-срез

def _depth_years(e: Entry) -> int | None:
    def year(p: str | None) -> int | None:
        if not p:
            return None
        m = re.match(r"^(\d{4})", p.strip())
        return int(m.group(1)) if m else None

    a, b = year(e.period_start), year(e.period_end)
    return (b - a) if (a and b and b >= a) else None


def write_excel(entries: list[Entry], path: str, limit: int) -> None:
    from collections import Counter

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(bold=True, color="FFFFFF", size=10)

    def put(ws, headers: list[str], rows: list[list], widths: list[int]) -> None:
        ws.append(headers)
        for c in ws[1]:
            c.fill, c.font = head_fill, head_font
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for r in rows:
            ws.append(r)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb = Workbook()

    # Лист 1 — сводка по источникам: то, ради чего скан затевался.
    ws = wb.active
    ws.title = "Сводка"
    rows = []
    for src in sorted({e.source for e in entries}):
        sub = [e for e in entries if e.source == src]
        freqs = Counter(e.frequency or "?" for e in sub)
        depths = [d for d in (_depth_years(e) for e in sub) if d is not None]
        starts = sorted(e.period_start for e in sub if e.period_start)
        dims = [e.n_dims for e in sub if e.n_dims]
        rows.append([
            src, sub[0].country, sub[0].agency, len(sub),
            ", ".join(f"{k}:{v}" for k, v in freqs.most_common(7)),
            starts[0] if starts else "—",
            f"{sum(depths) / len(depths):.0f}" if depths else "—",
            max(depths) if depths else "—",
            f"{sum(dims) / len(dims):.1f}" if dims else "нет метаданных",
            sum(1 for e in sub if (_depth_years(e) or 0) >= 25),
        ])
    put(ws,
        ["Источник", "Страна", "Ведомство", "Наборов в каталоге", "Частоты",
         "Самый ранний старт", "Средняя глубина, лет", "Макс. глубина, лет",
         "Среднее число осей", "Наборов с историей 25+ лет"],
        rows, [12, 20, 16, 18, 34, 16, 16, 16, 16, 18])

    # Лист 2 — сами наборы (срез; полный массив живёт в research.source_catalog).
    ws2 = wb.create_sheet("Наборы")
    entries_sorted = sorted(entries, key=lambda e: (-(_depth_years(e) or 0), e.source))
    body = [
        [e.source, e.country, e.dataset_id, e.title[:250], e.frequency or "",
         e.period_start or "", e.period_end or "", _depth_years(e) or "",
         e.n_dims or "", e.n_geo_levels or "", "; ".join(e.dims)[:200], e.url]
        for e in entries_sorted[:limit]
    ]
    put(ws2,
        ["Источник", "Страна", "ID набора", "Название", "Частота", "История с",
         "История по", "Глубина, лет", "Осей", "Гео-уровней", "Оси декомпозиции", "URL"],
        body, [11, 18, 20, 60, 9, 12, 12, 12, 8, 12, 44, 52])
    if len(entries_sorted) > limit:
        ws2.append([f"Показано {limit} из {len(entries_sorted)}; "
                    f"полный массив — research.source_catalog"])

    wb.save(path)


# ---------------------------------------------------------------- CLI

def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--sources", default="eurostat,ibge,ons",
                    help="через запятую: " + ", ".join(HARVESTERS))
    ap.add_argument("--enrich", action="store_true",
                    help="дотянуть метаданные по каждому набору (оси, период) — медленно")
    ap.add_argument("--limit", type=int, default=None, help="ограничить наборов на источник")
    ap.add_argument("--workers", type=int, default=10)
    ap.add_argument("--no-db", action="store_true")
    ap.add_argument("--from-db", action="store_true",
                    help="не ходить в сеть: собрать срез из research.source_catalog")
    ap.add_argument("--excel", default=None, help="путь для среза в .xlsx")
    ap.add_argument("--excel-limit", type=int, default=5000)
    args = ap.parse_args()

    names = [s.strip() for s in args.sources.split(",") if s.strip()]
    unknown = [n for n in names if n not in HARVESTERS]
    if unknown:
        print(f"неизвестные источники: {unknown}; доступны: {list(HARVESTERS)}")
        return 2

    if args.from_db:
        entries = read_db(names)
        print(f"[db] прочитано {len(entries)} строк")
        if args.excel and entries:
            write_excel(entries, args.excel, args.excel_limit)
            print(f"[excel] {args.excel}")
        return 0 if entries else 1

    all_entries: list[Entry] = []
    for name in names:
        t0 = time.time()
        try:
            got = list(HARVESTERS[name](args.enrich, args.limit, args.workers))
        except Exception as exc:  # noqa: BLE001 — источник упал, остальные продолжаем
            print(f"[{name}] FAIL {type(exc).__name__}: {exc}")
            continue
        all_entries.extend(got)
        with_hist = sum(1 for e in got if e.period_start)
        with_dims = sum(1 for e in got if e.n_dims)
        print(f"[{name}] {len(got)} наборов за {time.time() - t0:.1f}s "
              f"(с историей {with_hist}, с осями {with_dims})")

    if not all_entries:
        print("пусто — нечего писать")
        return 1

    if not args.no_db:
        try:
            print(f"[db] записано {write_db(all_entries)} строк в research.source_catalog")
        except Exception as exc:  # noqa: BLE001
            print(f"[db] FAIL {type(exc).__name__}: {exc}")
    if args.excel:
        write_excel(all_entries, args.excel, args.excel_limit)
        print(f"[excel] {args.excel}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
