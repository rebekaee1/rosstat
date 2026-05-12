"""Tests for CbrTradeGoodsMonthlyParser + CbrTradeServicesMonthlyParser."""

import io
from datetime import date, datetime

import openpyxl
import xlwt

from app.services.cbr_trade_goods_monthly_parser import (
    _parse_month,
    _parse_year,
    parse_trade_goods_monthly_xls,
)
from app.services.cbr_trade_services_monthly_parser import (
    _parse_header_date,
    parse_trade_services_monthly_xlsx,
)


def _build_goods_xls() -> bytes:
    """Build minimal XLS mirror of trade.xls лист 'Ежемесячные'.

    Layout (cols 0-16):
    - 0: Год | 1: Месяц | 2: Exp_total | 8: Imp_total | 14: Trade_balance
    """
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Ежемесячные")

    for ri, vals in enumerate([
        ["Внешняя торговля Российской Федерации"],
        ["(по методологии платёжного баланса)"],
        [""],
        ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "млн долларов США"],
        ["Период", "", "Экспорт товаров (ФОБ) "],
        ["Год", "Месяц"],
        ["", "месяц"],
    ]):
        for ci, val in enumerate(vals):
            ws.write(ri, ci, val)

    rows = [
        (2024, "Янв",  100.0, "x", 60.0, "x", 40.0, "x",  50.0, "x", 30.0, "x", 20.0, "x",  50.0, 30.0, 20.0),
        (2024, "Фев",  110.0, "x", 65.0, "x", 45.0, "x",  55.0, "x", 33.0, "x", 22.0, "x",  55.0, 32.0, 23.0),
        (2024, "Мар",  120.0, "x", 72.0, "x", 48.0, "x",  60.0, "x", 36.0, "x", 24.0, "x",  60.0, 36.0, 24.0),
        (2024, "Дек",  150.0, "x", 90.0, "x", 60.0, "x",  70.0, "x", 42.0, "x", 28.0, "x",  80.0, 48.0, 32.0),
        (2025, "Янв",  130.0, "x", 78.0, "x", 52.0, "x",  65.0, "x", 39.0, "x", 26.0, "x",  65.0, 39.0, 26.0),
    ]
    for ri, row in enumerate(rows, start=7):
        for ci, val in enumerate(row):
            ws.write(ri, ci, val)

    # Trailing notes (should be skipped by parser).
    ws.write(12, 0, "Примечание:")
    ws.write(13, 0, "Дата последнего обновления: 14 апреля 2026 года.")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_services_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "месяцы "

    ws.cell(row=2, column=1).value = "Внешняя торговля Российской Федерации услугами"

    ws.cell(row=4, column=2).value = datetime(2024, 1, 1)
    ws.cell(row=4, column=3).value = datetime(2024, 2, 1)
    ws.cell(row=4, column=4).value = datetime(2024, 12, 1)
    ws.cell(row=4, column=5).value = "янв.26\n(оценка)"
    ws.cell(row=4, column=6).value = "фев.26"

    rows = [
        (5, "Услуги ",       -1.0,  -1.5,  -2.0,  -1.8,  -2.1),
        (6, "Экспорт услуг",  5.0,   5.2,   6.0,   4.5,   4.8),
        (7, "Импорт услуг",   6.0,   6.7,   8.0,   6.3,   6.9),
        (8, "Транспортные услуги", 0.5, 0.6, 0.8, 0.4, 0.5),
        (9, "Экспорт",        2.0,   2.1,   2.3,   2.0,   2.1),
        (10, "Импорт ",       1.5,   1.5,   1.5,   1.6,   1.6),
    ]
    for r, label, *vals in rows:
        ws.cell(row=r, column=1).value = label
        for i, v in enumerate(vals):
            ws.cell(row=r, column=2 + i).value = v

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseMonth:
    """Russian short month name → 1-12."""

    def test_basic(self):
        assert _parse_month("Янв") == 1
        assert _parse_month("Дек") == 12

    def test_lowercase_and_whitespace(self):
        assert _parse_month("  май ") == 5
        assert _parse_month("ИЮН") == 6

    def test_unknown(self):
        assert _parse_month("XYZ") is None
        assert _parse_month("") is None
        assert _parse_month(None) is None


class TestParseYear:
    def test_basic(self):
        assert _parse_year(1997.0) == 1997
        assert _parse_year(2026) == 2026

    def test_string(self):
        assert _parse_year("1997") == 1997

    def test_out_of_range(self):
        assert _parse_year(1500) is None
        assert _parse_year("Примечание:") is None
        assert _parse_year("") is None


class TestParseGoodsXls:
    def test_exports(self):
        content = _build_goods_xls()
        points = parse_trade_goods_monthly_xls(content, "exports-monthly")
        assert len(points) == 5
        assert points[0].date == date(2024, 1, 1)
        assert points[0].value == 100.0
        assert points[-1].date == date(2025, 1, 1)
        assert points[-1].value == 130.0

    def test_imports(self):
        content = _build_goods_xls()
        points = parse_trade_goods_monthly_xls(content, "imports-monthly")
        assert len(points) == 5
        assert points[0].value == 50.0
        assert points[-1].value == 65.0

    def test_trade_balance(self):
        """col 14 (Сальдо) — должен совпадать с exports - imports по datapoints."""
        content = _build_goods_xls()
        points = parse_trade_goods_monthly_xls(content, "trade-balance-monthly")
        assert len(points) == 5
        assert points[0].value == 50.0  # 100 - 50
        assert points[1].value == 55.0  # 110 - 55
        assert points[-1].value == 65.0  # 130 - 65

    def test_skips_trailing_notes(self):
        """Rows с non-numeric col 0 ('Примечание:') не попадают в output."""
        content = _build_goods_xls()
        points = parse_trade_goods_monthly_xls(content, "exports-monthly")
        # Только 5 валидных rows, не 7 (учитывая «Примечание:»).
        assert len(points) == 5

    def test_unknown_target(self):
        content = _build_goods_xls()
        try:
            parse_trade_goods_monthly_xls(content, "bogus-target")
            assert False, "expected ValueError"
        except ValueError:
            pass


class TestParseHeaderDate:
    def test_datetime(self):
        assert _parse_header_date(datetime(2024, 5, 1)) == date(2024, 5, 1)

    def test_estimate_string(self):
        assert _parse_header_date("янв.26\n(оценка)") == date(2026, 1, 1)
        assert _parse_header_date("фев.26") == date(2026, 2, 1)

    def test_full_year(self):
        assert _parse_header_date("дек.2025") == date(2025, 12, 1)

    def test_invalid(self):
        assert _parse_header_date(None) is None
        assert _parse_header_date("") is None
        assert _parse_header_date("not a date") is None


class TestParseServicesXlsx:
    def test_exports(self):
        content = _build_services_xlsx()
        points = parse_trade_services_monthly_xlsx(
            content, "services-exports-monthly",
        )
        assert len(points) == 5
        assert points[0].date == date(2024, 1, 1)
        assert points[0].value == 5.0
        # Estimate column with «фев.26» parsed correctly.
        assert points[-1].date == date(2026, 2, 1)
        assert points[-1].value == 4.8

    def test_imports(self):
        content = _build_services_xlsx()
        points = parse_trade_services_monthly_xlsx(
            content, "services-imports-monthly",
        )
        assert len(points) == 5
        assert points[0].value == 6.0
        assert points[-1].value == 6.9

    def test_skips_non_target_rows(self):
        """Парсер берёт ТОЛЬКО row 'Экспорт услуг' / 'Импорт услуг',
        не подхватывает row 'Услуги ' (balance) и 'Экспорт' под транспортными.
        """
        content = _build_services_xlsx()
        exports = parse_trade_services_monthly_xlsx(
            content, "services-exports-monthly",
        )
        # «Экспорт услуг» row 6 → 5.0 на jan 2024 (а не «Услуги » row 5 → -1.0).
        assert exports[0].value == 5.0
        assert exports[0].value != -1.0
