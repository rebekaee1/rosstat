"""Tests for Rosstat IPI parser (canonical русский Rosstat ind_baza_*.xlsx)."""

import io
from datetime import date

import openpyxl
import pytest

from app.services.rosstat_ipi_parser import (
    DataPoint,
    chain_mom_to_index_2023_base,
    merge_mom_dicts,
    parse_rosstat_ipi_mom_xlsx,
)


def _make_ind_baza_xlsx(year: int, mom_values: list[float]) -> bytes:
    """Build minimal `ind_baza_*.xlsx` mimic with sheet '1' (MoM%, ПРОМЫШЛЕННОЕ
    ПРОИЗВОДСТВО row at row 6).

    Provides 12 monthly values for given year.
    """
    months = ["январь", "февраль", "март", "апрель", "май", "июнь",
              "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
    assert len(mom_values) == 12

    wb = openpyxl.Workbook()
    wb.active.title = "Содержание"
    ws = wb.create_sheet("1")
    ws.append(["К содержанию"])
    ws.append(["Индексы производства по отдельным видам деятельности"])
    ws.append(["(в % к предыдущему месяцу)"])
    ws.append(["Наименование ОКВЭД2", "Код ОКВЭД2", f"{year} год"] + [None] * 11)
    ws.append([None, None] + months)
    ws.append(["ПРОМЫШЛЕННОЕ ПРОИЗВОДСТВО (BCDE)", "BCDE "] + mom_values)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseRosstatIpiMom:
    def test_basic_extraction(self):
        content = _make_ind_baza_xlsx(2026, [76.3, 98.8, 112.4, 95, 102, 100, 99, 98, 100, 101, 100, 119])
        result = parse_rosstat_ipi_mom_xlsx(content)
        assert len(result) == 12
        assert result[date(2026, 1, 1)] == 76.3
        assert result[date(2026, 3, 1)] == 112.4
        assert result[date(2026, 12, 1)] == 119.0

    def test_handles_month_with_footnote(self):
        """Real rosstat file has 'январь1' (footnote marker). Regression test."""
        wb = openpyxl.Workbook()
        wb.active.title = "Содержание"
        ws = wb.create_sheet("1")
        ws.append(["x"])
        ws.append(["x"])
        ws.append(["x"])
        ws.append(["lbl", "code", "2026 год", None, None])
        ws.append([None, None, "январь1", "февраль1", "март1"])
        ws.append(["ПРОМ ПРОИЗ", "BCDE ", 76.3, 98.8, 112.4])
        buf = io.BytesIO(); wb.save(buf)
        result = parse_rosstat_ipi_mom_xlsx(buf.getvalue())
        assert result[date(2026, 1, 1)] == 76.3

    def test_missing_bcde_raises(self):
        wb = openpyxl.Workbook()
        wb.active.title = "Содержание"
        wb.create_sheet("1").append(["x"])
        buf = io.BytesIO(); wb.save(buf)
        with pytest.raises(ValueError, match="BCDE"):
            parse_rosstat_ipi_mom_xlsx(buf.getvalue())

    def test_missing_sheet_raises(self):
        wb = openpyxl.Workbook()
        buf = io.BytesIO(); wb.save(buf)
        with pytest.raises(ValueError, match="sheet '1' not found"):
            parse_rosstat_ipi_mom_xlsx(buf.getvalue())


class TestChainMomToIndex:
    def test_chain_with_2023_anchor(self):
        """Все значения = 100 → chained = 100 для всех месяцев → avg 2023 = 100 → scale = 1."""
        mom = {date(2023, m, 1): 100.0 for m in range(1, 13)}
        result = chain_mom_to_index_2023_base(mom)
        assert len(result) == 12
        assert all(p.value == 100.0 for p in result)

    def test_normalize_to_2023_avg(self):
        """Если MoM колеблется, среднее за 2023 должно нормализоваться к 100."""
        mom = {date(2023, m, 1): 102.0 for m in range(1, 13)}
        result = chain_mom_to_index_2023_base(mom)
        avg = sum(p.value for p in result) / 12
        assert abs(avg - 100.0) < 0.1

    def test_no_2023_data_raises(self):
        mom = {date(2024, 1, 1): 100.0}
        with pytest.raises(ValueError, match="2023"):
            chain_mom_to_index_2023_base(mom)

    def test_chain_continuity_across_years(self):
        """Декабрь 2023 → Январь 2024: index[jan24] = index[dec23] * MoM[jan24] / 100."""
        mom = {date(2023, m, 1): 100.0 for m in range(1, 13)}
        mom[date(2024, 1, 1)] = 75.0
        result = chain_mom_to_index_2023_base(mom)
        by_date = {p.date: p.value for p in result}
        assert abs(by_date[date(2024, 1, 1)] - 75.0) < 0.1

    def test_empty_input(self):
        assert chain_mom_to_index_2023_base({}) == []


class TestMergeMomDicts:
    def test_later_overrides_earlier(self):
        a = {date(2023, 1, 1): 100.0, date(2023, 2, 1): 99.0}
        b = {date(2023, 2, 1): 101.0, date(2023, 3, 1): 102.0}
        merged = merge_mom_dicts(a, b)
        assert merged[date(2023, 1, 1)] == 100.0
        assert merged[date(2023, 2, 1)] == 101.0
        assert merged[date(2023, 3, 1)] == 102.0
