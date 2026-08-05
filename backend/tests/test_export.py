"""Тесты серверного гейта скачиваний (ADR-0007 Phase 2)."""
from app.config import settings

BODY = {
    "format": "csv",
    "filename": "ряд.csv",
    "value_label": "Значение (%)",
    "points": [
        {"date": "2020-01-01", "actual": 1.2345},
        {"date": "2020-02-01", "actual": 2.0},
        {"date": "2020-03-01", "forecast": 3.1},
    ],
}


def test_anon_limited_to_quota(auth_client, monkeypatch):
    # Пинним лимит на 2, чтобы тест не зависел от дефолта (сейчас 0 —
    # жёсткая стена регистрации: гость не качает вовсе).
    monkeypatch.setattr(settings, "download_anon_limit", 2)

    r1 = auth_client.post("/api/v1/export/table", json=BODY)
    assert r1.status_code == 200
    assert "text/csv" in r1.headers["content-type"]
    assert "attachment" in r1.headers["content-disposition"]
    text = r1.content.decode("utf-8-sig")
    assert "Источник" in text or "источник" in text.lower() or "# " in text
    # русская запятая в значении
    assert "1,2345" in text or "1,23" in text

    r2 = auth_client.post("/api/v1/export/table", json=BODY)
    assert r2.status_code == 200

    r3 = auth_client.post("/api/v1/export/table", json=BODY)
    assert r3.status_code == 403
    assert r3.json()["detail"]["code"] == "download_limit"


def test_xlsx_has_description_sheet(auth_client, monkeypatch):
    monkeypatch.setattr(settings, "download_anon_limit", 1)
    body = {
        **BODY,
        "format": "xlsx",
        "filename": "ряд.xlsx",
        "indicator_name": "Тест",
        "unit": "%",
        "country": "Германия",
        "source": "Евростат",
    }
    r = auth_client.post("/api/v1/export/table", json=body)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"
    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(r.content), read_only=True)
    assert "Описание" in wb.sheetnames
    rows = list(wb["Описание"].iter_rows(values_only=True))
    blob = " ".join(f"{a} {b}" for a, b in rows if a).lower()
    assert "евростат" in blob or "источник" in blob
    assert "германия" in blob


def test_xlsx_format(auth_client, monkeypatch):
    # Дефолтный анонимный лимит = 0 (стена регистрации); этот тест проверяет
    # сам XLSX-формат, поэтому временно разрешаем одну гостевую выгрузку.
    monkeypatch.setattr(settings, "download_anon_limit", 1)
    body = {**BODY, "format": "xlsx", "filename": "ряд.xlsx"}
    r = auth_client.post("/api/v1/export/table", json=body)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"  # ZIP-сигнатура xlsx


def test_authed_unlimited(auth_client):
    reg = auth_client.post(
        "/api/v1/auth/register",
        json={"email": "dl@example.com", "password": "password123", "consent": True},
    )
    assert reg.status_code == 201
    # Больше анонимного лимита — все успешны.
    for _ in range(5):
        r = auth_client.post("/api/v1/export/table", json=BODY)
        assert r.status_code == 200


def test_validation_rejects_bad_format(auth_client):
    r = auth_client.post("/api/v1/export/table", json={**BODY, "format": "pdf"})
    assert r.status_code == 422
