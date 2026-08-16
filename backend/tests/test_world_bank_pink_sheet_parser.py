"""Unit tests for World Bank Pink Sheet monthly commodity parser."""

from __future__ import annotations

import asyncio
from datetime import date
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock

import openpyxl

from app.services import world_bank_pink_sheet_parser as wb_mod
from app.services.world_bank_pink_sheet_parser import (
    WorldBankPinkSheetParser,
    _discover_monthly_xlsx_url,
    _parse_pink_sheet_monthly,
)


def _sample_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Prices"
    ws["A1"] = "World Bank Commodity Price Data (The Pink Sheet)"
    ws["A5"] = None
    ws["B5"] = "Coal, Australian"
    ws["C5"] = "Copper"
    ws["B6"] = "($/mt)"
    ws["C6"] = "($/mt)"
    ws["A7"] = "1960M01"
    ws["B7"] = "…"
    ws["C7"] = 715.0
    ws["A8"] = "2026M06"
    ws["B8"] = 150.36
    ws["C8"] = 13552.04
    ws["A9"] = "2026M07"
    ws["B9"] = 131.9
    ws["C9"] = 13543.0
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_discover_monthly_xlsx_url():
    html = (
        '<a href="https://thedocs.worldbank.org/en/doc/abc-0050012026/'
        'related/CMO-Historical-Data-Monthly.xlsx">Monthly prices</a>'
    )
    assert _discover_monthly_xlsx_url(html).endswith(
        "CMO-Historical-Data-Monthly.xlsx"
    )


def test_parse_pink_sheet_skips_missing_and_supports_backfill():
    points = _parse_pink_sheet_monthly(
        _sample_xlsx(),
        "Coal, Australian",
        backfill_from=date(2026, 1, 1),
    )
    assert points == [
        (date(2026, 6, 1), 150.36),
        (date(2026, 7, 1), 131.9),
    ]


def test_parse_pink_sheet_copper_column():
    points = _parse_pink_sheet_monthly(_sample_xlsx(), "Copper")
    assert points[0] == (date(1960, 1, 1), 715.0)
    assert points[-1] == (date(2026, 7, 1), 13543.0)


def test_pink_sheet_parser_fetches_configured_column(monkeypatch):
    fetched_urls: list[str] = []

    def _fake_landing() -> str:
        return (
            '<a href="https://thedocs.worldbank.org/en/doc/x/'
            'related/CMO-Historical-Data-Monthly.xlsx">x</a>'
        )

    def _fake_xlsx(url: str) -> bytes:
        fetched_urls.append(url)
        return _sample_xlsx()

    monkeypatch.setattr(wb_mod, "_fetch_landing_html", _fake_landing)
    monkeypatch.setattr(wb_mod, "_fetch_xlsx_bytes", _fake_xlsx)

    indicator = SimpleNamespace(id=1, code="coal")
    fetch_log = SimpleNamespace(error_message=None)
    cfg = {
        "pink_sheet_column": "Coal, Australian",
        "backfill_from": "2026-01-01",
    }
    points, url = asyncio.run(
        WorldBankPinkSheetParser()._fetch_and_parse(
            AsyncMock(), indicator, cfg, fetch_log,
        )
    )
    assert fetched_urls
    assert url.endswith("CMO-Historical-Data-Monthly.xlsx")
    assert points[-1] == (date(2026, 7, 1), 131.9)


def test_pink_sheet_parser_requires_column():
    indicator = SimpleNamespace(id=1, code="coal")
    fetch_log = SimpleNamespace(error_message=None)
    points, url = asyncio.run(
        WorldBankPinkSheetParser()._fetch_and_parse(
            AsyncMock(), indicator, {}, fetch_log,
        )
    )
    assert points == []
    assert "pink_sheet_column" in (fetch_log.error_message or "")
    assert "commodity-markets" in url
