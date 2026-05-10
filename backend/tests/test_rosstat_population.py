"""Tests for Rosstat Population parsers (canonical русский Rosstat, без SDDS)."""

import io
from datetime import date

import openpyxl

from app.services.rosstat_population_parser import (
    DataPoint,
    merge_population_sources,
    parse_ok_popul_xlsx,
    parse_population_history_xlsx,
    parse_popul_components_xlsx,
)


def _make_ok_popul_xlsx() -> bytes:
    """Synthetic OkPopul_Comp{YYYY}_Site.xlsx mimicking real layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Всего"

    for _ in range(5):
        ws.append([None])
    ws.append([None, "населения", "общий", "в том числе:", None, "объем", "Численность населения", "в среднем"])
    ws.append([None, "на 1 января ", "прирост", "естественный ", "миграционный", "МТП", " 2025 г.", "за 2024 г."])
    ws.append([None, "2024 г."])
    ws.append([
        "Российская Федерация*",
        146150789,
        -30861,
        -599454,
        568593,
        0,
        146119928,
        146135359,
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_popul_components_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cover"
    ws2 = wb.create_sheet("1")

    for _ in range(7):
        ws2.append([None])

    ws2.append([2019, 146794.0, -31.1, -316.2, 285.1])
    ws2.append([2020, 146749.0, -596.3, -702.8, 106.5])
    ws2.append([2021, 145478.0, -669.8, -1038.8, 369.0])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_population_history_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"

    for _ in range(6):
        ws.append([None])
    ws.append([1897, None, None])
    ws.append(["в границах Российской империи", 128.2])
    ws.append(["в современных границах", 67.5])
    ws.append([1914, None, None])
    ws.append(["в границах Российской империи", 165.7])
    ws.append(["в современных границах", 89.9])
    ws.append([1970, 129.9])
    ws.append([1971, 130.6])
    ws.append(["20242)", 146.1])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseOkPopul:
    def test_returns_latest_year(self):
        content = _make_ok_popul_xlsx()
        result = parse_ok_popul_xlsx(content)
        assert result.date == date(2025, 1, 1)
        assert result.value == 146.12

    def test_value_in_millions(self):
        content = _make_ok_popul_xlsx()
        result = parse_ok_popul_xlsx(content)
        assert 100 < result.value < 200

    def test_handles_mixed_cyrillic_latin_label(self):
        """Real rosstat file has 'Федеpация' с латинской 'p' — regression case."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Всего"
        for _ in range(5):
            ws.append([None])
        ws.append([None, "n", "n", "n", "n", "n", "n", "n"])
        ws.append([None, "n", "n", "n", "n", "n", " 2025 г.", "за 2024 г."])
        ws.append([None, "2024 г."])
        ws.append([
            "Российская Федеpация*",  # 'p' латинская
            146150789, -30861, -599454, 568593, 0, 146119928, 146135359,
        ])
        buf = io.BytesIO()
        wb.save(buf)
        result = parse_ok_popul_xlsx(buf.getvalue())
        assert result.date == date(2025, 1, 1)
        assert result.value == 146.12


class TestParsePopulationHistory:
    def test_modern_borders_and_annual_rows(self):
        content = _make_population_history_xlsx()
        result = parse_population_history_xlsx(content)
        assert [p.date for p in result] == [
            date(1897, 1, 1),
            date(1914, 1, 1),
            date(1970, 1, 1),
            date(1971, 1, 1),
            date(2024, 1, 1),
        ]
        assert result[0].value == 67.5
        assert result[1].value == 89.9
        assert result[-1].value == 146.1


class TestMergePopulationSources:
    def test_later_source_wins(self):
        history = [DataPoint(date=date(2024, 1, 1), value=146.1)]
        components = [DataPoint(date=date(2024, 1, 1), value=146.15)]
        latest = [DataPoint(date=date(2025, 1, 1), value=146.12)]
        result = merge_population_sources(history, components, latest)
        assert len(result) == 2
        assert result[0].value == 146.15
        assert result[1].value == 146.12

    def test_no_overlap_concatenates(self):
        old = [DataPoint(date=date(1897, 1, 1), value=67.5)]
        new = [DataPoint(date=date(2025, 1, 1), value=146.12)]
        result = merge_population_sources(old, new)
        assert len(result) == 2
        assert result[0].date == date(1897, 1, 1)
        assert result[1].date == date(2025, 1, 1)

    def test_empty_sources_skipped(self):
        result = merge_population_sources([], [DataPoint(date=date(2024, 1, 1), value=146.1)], [])
        assert len(result) == 1


class TestParsePopulComponents:
    def test_population(self):
        content = _make_popul_components_xlsx()
        result = parse_popul_components_xlsx(content)
        pop = result["population"]
        assert len(pop) == 3
        assert pop[0].date == date(2019, 1, 1)
        assert pop[0].value == 146.79  # 146794 / 1000

    def test_natural_growth(self):
        content = _make_popul_components_xlsx()
        result = parse_popul_components_xlsx(content)
        natural = result["natural-growth"]
        assert len(natural) == 3
        assert natural[0].value == -316.2
        assert natural[2].value == -1038.8

    def test_migration(self):
        content = _make_popul_components_xlsx()
        result = parse_popul_components_xlsx(content)
        migration = result["migration"]
        assert len(migration) == 3
        assert migration[0].value == 285.1

    def test_total_growth(self):
        content = _make_popul_components_xlsx()
        result = parse_popul_components_xlsx(content)
        total = result["total-growth"]
        assert len(total) == 3
        assert total[0].value == -31.1
