"""Tests for Rosstat GDP parser (canonical русский Rosstat, ADR-0004)."""

import io
from datetime import date

import openpyxl
import pytest

from app.services.rosstat_gdp_parser import (
    DataPoint,
    _parse_quarter_header,
    _parse_ru_number,
    parse_rosstat_gdp_quarter_grid_xlsx,
    parse_rosstat_gdp_use_xls,
    splice_at_overlap,
)


def _make_official_quarter_grid_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "9"
    ws.append(["К содержанию"])
    ws.append(["Валовой внутренний продукт (в ценах 2021г., млрд.руб.)"])
    ws.append([2011, None, None, None, 2012, None, None, None])
    ws.append(["I квартал", "II квартал", "III квартал", "IV квартал", "I квартал", "II квартал", "III квартал", "IV квартал"])
    ws.append([26368.3866, 28407.6171, 29882.9625, 32019.0647, 27871.2704, 29785.6564, 30925.3725, 32790.9561])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseQuarterHeader:
    def test_valid(self):
        assert _parse_quarter_header("Q1-2024") == date(2024, 3, 1)
        assert _parse_quarter_header("Q2-2024") == date(2024, 6, 1)
        assert _parse_quarter_header("Q3-2024**") == date(2024, 9, 1)
        assert _parse_quarter_header("Q4-2023") == date(2023, 12, 1)

    def test_invalid(self):
        assert _parse_quarter_header("01.2024") is None
        assert _parse_quarter_header("foo") is None
        assert _parse_quarter_header("") is None
        assert _parse_quarter_header(None) is None

    def test_out_of_range(self):
        assert _parse_quarter_header("Q5-2024") is None
        assert _parse_quarter_header("Q0-2024") is None


class TestParseOfficialGdpQuarterGrid:
    def test_real_gdp_sheet(self):
        content = _make_official_quarter_grid_xlsx()
        result = parse_rosstat_gdp_quarter_grid_xlsx(content, "9")
        assert len(result) == 8
        assert result[0].date == date(2011, 3, 1)
        assert result[0].value == 26368.4
        assert result[3].date == date(2011, 12, 1)
        assert result[4].date == date(2012, 3, 1)

    def test_nominal_gdp_sheet_2(self):
        """Sheet '2' of VVP_kvartal_s_1995-2025.xlsx — nominal GDP в текущих ценах
        (ОКВЭД2, с 2011). Same row layout as sheet 9, configurable via gdp_sheet."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "2"
        ws.append(["К содержанию"])
        ws.append(["Валовой внутрений продукт1) (в текущих ценах, млрд.руб.)"])
        ws.append([2024, None, None, None, "20252)", None, None, None])
        ws.append(["I квартал", "II квартал", "III квартал", "IV квартал", "I квартал", "II квартал", "III квартал", "IV квартал"])
        ws.append([41410.7, 44389.2, 45606.9, 52079.4, 47950.4, 49284.8, 54671.8, 62354.1])

        buf = io.BytesIO()
        wb.save(buf)
        result = parse_rosstat_gdp_quarter_grid_xlsx(buf.getvalue(), "2")
        assert len(result) == 8
        assert result[7].date == date(2025, 12, 1)
        assert result[7].value == 62354.1
        assert result[4].date == date(2025, 3, 1)
        assert result[4].value == 47950.4


def _make_gdp_use_xls() -> bytes:
    """Build legacy .xls (OLE2) sample with ВВП-by-use multi-row layout.

    Mirrors `GDP-quarters-of-use-*.xls` sheet '2' (ОКВЭД2, 2011+):
    rows 4/7/8/11 = ВВП / consumption-HH / government / GFCF.
    """
    import xlwt

    wb = xlwt.Workbook()
    ws = wb.add_sheet("2")

    ws.write(0, 0, "К содержанию")
    ws.write(1, 0, "Элементы использования валового внутреннего продукта")
    for i, year in enumerate([2024, 2025]):
        ws.write(2, 1 + i * 4, year)
    quarters = ["I квартал", "II квартал", "III квартал", "IV квартал"]
    for offset in (0, 1):
        for j, q in enumerate(quarters):
            ws.write(3, 1 + offset * 4 + j, q)

    ws.write(4, 0, "Валовой внутренний продукт")
    for col, val in enumerate([41410.7, 44389.2, 45606.9, 52079.4, 47950.4, 49284.8, 54671.8, 62354.1]):
        ws.write(4, 1 + col, val)

    ws.write(7, 0, "домашних хозяйств")
    for col, val in enumerate([22803.0, 23891.7, 25694.3, 26660.6, 25265.0, 26174.9, 28038.6, 29281.1]):
        ws.write(7, 1 + col, val)

    ws.write(8, 0, "государственного управления")
    for col, val in enumerate([9101.9, 9119.5, 9249.3, 9592.6, 10114.7, 10034.8, 10327.7, 10448.3]):
        ws.write(8, 1 + col, val)

    ws.write(11, 0, "валовое накопление основного капитала")
    for col, val in enumerate([7397.1, 9843.4, 11056.0, 17091.4, 8530.3, 10599.8, 12165.1, 18176.8]):
        ws.write(11, 1 + col, val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseRosstatGdpUseXls:
    def test_consumption_row(self):
        try:
            content = _make_gdp_use_xls()
        except ImportError:
            pytest.skip("xlwt not installed (legacy .xls writer for tests only)")
        result = parse_rosstat_gdp_use_xls(content, sheet_name="2", value_row_index=7)
        assert len(result) == 8
        assert result[7].date == date(2025, 12, 1)
        assert result[7].value == 29281.1
        assert result[0].date == date(2024, 3, 1)
        assert result[0].value == 22803.0

    def test_government_row(self):
        try:
            content = _make_gdp_use_xls()
        except ImportError:
            pytest.skip("xlwt not installed")
        result = parse_rosstat_gdp_use_xls(content, sheet_name="2", value_row_index=8)
        assert len(result) == 8
        assert result[7].value == 10448.3

    def test_investment_row(self):
        try:
            content = _make_gdp_use_xls()
        except ImportError:
            pytest.skip("xlwt not installed")
        result = parse_rosstat_gdp_use_xls(content, sheet_name="2", value_row_index=11)
        assert len(result) == 8
        assert result[7].value == 18176.8
        assert result[3].value == 17091.4

    def test_dates_sorted(self):
        try:
            content = _make_gdp_use_xls()
        except ImportError:
            pytest.skip("xlwt not installed")
        result = parse_rosstat_gdp_use_xls(content, sheet_name="2", value_row_index=4)
        dates = [p.date for p in result]
        assert dates == sorted(dates)


class TestParseRuNumber:
    """Russian Excel ловушки: запятая, footnote-suffix, NBSP, str/int/float mix."""

    def test_int_float(self):
        assert _parse_ru_number(123) == 123.0
        assert _parse_ru_number(123.45) == 123.45

    def test_russian_decimal(self):
        assert _parse_ru_number("1662,8") == 1662.8

    def test_footnote_suffix_after_value(self):
        """Real Rosstat case: '1662,82)' = 1662.8 + footnote 2)."""
        assert _parse_ru_number("1662,82)") == 1662.8

    def test_footnote_strips_only_single_digit(self):
        """Rosstat-конвенция: footnote nums начинаются с 1 (single digit). Multi-digit
        footnotes не встречаются в rosstat, но даже если — strip только one digit."""
        assert _parse_ru_number("100,510)") == 100.51

    def test_nbsp_thousands(self):
        assert _parse_ru_number("1\u00a0234,5") == 1234.5
        assert _parse_ru_number("1 234,5") == 1234.5

    def test_empty_and_none(self):
        assert _parse_ru_number(None) is None
        assert _parse_ru_number("") is None
        assert _parse_ru_number("   ") is None

    def test_garbage_returns_none(self):
        assert _parse_ru_number("foo") is None
        assert _parse_ru_number("abc,def") is None

    def test_dot_decimal_kept(self):
        """Англоязычная запись тоже работает."""
        assert _parse_ru_number("1234.5") == 1234.5


class TestSpliceAtOverlap:
    """ratio-splice для GDP history extension до 1995 (ADR-0004)."""

    def test_basic_ratio_calibration(self):
        """ratio = mean(modern_2011) / mean(history_2011); historical scaled."""
        history = [
            DataPoint(date=date(2010, 12, 1), value=100.0),
            DataPoint(date=date(2011, 3, 1), value=200.0),
            DataPoint(date=date(2011, 6, 1), value=200.0),
        ]
        modern = [
            DataPoint(date=date(2011, 3, 1), value=220.0),
            DataPoint(date=date(2011, 6, 1), value=220.0),
            DataPoint(date=date(2012, 3, 1), value=240.0),
        ]
        result = splice_at_overlap(history, modern, overlap_year=2011)
        assert len(result) == 4
        assert result[0].date == date(2010, 12, 1)
        assert result[0].value == 110.0  # 100 * (220/200)
        assert result[1].date == date(2011, 3, 1)
        assert result[1].value == 220.0  # modern wins on overlap
        assert result[3].date == date(2012, 3, 1)
        assert result[3].value == 240.0

    def test_real_gdp_2011_quarters(self):
        """Real-world: nominal GDP 2011 sheet 1 vs sheet 2 → ratio ~1.074."""
        sheet1_2011 = [11954.2, 13376.4, 14732.9, 15903.7]
        sheet2_2011 = [13024.8, 14434.8, 15745.6, 16908.8]
        history = [DataPoint(date=date(2010, q * 3, 1), value=v) for q, v in enumerate([1000.0, 1100.0, 1200.0, 1300.0], start=1)]
        history += [DataPoint(date=date(2011, q * 3, 1), value=v) for q, v in enumerate(sheet1_2011, start=1)]
        modern = [DataPoint(date=date(2011, q * 3, 1), value=v) for q, v in enumerate(sheet2_2011, start=1)]
        result = splice_at_overlap(history, modern, overlap_year=2011)
        ratio = (sum(sheet2_2011) / 4) / (sum(sheet1_2011) / 4)
        assert abs(ratio - 1.074) < 0.001
        scaled_2010_q1 = next(p for p in result if p.date == date(2010, 3, 1))
        assert abs(scaled_2010_q1.value - 1000.0 * ratio) < 0.5

    def test_modern_priority_on_overlap(self):
        """Modern points в overlap_year полностью замещают history."""
        history = [
            DataPoint(date=date(2011, 3, 1), value=999.0),
            DataPoint(date=date(2011, 6, 1), value=999.0),
        ]
        modern = [
            DataPoint(date=date(2011, 3, 1), value=100.0),
            DataPoint(date=date(2011, 6, 1), value=200.0),
        ]
        result = splice_at_overlap(history, modern, overlap_year=2011)
        assert all(p.value in (100.0, 200.0) for p in result)
        assert 999.0 not in [p.value for p in result]

    def test_history_extension_only(self):
        """Если history дальше overlap-года — те точки не берутся (modern wins)."""
        history = [
            DataPoint(date=date(2010, 3, 1), value=100.0),
            DataPoint(date=date(2011, 3, 1), value=200.0),
            DataPoint(date=date(2012, 3, 1), value=300.0),  # после overlap, modern должен победить
        ]
        modern = [
            DataPoint(date=date(2011, 3, 1), value=220.0),
            DataPoint(date=date(2012, 3, 1), value=320.0),
        ]
        result = splice_at_overlap(history, modern, overlap_year=2011)
        assert len(result) == 3
        years = sorted({p.date.year for p in result})
        assert years == [2010, 2011, 2012]
        assert next(p for p in result if p.date == date(2012, 3, 1)).value == 320.0

    def test_raises_no_overlap_history(self):
        with pytest.raises(ValueError, match="history не содержит точек"):
            splice_at_overlap(
                [DataPoint(date=date(2010, 3, 1), value=100.0)],
                [DataPoint(date=date(2011, 3, 1), value=200.0)],
                overlap_year=2011,
            )

    def test_raises_no_overlap_modern(self):
        with pytest.raises(ValueError, match="modern не содержит точек"):
            splice_at_overlap(
                [DataPoint(date=date(2011, 3, 1), value=100.0)],
                [DataPoint(date=date(2012, 3, 1), value=200.0)],
                overlap_year=2011,
            )

    def test_raises_zero_history_mean(self):
        with pytest.raises(ValueError, match="history mean == 0"):
            splice_at_overlap(
                [DataPoint(date=date(2011, 3, 1), value=0.0)],
                [DataPoint(date=date(2011, 3, 1), value=100.0)],
                overlap_year=2011,
            )

    def test_result_sorted_by_date(self):
        history = [
            DataPoint(date=date(2010, 6, 1), value=100.0),
            DataPoint(date=date(2010, 3, 1), value=100.0),
            DataPoint(date=date(2011, 3, 1), value=100.0),
        ]
        modern = [
            DataPoint(date=date(2012, 3, 1), value=100.0),
            DataPoint(date=date(2011, 3, 1), value=100.0),
        ]
        result = splice_at_overlap(history, modern, overlap_year=2011)
        dates = [p.date for p in result]
        assert dates == sorted(dates)
