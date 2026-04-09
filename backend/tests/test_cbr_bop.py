"""Tests for CBR BOP (Balance of Payments) XLSX parser."""

import io
from datetime import date

import openpyxl
import pytest

from app.services.cbr_bop_parser import _parse_quarter_header, parse_bop_xlsx


def _make_bop_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Кварталы"

    for _ in range(6):
        ws.append([None])

    ws.cell(row=7, column=1, value="Счет текущих операций")
    ws.cell(row=7, column=2, value="1 квартал 2024 г.")
    ws.cell(row=7, column=3, value="2 квартал 2024 г.")

    ws.cell(row=7, column=2, value=1000.0)
    ws.cell(row=7, column=3, value=1200.0)

    ws.cell(row=6, column=2, value="1 квартал 2024 г.")
    ws.cell(row=6, column=3, value="2 квартал 2024 г.")

    ws.cell(row=13, column=1, value="Товары")
    ws.cell(row=13, column=2, value=5000.0)
    ws.cell(row=13, column=3, value=6000.0)

    ws.cell(row=14, column=1, value="Экспорт")
    ws.cell(row=14, column=2, value=8000.0)
    ws.cell(row=14, column=3, value=9000.0)

    ws.cell(row=15, column=1, value="Импорт")
    ws.cell(row=15, column=2, value=3000.0)
    ws.cell(row=15, column=3, value=3000.0)

    ws.cell(row=26, column=1, value="Услуги")
    ws.cell(row=26, column=2, value=500.0)
    ws.cell(row=26, column=3, value=600.0)

    ws.cell(row=27, column=1, value="Экспорт")
    ws.cell(row=27, column=2, value=2000.0)
    ws.cell(row=27, column=3, value=2200.0)

    ws.cell(row=28, column=1, value="Импорт")
    ws.cell(row=28, column=2, value=1500.0)
    ws.cell(row=28, column=3, value=1600.0)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestParseQuarterHeader:
    def test_valid(self):
        assert _parse_quarter_header("1 квартал 2024 г.") == date(2024, 3, 1)
        assert _parse_quarter_header("2 квартал 2024 г.") == date(2024, 6, 1)
        assert _parse_quarter_header("3 квартал 2023 г.") == date(2023, 9, 1)
        assert _parse_quarter_header("4 квартал 2023 г.") == date(2023, 12, 1)

    def test_invalid(self):
        assert _parse_quarter_header(None) is None
        assert _parse_quarter_header("") is None
        assert _parse_quarter_header("foo") is None
        assert _parse_quarter_header("Q1-2024") is None


class TestParseBopXlsx:
    def test_exports(self):
        content = _make_bop_xlsx()
        result = parse_bop_xlsx(content, "exports")
        assert len(result) == 2
        assert result[0].value == 8000.0
        assert result[1].value == 9000.0

    def test_imports(self):
        content = _make_bop_xlsx()
        result = parse_bop_xlsx(content, "imports")
        assert len(result) == 2
        assert result[0].value == 3000.0

    def test_trade_balance(self):
        content = _make_bop_xlsx()
        result = parse_bop_xlsx(content, "trade-balance")
        assert len(result) == 2
        assert result[0].value == 5000.0

    def test_services_exports(self):
        content = _make_bop_xlsx()
        result = parse_bop_xlsx(content, "services-exports")
        assert len(result) == 2
        assert result[0].value == 2000.0
        assert result[1].value == 2200.0

    def test_services_imports(self):
        content = _make_bop_xlsx()
        result = parse_bop_xlsx(content, "services-imports")
        assert len(result) == 2
        assert result[0].value == 1500.0

    def test_invalid_target(self):
        content = _make_bop_xlsx()
        with pytest.raises(ValueError, match="Unknown BOP target"):
            parse_bop_xlsx(content, "invalid")
