"""Серверная выгрузка таблиц (Excel/CSV) с гейтом по лимиту (ADR-0007 Phase 2).

Генерация файла перенесена с клиента на бэкенд: это (1) даёт жёсткий лимит
гостевых скачиваний, (2) снимает ~430 КБ xlsx из фронтового бандла. Трансформы
рядов (режимы графика) остаются на клиенте — сюда приходят уже готовые точки.

Числа в CSV — через ``display.format_number_ru`` (русская запятая). В шапке
файла — название, единица, частота, страна, источник и дата выгрузки.
"""
import io
import logging
from datetime import datetime
from urllib.parse import quote
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Request, Response
from pydantic import BaseModel, field_validator
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.database import get_db
from app.models import User
from app.security.auth import get_optional_user
from app.security import download_quota as dq
from app.services.display import format_number_ru, today_msk

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/export", tags=["export"])

_MAX_POINTS = 100_000
_MSK = ZoneInfo("Europe/Moscow")


class ExportPoint(BaseModel):
    date: str
    actual: float | None = None
    forecast: float | None = None


class ExportMeta(BaseModel):
    """Опциональные поля provenance для шапки файла."""

    indicator_name: str | None = None
    unit: str | None = None
    frequency: str | None = None
    country: str | None = None
    source: str | None = None
    source_url: str | None = None


class ExportIn(BaseModel):
    format: str
    filename: str
    value_label: str = "Значение"
    points: list[ExportPoint]
    # Плоские поля (удобно для клиента) + вложенный meta.
    indicator_name: str | None = None
    unit: str | None = None
    frequency: str | None = None
    country: str | None = None
    source: str | None = None
    source_url: str | None = None
    meta: ExportMeta | None = None

    @field_validator("format")
    @classmethod
    def _fmt(cls, v: str) -> str:
        v = (v or "").lower()
        if v not in ("xlsx", "csv"):
            raise ValueError("format must be xlsx or csv")
        return v

    @field_validator("points")
    @classmethod
    def _points(cls, v: list) -> list:
        if not v:
            raise ValueError("no points")
        if len(v) > _MAX_POINTS:
            raise ValueError("too many points")
        return v


def _round(x: float | None) -> float | None:
    return None if x is None else round(float(x), 4)


def _point_year(date_str: str) -> int | None:
    """Год из ISO-даты точки ('2024-01-01', '2024', '2024-W03') — для гейта глубины."""
    head = (date_str or "").strip()[:4]
    return int(head) if head.isdigit() else None


def _limit_history(points: list[ExportPoint]) -> list[ExportPoint]:
    """Гостю отдаём только последние N лет истории; полный период — за регистрацию.

    Отсчёт от самой поздней точки набора (а не от текущей даты): прогноз уходит
    в будущее, поэтому ориентир — максимальный год среди переданных точек.
    """
    years = settings.download_anon_history_years
    if years <= 0:
        return points
    yrs = [y for p in points if (y := _point_year(p.date)) is not None]
    if not yrs:
        return points
    cutoff = max(yrs) - years
    limited = [p for p in points if (y := _point_year(p.date)) is None or y > cutoff]
    return limited or points


def _split(points: list[ExportPoint]):
    facts = [(p.date, _round(p.actual)) for p in points if p.actual is not None]
    forecasts = [
        (p.date, _round(p.forecast))
        for p in points
        if p.forecast is not None and p.actual is None
    ]
    return facts, forecasts


def _resolve_meta(body: ExportIn | None = None, **kwargs) -> dict[str, str]:
    """Собрать поля шапки из body / kwargs."""
    m = (body.meta if body is not None else None) or ExportMeta()
    get = lambda k: (
        kwargs.get(k)
        or (getattr(body, k, None) if body is not None else None)
        or getattr(m, k, None)
        or ""
    )
    exported_at = datetime.now(_MSK).strftime("%Y-%m-%d %H:%M %Z")
    return {
        "indicator_name": str(get("indicator_name") or "").strip(),
        "unit": str(get("unit") or "").strip(),
        "frequency": str(get("frequency") or "").strip(),
        "country": str(get("country") or "").strip(),
        "source": str(get("source") or "").strip(),
        "source_url": str(get("source_url") or "").strip(),
        "exported_at": exported_at,
        "exported_date": today_msk().isoformat(),
    }


def _meta_rows(meta: dict[str, str], value_label: str) -> list[tuple[str, str]]:
    """Пары (поле, значение) для шапки файла."""
    name = meta.get("indicator_name") or value_label or "Значение"
    rows: list[tuple[str, str]] = [
        ("Показатель", name),
    ]
    if meta.get("unit"):
        rows.append(("Единица", meta["unit"]))
    if meta.get("frequency"):
        rows.append(("Частота", meta["frequency"]))
    if meta.get("country"):
        rows.append(("Страна", meta["country"]))
    source = meta.get("source") or "Евростат"
    rows.append(("Источник", source))
    if meta.get("source_url"):
        rows.append(("URL источника", meta["source_url"]))
    rows.append(("Дата выгрузки", meta.get("exported_at") or meta.get("exported_date") or ""))
    return rows


def _format_csv_value(val: float | None) -> str:
    if val is None:
        return ""
    return format_number_ru(val)


def _build_xlsx(
    facts,
    forecasts,
    value_label: str,
    meta: dict[str, str] | None = None,
) -> bytes:
    from openpyxl import Workbook

    meta = meta or _resolve_meta()
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Описание"
    for label, value in _meta_rows(meta, value_label):
        ws_meta.append([label, value])

    ws = wb.create_sheet("Факт", 1)
    ws.append(["Дата", value_label])
    for date, val in facts:
        # Excel хранит число нативно; подпись единицы — в «Описание».
        ws.append([date, val])
    if forecasts:
        ws2 = wb.create_sheet("Прогноз")
        ws2.append(["Дата", f"Прогноз {value_label}"])
        for date, val in forecasts:
            ws2.append([date, val])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_csv(
    facts,
    forecasts,
    value_label: str,
    meta: dict[str, str] | None = None,
) -> bytes:
    meta = meta or _resolve_meta()
    lines: list[str] = []
    for label, value in _meta_rows(meta, value_label):
        # экранируем ';' в значениях
        safe = str(value).replace(";", ",")
        lines.append(f"# {label};{safe}")
    lines.append(";".join(["Дата", value_label, "Тип"]))
    for date, val in facts:
        lines.append(";".join([date, _format_csv_value(val), "факт"]))
    for date, val in forecasts:
        lines.append(";".join([date, _format_csv_value(val), "прогноз"]))
    return ("\ufeff" + "\n".join(lines)).encode("utf-8")


def _content_disposition(filename: str) -> str:
    ascii_fallback = "export." + (filename.rsplit(".", 1)[-1] if "." in filename else "dat")
    return f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{quote(filename)}"


@router.get("/quota")
async def export_quota(
    request: Request,
    user: User | None = Depends(get_optional_user),
):
    """Сколько гостевых выгрузок осталось — для состояния кнопок UI (без инкремента).

    Авторизованный пользователь: unlimited=True. Гость: remaining из счётчика fe_dl.
    """
    if user is not None:
        return {"unlimited": True, "remaining": None, "limit": settings.download_anon_limit,
                "history_years": 0}
    dl_id = request.cookies.get(dq.DL_COOKIE)
    remaining = await dq.remaining_anon_downloads(dl_id)
    return {"unlimited": False, "remaining": remaining, "limit": settings.download_anon_limit,
            "history_years": settings.download_anon_history_years}


@router.post("/table")
async def export_table(
    body: ExportIn,
    request: Request,
    user: User | None = Depends(get_optional_user),
    db: AsyncSession = Depends(get_db),
):
    set_cookie_id: str | None = None
    if user is None:
        dl_id = request.cookies.get(dq.DL_COOKIE)
        if not dl_id:
            dl_id = dq.new_download_id()
            set_cookie_id = dl_id
        allowed = await dq.consume_anon_download(dl_id)
        if not allowed:
            raise HTTPException(
                status_code=403,
                detail={
                    "code": "download_limit",
                    "message": "Лимит бесплатных выгрузок исчерпан. Войдите в аккаунт для безлимитного скачивания.",
                },
            )

    # Полный период истории — бонус за регистрацию: гостю обрезаем глубину.
    points = body.points if user is not None else _limit_history(body.points)
    facts, forecasts = _split(points)
    meta = _resolve_meta(body)
    # Если клиент не передал имя — вытащим из value_label.
    if not meta.get("indicator_name") and body.value_label:
        meta["indicator_name"] = body.value_label
    if body.format == "xlsx":
        data = _build_xlsx(facts, forecasts, body.value_label, meta)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        data = _build_csv(facts, forecasts, body.value_label, meta)
        media = "text/csv; charset=utf-8"

    resp = Response(content=data, media_type=media)
    resp.headers["Content-Disposition"] = _content_disposition(body.filename)
    if user is None:
        remaining = await dq.remaining_anon_downloads(set_cookie_id or request.cookies.get(dq.DL_COOKIE))
        resp.headers["X-Download-Remaining"] = str(remaining)
        resp.headers["Access-Control-Expose-Headers"] = "X-Download-Remaining"
    if set_cookie_id is not None:
        kw = {"httponly": True, "secure": settings.auth_cookie_secure, "samesite": "lax", "path": "/"}
        if settings.auth_cookie_domain:
            kw["domain"] = settings.auth_cookie_domain
        resp.set_cookie(dq.DL_COOKIE, set_cookie_id, max_age=settings.download_anon_window_seconds, **kw)
    return resp
