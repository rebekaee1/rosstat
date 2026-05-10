"""ETL: Росстат population XLSX → IndicatorData (canonical русский Rosstat, без SDDS).

Три файла из rosstat.gov.ru/folder/12781:

1. `Popul_1897+.xlsx` — историческая годовая численность 1897-2024, млн человек.
   Sheet "Лист1", rows 7+. Для 1897 и 1914 — две строки (имперские vs совр. границы),
   берём «в современных границах».

2. `Popul components_1990+.xlsx` — годовые компоненты прироста с 1990:
   население на 1 января (тыс.), общий/естественный/миграционный прирост (тыс.).
   Sheet "1", rows 8+. Дает 4 индикатора: population, total/natural/migration growth.

3. `OkPopul_Comp{YYYY}_Site.xlsx` — ежегодная «оценка численности постоянного
   населения на 1 января», sheet "Всего" row РФ. Один новый point per year =
   население на 1 января текущего года (после обновления файла rosstat'ом).
"""

from __future__ import annotations

import asyncio
import io
import logging
import re
from dataclasses import dataclass
from datetime import date
from typing import ClassVar

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import FetchLog, Indicator
from app.services.base_parser import BaseParser
from app.services.data_validator import validate_points
from app.services.rosstat_sdds_fetcher import (
    fetch_rosstat_okpopul,
    fetch_rosstat_static_xlsx,
)

logger = logging.getLogger(__name__)


@dataclass
class DataPoint:
    date: date
    value: float


_YEAR_RE = re.compile(r"(\d{4})")


def _extract_year(cell) -> int | None:
    """Extract 4-digit year from a cell that may contain footnote markers like '20222),3)'."""
    if cell is None:
        return None
    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
        y = int(cell)
        return y if 1800 <= y <= 2100 else None
    m = _YEAR_RE.match(str(cell).strip())
    if not m:
        return None
    y = int(m.group(1))
    return y if 1800 <= y <= 2100 else None


def parse_ok_popul_xlsx(content: bytes) -> DataPoint:
    """Parse rosstat русский `OkPopul_Comp{YYYY}_Site.xlsx` → одна точка
    «население РФ на 1 января {YYYY}» в млн человек.

    Sheet "Всего": ищем row начинающийся с «Российская Федерация», в этой строке
    ищем колонку с числом 100M-200M (тех.значение РФ population) у которой выше
    в шапке стоит "{YYYY} г.". Берём наибольший year (новый файл публикует значения
    на 1 января нового года в правой части и предыдущего в левой).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        if "Всего" not in wb.sheetnames:
            raise ValueError("OkPopul: sheet 'Всего' not found")
        ws = wb["Всего"]
        rows_data = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    rf_row_idx: int | None = None
    for r_idx, row in enumerate(rows_data):
        cell = str((row[0] if row else "") or "").strip().lower()
        # Rosstat файлы регулярно мешают кириллицу с латиницей в "Федеpация"
        # (латинская "p"); проверяем только устойчивый префикс.
        if cell.startswith("российская"):
            rf_row_idx = r_idx
            break
    if rf_row_idx is None:
        raise ValueError("OkPopul: 'Российская …' row not found")

    candidates: list[tuple[date, float]] = []
    rf_data = rows_data[rf_row_idx]
    for c_idx, val in enumerate(rf_data):
        if isinstance(val, bool) or not isinstance(val, (int, float)):
            continue
        v = float(val)
        if not (100_000_000 < v < 200_000_000):
            continue
        for hr in range(max(0, rf_row_idx - 5), rf_row_idx):
            if c_idx >= len(rows_data[hr]):
                continue
            cell_text = str(rows_data[hr][c_idx] or "")
            m = re.search(r"(\d{4})\s*г\.", cell_text)
            if m:
                year = int(m.group(1))
                if 2020 <= year <= 2100:
                    candidates.append((date(year, 1, 1), round(v / 1_000_000, 2)))
                    break

    if not candidates:
        raise ValueError("OkPopul: no valid (year, pop) candidate found in РФ row")

    latest = max(candidates, key=lambda x: x[0])
    return DataPoint(date=latest[0], value=latest[1])


def parse_population_history_xlsx(content: bytes) -> list[DataPoint]:
    """Parse Popul_1897+.xlsx → annual population in millions.

    For 1897 and 1914 Rosstat gives two rows; use the row in modern borders,
    because the rest of the product is Russia in current/modern boundaries.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb.worksheets[0]
        rows_data: list[list] = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append(list(row))
    finally:
        wb.close()

    points: list[DataPoint] = []
    pending_year: int | None = None
    for row in rows_data[6:]:
        if not row:
            continue

        first = row[0]
        year = _extract_year(first)
        if year is not None:
            if len(row) > 1 and row[1] is not None:
                try:
                    points.append(DataPoint(date=date(year, 1, 1), value=round(float(row[1]), 2)))
                except (ValueError, TypeError):
                    pass
                pending_year = None
            else:
                pending_year = year
            continue

        label = str(first or "").lower().replace("\xa0", " ").strip()
        if pending_year and "современных границах" in label and len(row) > 1 and row[1] is not None:
            try:
                points.append(DataPoint(date=date(pending_year, 1, 1), value=round(float(row[1]), 2)))
            except (ValueError, TypeError):
                pass
            pending_year = None

    return sorted(points, key=lambda p: p.date)


def merge_population_sources(*sources: list[DataPoint]) -> list[DataPoint]:
    """Merge ordered population series. Later args win on date overlap.

    Use case: merge_population_sources(history, components, [latest_okpopul]) — components
    overrides history on 1990+ overlap, OkPopul overrides components on the most-recent
    январь.
    """
    by_date: dict[date, DataPoint] = {}
    for src in sources:
        by_date.update({p.date: p for p in src})
    return [by_date[d] for d in sorted(by_date)]


def parse_popul_components_xlsx(content: bytes) -> dict[str, list[DataPoint]]:
    """Parse Popul components_1990+.xlsx → population, natural growth, total growth, migration.

    Sheet "1", rows starting from row 8:
    Col A: Year (int)
    Col B: Population Jan 1 (thousands)
    Col C: Total change (thousands)
    Col D: Natural change (thousands)
    Col E: Migration change (thousands)
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        sheets = wb.sheetnames
        ws = wb["1"] if "1" in sheets else wb.worksheets[1]
        rows_data: list[list] = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append(list(row))
    finally:
        wb.close()

    population: list[DataPoint] = []
    total_growth: list[DataPoint] = []
    natural_growth: list[DataPoint] = []
    migration: list[DataPoint] = []

    for row in rows_data[7:]:
        if not row or row[0] is None:
            continue
        year = _extract_year(row[0])
        if year is None:
            continue

        d = date(year, 1, 1)

        if row[1] is not None:
            try:
                pop_thousands = float(row[1])
                population.append(DataPoint(date=d, value=round(pop_thousands / 1000, 2)))
            except (ValueError, TypeError):
                pass

        if row[2] is not None:
            try:
                total_growth.append(DataPoint(date=d, value=round(float(row[2]), 1)))
            except (ValueError, TypeError):
                pass

        if row[3] is not None:
            try:
                natural_growth.append(DataPoint(date=d, value=round(float(row[3]), 1)))
            except (ValueError, TypeError):
                pass

        if row[4] is not None:
            try:
                migration.append(DataPoint(date=d, value=round(float(row[4]), 1)))
            except (ValueError, TypeError):
                pass

    return {
        "population": sorted(population, key=lambda p: p.date),
        "total-growth": sorted(total_growth, key=lambda p: p.date),
        "natural-growth": sorted(natural_growth, key=lambda p: p.date),
        "migration": sorted(migration, key=lambda p: p.date),
    }


INDICATOR_SOURCE_MAP: dict[str, str] = {
    "population": "historical",
    "population-total-growth": "components",
    "population-natural-growth": "components",
    "population-migration": "components",
}

COMPONENT_SERIES_MAP: dict[str, str] = {
    "population-total-growth": "total-growth",
    "population-natural-growth": "natural-growth",
    "population-migration": "migration",
}


class RosstatPopulationParser(BaseParser):
    parser_type: ClassVar[str] = "rosstat_population"

    async def _fetch_and_parse(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
    ) -> tuple[list, str]:
        code = indicator.code
        source = INDICATOR_SOURCE_MAP.get(code, "components")

        if source == "historical":
            history_content, history_url = await asyncio.to_thread(
                fetch_rosstat_static_xlsx, "population_history"
            )
            components_content, components_url = await asyncio.to_thread(
                fetch_rosstat_static_xlsx, "popul_components"
            )
            history_points = await asyncio.to_thread(parse_population_history_xlsx, history_content)
            components_series = await asyncio.to_thread(parse_popul_components_xlsx, components_content)
            components_pop = components_series["population"]

            latest_points: list[DataPoint] = []
            sources_url = f"{history_url}; {components_url}"
            try:
                ok_content, ok_url = await asyncio.to_thread(fetch_rosstat_okpopul)
                latest_points = [await asyncio.to_thread(parse_ok_popul_xlsx, ok_content)]
                sources_url = f"{history_url}; {components_url}; {ok_url}"
            except Exception as e:
                logger.warning(
                    "OkPopul yearly fetch failed (latest year may lag until next OkPopul publish): %s",
                    e,
                )

            points = merge_population_sources(history_points, components_pop, latest_points)
            return points, sources_url

        content, final_url = await asyncio.to_thread(
            fetch_rosstat_static_xlsx, "popul_components"
        )
        all_series = await asyncio.to_thread(parse_popul_components_xlsx, content)
        series_key = COMPONENT_SERIES_MAP.get(code, "population")
        return all_series.get(series_key, []), final_url

    def _validate(self, points: list, cfg: dict) -> list:
        return validate_points(points, cfg)
