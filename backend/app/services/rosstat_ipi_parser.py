"""ETL: Росстат Industrial Production Index → IndicatorData (canonical русский Rosstat).

Два файла из rosstat.gov.ru/enterprise_industrial:

1. `ind_baza_2018_12-2025.xlsx` — финальная публикация база 2018, монtnly MoM%,
   2015-Dec 2025. Sheet "1" row 6 (BCDE = ПРОМЫШЛЕННОЕ ПРОИЗВОДСТВО).

2. `ind_baza_2023_{MM}-{YYYY}.xlsx` — текущая публикация база 2023, переиздаётся
   ежемесячно. Sheet "1" row 6, MoM% для 2026+.

Стратегия (path P / compat ADR-0004): rosstat публикует MoM% индексы, наша DB
исторически хранит cumulative monthly index с anchor `2023 average = 100`.
Парсер делает chain MoM → cumulative → нормализует так, что среднее за 2023 г. = 100.
Результат идентичен формату хранения SDDS (verified bit-for-bit on local stack).
Frontend value formatter, chart unit, derived `ipi-yoy` и forecast model не трогаются.
"""

from __future__ import annotations

import asyncio
import io
import logging
import re
from dataclasses import dataclass
from datetime import date
from typing import ClassVar

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import FetchLog, Indicator
from app.services.base_parser import BaseParser
from app.services.data_validator import validate_points
from app.services.rosstat_sdds_fetcher import (
    fetch_rosstat_ipi_current,
    fetch_rosstat_static_xlsx,
)

logger = logging.getLogger(__name__)


@dataclass
class DataPoint:
    date: date
    value: float


_MONTH_NAME_TO_NUM = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4, "май": 5, "июнь": 6,
    "июль": 7, "август": 8, "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}

_YEAR_RE = re.compile(r"(\d{4})")
_MONTH_PREFIX_RE = re.compile(r"([а-я]+)")


def parse_rosstat_ipi_mom_xlsx(content: bytes) -> dict[date, float]:
    """Parse rosstat ind_baza_*.xlsx sheet "1" → {date: MoM_percent} for ПРОМЫШЛЕННОЕ
    ПРОИЗВОДСТВО (BCDE) row.

    Layout:
      row 4: year headers (sparse — only at first month of each year)
      row 5: month names (январь/февраль/...) — могут иметь footnote markers (январь1)
      row 6 (or detected dynamically): BCDE row with monthly MoM% values starting col 3
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        if "1" not in wb.sheetnames:
            raise ValueError("Rosstat IPI: sheet '1' not found")
        ws = wb["1"]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    bcde_row_idx: int | None = None
    for r_idx, row in enumerate(rows):
        if len(row) > 1 and row[1] and "BCDE" in str(row[1]):
            bcde_row_idx = r_idx
            break
    if bcde_row_idx is None:
        raise ValueError("Rosstat IPI: BCDE (ПРОМЫШЛЕННОЕ ПРОИЗВОДСТВО) row not found")

    if len(rows) < 5:
        raise ValueError(f"Rosstat IPI: too few rows ({len(rows)})")

    year_row = rows[3]
    month_row = rows[4]
    data_row = rows[bcde_row_idx]

    result: dict[date, float] = {}
    current_year: int | None = None

    for c_idx in range(2, len(month_row)):
        if c_idx < len(year_row) and year_row[c_idx] is not None:
            m = _YEAR_RE.search(str(year_row[c_idx]))
            if m:
                year = int(m.group(1))
                if 1990 <= year <= 2100:
                    current_year = year

        if current_year is None:
            continue

        month_name = str(month_row[c_idx] or "").strip().lower()
        m = _MONTH_PREFIX_RE.match(month_name)
        if not m:
            continue
        month_num = _MONTH_NAME_TO_NUM.get(m.group(1))
        if month_num is None:
            continue

        val = data_row[c_idx] if c_idx < len(data_row) else None
        if val is None:
            continue
        try:
            result[date(current_year, month_num, 1)] = float(val)
        except (ValueError, TypeError):
            pass

    return result


def chain_mom_to_index_2023_base(mom_data: dict[date, float]) -> list[DataPoint]:
    """Chain MoM% to cumulative monthly index, нормализованный так что среднее
    значение за 2023 г. = 100.

    Формула:
      index[t] = index[t-1] * MoM[t] / 100  (стартовое значение arbitrary 100,
                                              нормализуется ниже)
      scale = 100 / mean(index[2023-01..2023-12])
      output[t] = index[t] * scale

    Если в данных нет 2023 г. (требуется для anchor) — raises ValueError.
    """
    if not mom_data:
        return []

    sorted_dates = sorted(mom_data.keys())
    chained: dict[date, float] = {}
    prev = 100.0
    for d in sorted_dates:
        cur = prev * mom_data[d] / 100.0
        chained[d] = cur
        prev = cur

    vals_2023 = [v for d, v in chained.items() if d.year == 2023]
    if not vals_2023:
        raise ValueError(
            "Rosstat IPI chain: cannot anchor — нет данных за 2023 г. "
            "(нужно для нормализации `2023 avg = 100`)"
        )
    avg_2023 = sum(vals_2023) / len(vals_2023)
    scale = 100.0 / avg_2023

    return [DataPoint(date=d, value=round(chained[d] * scale, 1)) for d in sorted_dates]


def merge_mom_dicts(*dicts: dict[date, float]) -> dict[date, float]:
    """Merge MoM dicts: later args override earlier on date overlap."""
    merged: dict[date, float] = {}
    for d in dicts:
        merged.update(d)
    return merged


class RosstatIpiParser(BaseParser):
    parser_type: ClassVar[str] = "rosstat_sdds_ipi"

    async def _fetch_and_parse(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
    ) -> tuple[list, str]:
        hist_content, hist_url = await asyncio.to_thread(
            fetch_rosstat_static_xlsx, "ipi_historical_2018"
        )
        hist_mom = await asyncio.to_thread(parse_rosstat_ipi_mom_xlsx, hist_content)

        sources_url = hist_url
        cur_mom: dict[date, float] = {}
        try:
            cur_content, cur_url = await asyncio.to_thread(fetch_rosstat_ipi_current)
            cur_mom = await asyncio.to_thread(parse_rosstat_ipi_mom_xlsx, cur_content)
            sources_url = f"{hist_url}; {cur_url}"
        except Exception as e:
            logger.warning(
                "Rosstat IPI current ind_baza_2023_*.xlsx fetch failed (свежие месяцы могут "
                "отсутствовать пока rosstat не опубликует новый файл): %s",
                e,
            )

        merged_mom = merge_mom_dicts(hist_mom, cur_mom)
        points = chain_mom_to_index_2023_base(merged_mom)
        return points, sources_url

    def _validate(self, points: list, cfg: dict) -> list:
        return validate_points(points, cfg)
