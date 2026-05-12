"""ETL: ЦБ РФ — внешняя торговля **услугами** monthly → IndicatorData.

Источник: `https://www.cbr.ru/vfs/statistics/credit_statistics/trade/trade_monthly.xlsx`
Лист «месяцы » (с trailing-пробелом в имени!) содержит месячные значения
экспорта/импорта услуг с **2018-01** до текущего месяца.

Layout — **transposed**:
- row 4 (1-based): header row с датами в col 2..N. Datetime objects или строки
  типа «янв.26\\n(оценка)» для предварительных оценок последних 1-2 месяцев.
- col 1: label row (например «Услуги », «Экспорт услуг», «Импорт услуг»,
  «Транспортные услуги», и т.д.)

Target rows:
- `services-exports-monthly` → row label «Экспорт услуг»
- `services-imports-monthly` → row label «Импорт услуг»
- (`services-balance-monthly` row «Услуги » пока не нужен — see seed_data.py
  T3 scope: minimal_no_derived).
"""

from __future__ import annotations

import asyncio
import io
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import ClassVar

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import FetchLog, Indicator
from app.services.base_parser import BaseParser
from app.services.http_client import create_session

logger = logging.getLogger(__name__)

TRADE_SERVICES_URL = (
    "https://www.cbr.ru/vfs/statistics/credit_statistics/trade/trade_monthly.xlsx"
)

_MONTH_RU_SHORT = {
    "янв": 1, "фев": 2, "мар": 3, "апр": 4, "май": 5, "июн": 6,
    "июл": 7, "авг": 8, "сен": 9, "окт": 10, "ноя": 11, "дек": 12,
}

_TARGET_LABEL = {
    "services-exports-monthly": "экспорт услуг",
    "services-imports-monthly": "импорт услуг",
}

_HEADER_STR_RE = re.compile(
    r"([а-яё]{3,4})\.?\s*(\d{2,4})", re.IGNORECASE,
)


@dataclass
class DataPoint:
    date: date
    value: float


def _parse_header_date(cell: object) -> date | None:
    """Convert header cell to first-of-month date.

    Accepts:
    - datetime (most common): 2018-01-01 00:00:00 → date(2018, 1, 1)
    - string like 'янв.26\\n(оценка)' or 'фев.26' → date(2026, 2, 1)
    """
    if cell is None:
        return None
    if isinstance(cell, datetime):
        return date(cell.year, cell.month, 1)
    if not isinstance(cell, str):
        return None
    s = cell.strip().lower()
    m = _HEADER_STR_RE.search(s)
    if not m:
        return None
    month_key = m.group(1)[:3]
    if month_key not in _MONTH_RU_SHORT:
        return None
    month = _MONTH_RU_SHORT[month_key]
    yr_raw = m.group(2)
    year = int(yr_raw)
    if year < 100:
        year = 2000 + year
    if year < 1990 or year > 2100:
        return None
    return date(year, month, 1)


def fetch_trade_services_xlsx() -> tuple[bytes, str]:
    session = create_session()
    try:
        resp = session.get(TRADE_SERVICES_URL, timeout=90)
        resp.raise_for_status()
        if resp.content[:4] != b"PK\x03\x04":
            raise ValueError("trade_monthly.xlsx: not a valid XLSX archive")
        logger.info(
            "Downloaded trade_monthly.xlsx: %d KB", len(resp.content) // 1024,
        )
        return resp.content, TRADE_SERVICES_URL
    finally:
        session.close()


def parse_trade_services_monthly_xlsx(
    content: bytes, target: str,
) -> list[DataPoint]:
    """Parse trade_monthly.xlsx лист «месяцы» для одного target.

    target: один из 'services-exports-monthly', 'services-imports-monthly'.
    """
    if target not in _TARGET_LABEL:
        raise ValueError(f"Unknown services trade target: {target}")
    label_needle = _TARGET_LABEL[target]

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = None
    for sn in wb.sheetnames:
        if "месяц" in sn.lower():
            ws = wb[sn]
            break
    if ws is None:
        raise ValueError("trade_monthly.xlsx: лист 'месяцы' не найден")

    header_row_idx = None
    dates: list[tuple[int, date]] = []
    for ri in range(1, min(10, ws.max_row + 1)):
        candidate_dates: list[tuple[int, date]] = []
        for ci in range(2, ws.max_column + 1):
            d = _parse_header_date(ws.cell(row=ri, column=ci).value)
            if d:
                candidate_dates.append((ci, d))
        if len(candidate_dates) >= 3:
            header_row_idx = ri
            dates = candidate_dates
            break

    if header_row_idx is None or not dates:
        raise ValueError(
            "trade_monthly.xlsx: header row с датами не найден в первых 10 строках",
        )

    data_row_idx = None
    for ri in range(header_row_idx + 1, ws.max_row + 1):
        label = ws.cell(row=ri, column=1).value
        if not label:
            continue
        label_norm = str(label).strip().lower()
        if label_norm == label_needle:
            data_row_idx = ri
            break

    if data_row_idx is None:
        raise ValueError(
            f"trade_monthly.xlsx: row для '{label_needle}' не найдена",
        )

    points: list[DataPoint] = []
    for ci, d in dates:
        raw = ws.cell(row=data_row_idx, column=ci).value
        if raw is None or raw == "":
            continue
        try:
            val = float(raw)
        except (TypeError, ValueError):
            continue
        points.append(DataPoint(date=d, value=round(val, 2)))

    points.sort(key=lambda p: p.date)
    logger.info(
        "trade_monthly.xlsx '%s' (row=%d, label='%s'): %d points (%s → %s)",
        target,
        data_row_idx,
        label_needle,
        len(points),
        points[0].date if points else "?",
        points[-1].date if points else "?",
    )
    return points


class CbrTradeServicesMonthlyParser(BaseParser):
    """ETL parser для месячной внешней торговли услугами (ЦБ trade_monthly.xlsx)."""

    parser_type: ClassVar[str] = "cbr_trade_services_monthly"

    async def _fetch_and_parse(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
    ) -> tuple[list, str]:
        target = cfg.get("bop_target", indicator.code)
        content, final_url = await asyncio.to_thread(fetch_trade_services_xlsx)
        points = await asyncio.to_thread(
            parse_trade_services_monthly_xlsx, content, target,
        )
        return points, final_url
