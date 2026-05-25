"""ETL: CBR Monetary Aggregates XLSX parser.

Парсит `https://www.cbr.ru/vfs/statistics/credit_statistics/monetary_agg.xlsx` —
официальный XLSX от ЦБ «Денежные агрегаты», публикуется ~25 числа каждого
месяца с **апрельскими** данными на 1-е мая (т.е. в XLSX дата 2026-05-01 =
состояние конца апреля 2026).

Почему отдельный парсер, а не DataService:
- CBR DataService (publicationId=5/datasetId=7) обновляется с лагом ~2-4
  недели **после** XLSX. На 25 мая 2026 DataService отдавал последнюю
  точку 2026-04-01 (=март), а XLSX уже содержал 2026-05-01 (=апрель). См.
  правка 2026-05-25 (Никита: «теперь стало за март, но апреля все ещё нет»).
- XLSX покрывает 5 наших индикаторов одной выгрузкой: m0, m1, m2,
  deposits-individual, deposits-business. Меньше HTTP-запросов, меньше
  парсеров, меньше точек отказа.
- consumer-credit/business-credit (publicationId=20/22) остаются на
  DataService — другие источники, не агрегированы в этом XLSX.

Семантика дат: ЦБ записывает «остаток на 1-е число» = состояние конца
**предыдущего** месяца → применяем `date_offset_months = -1` (как для
DataService monetary, см. правка 2026-05-25). 2026-05-01 → 2026-04-01.

Структура XLSX (sheet «Денежные агрегаты»):
  row 1   — даты (datetime), начиная с column index 1
  row 2   — Денежный агрегат М0
  row 5   — Переводные депозиты нефинансовых организаций
  row 6   — Переводные депозиты домашних хозяйств
  row 9   — Денежный агрегат М1
  row 12  — Другие депозиты нефинансовых организаций
  row 13  — Другие депозиты домашних хозяйств
  row 14  — Денежный агрегат М2
  row 17  — Депозиты в иностранной валюте нефинансовых организаций
  row 18  — Депозиты в иностранной валюте домашних хозяйств
  row 20  — Денежный агрегат M2X (широкая денежная масса)

Конфиг в indicator.model_config_json:
{
  "monetary_agg": {
    "indicator": "M2"  # или: M0, M1, M2X, deposits-individual, deposits-business
  },
  "backfill_from_year": 2010   # optional
}
"""

from __future__ import annotations

import asyncio
import io
import logging
from datetime import date, datetime
from typing import ClassVar

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import FetchLog, Indicator
from app.services.base_parser import BaseParser
from app.services.http_client import create_session

logger = logging.getLogger(__name__)

XLSX_URL = "https://www.cbr.ru/vfs/statistics/credit_statistics/monetary_agg.xlsx"

# Маппинг indicator → список row-индексов (1-based в openpyxl), значения которых
# складываются. Для агрегатов M0/M1/M2/M2X — одна строка; для deposits-* —
# сумма трёх (переводные + другие в рублях + в иностранной валюте).
ROW_MAP: dict[str, list[int]] = {
    "M0": [2],
    "M1": [9],
    "M2": [14],
    "M2X": [20],
    "deposits-business":   [5, 12, 17],
    "deposits-individual": [6, 13, 18],
}


def _shift_month(d: date, months: int) -> date:
    """Сдвиг даты на N месяцев, всегда 1-е число."""
    y, m = d.year, d.month + months
    while m <= 0:
        m += 12
        y -= 1
    while m > 12:
        m -= 12
        y += 1
    return date(y, m, 1)


def fetch_monetary_agg(
    indicator_key: str,
    year_from: int = 1993,
    date_offset_months: int = -1,
) -> list[tuple[date, float]]:
    """Скачать XLSX и извлечь ряд для одного indicator_key."""
    rows = ROW_MAP.get(indicator_key)
    if not rows:
        raise ValueError(f"Unknown monetary_agg indicator {indicator_key!r}; available: {sorted(ROW_MAP)}")

    session = create_session()
    try:
        resp = session.get(XLSX_URL, timeout=60)
        resp.raise_for_status()
        xlsx_bytes = resp.content
    finally:
        session.close()

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    if "Денежные агрегаты" not in wb.sheetnames:
        raise ValueError(f"Sheet «Денежные агрегаты» not found in monetary_agg.xlsx; available: {wb.sheetnames}")
    ws = wb["Денежные агрегаты"]

    cells = list(ws.iter_rows(values_only=True))
    if not cells:
        return []

    header = cells[0]
    # Cборка по столбцам: column index i → date header[i], значение = сумма по rows.
    points: dict[date, float] = {}
    for col_i in range(1, len(header)):
        dt_raw = header[col_i]
        if not isinstance(dt_raw, datetime):
            continue
        raw_date = dt_raw.date().replace(day=1)
        if raw_date.year < year_from:
            continue
        total = 0.0
        any_present = False
        for r in rows:
            if r - 1 >= len(cells):
                continue
            val = cells[r - 1][col_i] if col_i < len(cells[r - 1]) else None
            if val is None:
                continue
            try:
                total += float(val)
                any_present = True
            except (TypeError, ValueError):
                continue
        if not any_present:
            continue
        shifted = _shift_month(raw_date, date_offset_months)
        points[shifted] = round(total, 4)

    return sorted(points.items())


class CbrMonetaryAggParser(BaseParser):
    parser_type: ClassVar[str] = "cbr_monetary_agg_xlsx"

    async def _fetch_and_parse(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
    ) -> tuple[list, str]:
        ma_cfg = cfg.get("monetary_agg")
        if not ma_cfg:
            raise ValueError("Missing 'monetary_agg' in model_config_json")
        indicator_key = ma_cfg.get("indicator")
        if indicator_key not in ROW_MAP:
            raise ValueError(
                f"Unknown 'monetary_agg.indicator' {indicator_key!r}; "
                f"available: {sorted(ROW_MAP)}"
            )

        year_from = int(cfg.get("backfill_from_year", 1993))
        date_offset = int(ma_cfg.get("date_offset_months", -1))

        points = await asyncio.to_thread(
            fetch_monetary_agg, indicator_key, year_from, date_offset,
        )
        return points, XLSX_URL
