"""ETL: Росстат SDDS Price Indices XLSX → IndicatorData (PPI).

Файл: SDDS_price indices_{year}.xlsx
Лист: "Price indices_2010"
Строки:
  Row 1: заголовки дат (MM.YYYY)
  Row 2: Consumer Price Index (2010=100)
  Row 3: Producer Price Index (2010=100)
"""

from __future__ import annotations

import asyncio
import io
import logging
from dataclasses import dataclass
from datetime import date, datetime, timezone
from typing import ClassVar

import openpyxl
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import FetchLog, Indicator, IndicatorData
from app.services.base_parser import BaseParser
from app.services.upsert import upsert_indicator_data
from app.services.data_validator import validate_points
from app.services.forecast_pipeline import retrain_indicator_forecast
from app.services.rosstat_sdds_fetcher import fetch_sdds_xlsx
from app.core.cache import cache_invalidate_indicator

logger = logging.getLogger(__name__)


@dataclass
class DataPoint:
    date: date
    value: float


def _parse_header_date(header: str) -> date | None:
    if not header or not isinstance(header, str):
        return None
    parts = header.strip().split(".")
    if len(parts) != 2:
        return None
    try:
        month, year = int(parts[0]), int(parts[1])
        if 1 <= month <= 12 and 1990 <= year <= 2100:
            return date(year, month, 1)
    except (ValueError, TypeError):
        pass
    return None


def parse_ppi_xlsx(content: bytes) -> list[DataPoint]:
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    ws = None
    for name in wb.sheetnames:
        if "price" in name.lower() and "2010" in name:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.worksheets[0]

    rows_data: list[list] = []
    for row in ws.iter_rows(values_only=True):
        rows_data.append(list(row))
    wb.close()

    if len(rows_data) < 3:
        raise ValueError(f"PPI XLSX: expected >=3 rows, got {len(rows_data)}")

    dates: list[tuple[int, date]] = []
    for col_idx in range(2, len(rows_data[0])):
        header = rows_data[0][col_idx]
        d = _parse_header_date(str(header) if header else "")
        if d:
            dates.append((col_idx, d))

    if not dates:
        raise ValueError("PPI XLSX: no valid date headers found")

    ppi_row_idx = None
    for i, row in enumerate(rows_data):
        label = str(row[0] or "").lower()
        if "producer" in label and "price" in label:
            ppi_row_idx = i
            break
    if ppi_row_idx is None:
        ppi_row_idx = 2

    ppi_row = rows_data[ppi_row_idx]
    points: list[DataPoint] = []
    for col_idx, d in dates:
        val = ppi_row[col_idx] if col_idx < len(ppi_row) else None
        if val is not None:
            try:
                points.append(DataPoint(date=d, value=round(float(val), 2)))
            except (ValueError, TypeError):
                pass

    points.sort(key=lambda p: p.date)
    return points


class RosstatPpiParser(BaseParser):
    parser_type: ClassVar[str] = "rosstat_sdds_ppi"

    async def run(self, db: AsyncSession, indicator: Indicator, fetch_log: FetchLog) -> None:
        code = indicator.code
        try:
            content, final_url = await asyncio.to_thread(fetch_sdds_xlsx, "prices")
            fetch_log.source_url = final_url[:500]

            points = await asyncio.to_thread(parse_ppi_xlsx, content)

            cfg = indicator.model_config_json or {}
            points = validate_points(points, cfg)

            if not points:
                fetch_log.status = "no_new_data"
                fetch_log.error_message = "Parser returned 0 data points"
                fetch_log.completed_at = datetime.now(timezone.utc)
                await db.commit()
                return

            count_before = (await db.execute(
                select(func.count(IndicatorData.id))
                .where(IndicatorData.indicator_id == indicator.id)
            )).scalar() or 0

            for p in points:
                await db.execute(upsert_indicator_data(indicator.id, p.date, p.value))

            await db.flush()
            count_after = (await db.execute(
                select(func.count(IndicatorData.id))
                .where(IndicatorData.indicator_id == indicator.id)
            )).scalar() or 0

            records_added = count_after - count_before
            fetch_log.records_added = records_added
            logger.info("PPI '%s': +%d rows (total %d)", code, records_added, count_after)

            steps = int(cfg.get("forecast_steps", 0) or 0)
            if steps > 0 and records_added > 0:
                await retrain_indicator_forecast(db, indicator)

            if records_added > 0:
                await cache_invalidate_indicator(code)

            fetch_log.status = "success" if records_added > 0 else "no_new_data"
            fetch_log.completed_at = datetime.now(timezone.utc)
            await db.commit()

        except Exception as e:
            logger.exception("ETL failed for '%s'", code)
            await db.rollback()
            fetch_log.status = "failed"
            fetch_log.error_message = str(e)[:500]
            fetch_log.completed_at = datetime.now(timezone.utc)
            db.add(fetch_log)
            await db.commit()
