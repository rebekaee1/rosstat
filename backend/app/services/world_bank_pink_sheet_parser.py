"""ETL parser for World Bank Commodity Markets «Pink Sheet» monthly prices.

Source: World Bank Prospects Group, Commodity Markets (The Pink Sheet).
Landing page (URL of the XLSX changes every month — discover, do not hardcode):
  https://www.worldbank.org/en/research/commodity-markets

Workbook sheet ``Monthly Prices``: row of commodity names, row of units,
then rows ``YYYYMmm`` with nominal USD period averages.

Config (``model_config_json``):
  pink_sheet_column  — exact header text, e.g. ``Coal, Australian``
  backfill_from      — optional ISO date; earlier months dropped after parse

License basis (verified 2026-08-16): World Bank datasets default to
Creative Commons Attribution 4.0 (CC-BY 4.0) with attribution to the World
Bank; see https://data.worldbank.org/summary-terms-of-use. Public ``source``
field must say «Всемирный банк», not vendor feed names from the Description
sheet.

``replace_series = True``: a full monthly snapshot replaces any leftover
daily Yahoo history for the same indicator code.
"""
from __future__ import annotations

import asyncio
import logging
import re
from datetime import date
from io import BytesIO
from typing import ClassVar

import httpx
import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import FetchLog, Indicator
from app.services.base_parser import BaseParser

logger = logging.getLogger(__name__)

_LANDING_URL = "https://www.worldbank.org/en/research/commodity-markets"
_UA = "ForecastEconomy/1.0 (+https://forecasteconomy.com)"
_MONTHLY_XLSX_RE = re.compile(
    r"https://thedocs\.worldbank\.org/en/doc/[^\"'\s]+/CMO-Historical-Data-Monthly\.xlsx",
    re.IGNORECASE,
)
_PERIOD_RE = re.compile(r"^(\d{4})M(\d{2})$")
_MISSING = {"", "…", "...", "….", "-", "—", "na", "n.a.", "n/a"}


def _discover_monthly_xlsx_url(html: str) -> str | None:
    match = _MONTHLY_XLSX_RE.search(html)
    return match.group(0) if match else None


def _fetch_landing_html() -> str:
    with httpx.Client(
        timeout=60.0,
        headers={"User-Agent": _UA},
        follow_redirects=True,
    ) as client:
        response = client.get(_LANDING_URL)
        response.raise_for_status()
        return response.text


def _fetch_xlsx_bytes(url: str) -> bytes:
    with httpx.Client(
        timeout=120.0,
        headers={"User-Agent": _UA},
        follow_redirects=True,
    ) as client:
        response = client.get(url)
        response.raise_for_status()
        return response.content


def _parse_period(raw: object) -> date | None:
    if raw is None:
        return None
    text = str(raw).strip()
    match = _PERIOD_RE.match(text)
    if not match:
        return None
    year = int(match.group(1))
    month = int(match.group(2))
    if month < 1 or month > 12:
        return None
    return date(year, month, 1)


def _parse_value(raw: object) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if text.lower() in _MISSING or text in _MISSING:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return None


def _parse_pink_sheet_monthly(
    content: bytes,
    column_name: str,
    *,
    backfill_from: date | None = None,
) -> list[tuple[date, float]]:
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    if "Monthly Prices" not in wb.sheetnames:
        logger.warning("Pink Sheet workbook missing 'Monthly Prices' sheet")
        return []
    ws = wb["Monthly Prices"]

    name_row = next(ws.iter_rows(min_row=5, max_row=5, values_only=True), None)
    if not name_row:
        return []

    col_idx: int | None = None
    for i, name in enumerate(name_row):
        if name is not None and str(name).strip() == column_name:
            col_idx = i
            break
    if col_idx is None:
        logger.warning("Pink Sheet column %r not found", column_name)
        return []

    out: list[tuple[date, float]] = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row or col_idx >= len(row):
            continue
        d = _parse_period(row[0])
        if d is None:
            continue
        if backfill_from is not None and d < backfill_from:
            continue
        value = _parse_value(row[col_idx])
        if value is None:
            continue
        out.append((d, value))
    return out


class WorldBankPinkSheetParser(BaseParser):
    parser_type: ClassVar[str] = "world_bank_pink_sheet"
    replace_series: ClassVar[bool] = True

    async def _fetch_and_parse(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
    ) -> tuple[list, str]:
        column = str(cfg.get("pink_sheet_column") or "").strip()
        if not column:
            fetch_log.error_message = "pink_sheet_column missing in model_config_json"
            return [], _LANDING_URL

        backfill_from: date | None = None
        raw_from = cfg.get("backfill_from")
        if raw_from:
            try:
                backfill_from = date.fromisoformat(str(raw_from))
            except ValueError:
                logger.warning(
                    "Invalid backfill_from %r for %s — ignoring",
                    raw_from,
                    indicator.code,
                )

        try:
            html = await asyncio.to_thread(_fetch_landing_html)
            xlsx_url = _discover_monthly_xlsx_url(html)
            if not xlsx_url:
                fetch_log.error_message = "Pink Sheet monthly XLSX URL not found on landing page"
                return [], _LANDING_URL
            content = await asyncio.to_thread(_fetch_xlsx_bytes, xlsx_url)
        except Exception as exc:
            fetch_log.error_message = f"Pink Sheet fetch failed: {exc}"[:500]
            return [], _LANDING_URL

        points = _parse_pink_sheet_monthly(
            content, column, backfill_from=backfill_from,
        )
        return points, xlsx_url
