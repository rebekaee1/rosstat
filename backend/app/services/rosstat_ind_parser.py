"""
Parser for Rosstat monthly indicator compilation: ind_MM-YYYY.xlsx.

Source: rosstat.gov.ru/storage/mediabank/ind_MM-YYYY.xlsx
Sheets used:
  - "1.12 " → Retail trade turnover (млрд руб.)
  - "1.8 "  → Housing commissioned (млн кв.м)

Structure:
  Row 0-1: headers (Year, Quarters, Months)
  Row 2-3: section title
  Row 4+: data (Year | Annual | Q1-Q4 | Jan-Dec)
  Months in cols G-R (indices 6-17)
"""

from __future__ import annotations

import io
import logging
import math
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import ClassVar

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models import Indicator, FetchLog
from app.services.http_client import create_session
from app.services.base_parser import BaseParser
from app.services.calculation_engine import calculation_engine
from app.services.upsert import prune_indicator_dates_not_in

logger = logging.getLogger(__name__)

BASE_URL = "https://rosstat.gov.ru/storage/mediabank/"


@dataclass
class DataPoint:
    date: date
    value: float


def _to_float(cell) -> float | None:
    if cell is None:
        return None
    s = str(cell).strip().replace("\u2212", "-").replace(",", ".").replace("\xa0", "").replace(" ", "")
    if s in ("", "…", "-", "..."):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _extract_year(cell) -> int | None:
    if cell is None:
        return None
    s = str(cell).strip()
    m = re.match(r"(\d{4})", s)
    if m:
        y = int(m.group(1))
        if 1990 <= y <= 2100:
            return y
    return None


def parse_ind_sheet(content: bytes, sheet_name: str) -> list[DataPoint]:
    """Parse a sheet from ind_MM-YYYY.xlsx → monthly DataPoints."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = None
        for name in wb.sheetnames:
            if sheet_name.strip() in name.strip():
                ws = wb[name]
                break
        if ws is None:
            raise ValueError(f"Sheet '{sheet_name}' not found in {wb.sheetnames[:10]}")

        rows_data = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    points = []
    seen: set[tuple[int, int]] = set()
    data_started = False

    QUARTER_START_MONTH = {0: 1, 1: 4, 2: 7, 3: 10}

    for row in rows_data:
        if not row or len(row) < 3:
            continue

        first_cell = str(row[0] or "").strip().lower()
        if data_started and ("в %" in first_cell or "percent" in first_cell):
            break

        year = _extract_year(row[0])
        if year is None:
            continue

        data_started = True

        has_monthly = False
        for month_idx in range(12):
            col = 6 + month_idx
            if col >= len(row):
                break
            key = (year, month_idx + 1)
            if key in seen:
                continue
            val = _to_float(row[col])
            if val is not None and val > 0:
                seen.add(key)
                points.append(DataPoint(
                    date=date(year, month_idx + 1, 1),
                    value=round(val, 2),
                ))
                has_monthly = True

        if not has_monthly:
            for q_idx in range(4):
                col = 2 + q_idx
                if col >= len(row):
                    break
                month = QUARTER_START_MONTH[q_idx]
                key = (year, month)
                if key in seen:
                    continue
                val = _to_float(row[col])
                if val is not None and val > 0:
                    seen.add(key)
                    points.append(DataPoint(
                        date=date(year, month, 1),
                        value=round(val, 2),
                    ))

    return sorted(points, key=lambda p: p.date)


_QUARTER_START_MONTH = {0: 1, 1: 4, 2: 7, 3: 10}


def collapse_flow_to_quarterly(points: list[DataPoint]) -> list[DataPoint]:
    """Свернуть помесячные потоки в квартальные суммы на якорях квартала.

    Лист 1.6 (инвестиции в ОК) до ~2016 отдаёт помесячные потоки; с 2016 — только
    квартальные колонки. Без свёртки YoY сравнивает Q1 с одним месяцем прошлого
    года и даёт ложный всплеск (~300% в 2016).
    """
    buckets: dict[tuple[int, int], float] = {}
    for p in points:
        q = (p.date.month - 1) // 3
        key = (p.date.year, q)
        buckets[key] = buckets.get(key, 0.0) + p.value

    collapsed = [
        DataPoint(
            date=date(year, _QUARTER_START_MONTH[q], 1),
            value=round(total, 2),
        )
        for (year, q), total in sorted(buckets.items())
    ]
    return collapsed


def _fetch_latest_ind(session) -> tuple[bytes, str]:
    """Try to download the latest ind_MM-YYYY.xlsx by trying recent months."""
    now = datetime.now()
    candidates = []
    for offset in range(6):
        m = now.month - offset
        y = now.year
        while m <= 0:
            m += 12
            y -= 1
        candidates.append(f"ind_{m:02d}-{y}.xlsx")

    for fn in candidates:
        url = BASE_URL + fn
        try:
            resp = session.get(url, timeout=90)
            ct = resp.headers.get("content-type", "")
            if "html" in ct.lower() and resp.status_code == 200:
                logger.warning("Got HTML instead of XLSX from %s", url)
                continue
            if resp.status_code == 200 and resp.content[:4] == b"PK\x03\x04":
                logger.info("Downloaded %s (%d KB)", fn, len(resp.content) // 1024)
                return resp.content, url
        except Exception as e:
            logger.debug("Download failed for %s: %s", url, e)
            continue

    raise ValueError(f"ind XLSX not found (tried {candidates})")


SHEET_MAP = {
    "retail-trade": "1.12 ",
    "housing-commissioned": "1.8 ",
    "construction-work": "1.7 ",
    "capital-investment": "1.6 ",
}


class RosstatIndParser(BaseParser):
    parser_type: ClassVar[str] = "rosstat_ind_monthly"

    async def _fetch_and_parse(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
    ) -> tuple[list, str]:
        code = indicator.code
        sheet_name = cfg.get("ind_sheet", SHEET_MAP.get(code, ""))
        if not sheet_name:
            raise ValueError(f"No sheet mapping for indicator {code}")

        session = create_session()
        try:
            session.verify = settings.rosstat_ca_cert
            content, url = _fetch_latest_ind(session)
        finally:
            session.close()

        return parse_ind_sheet(content, sheet_name), url

    def _validate(self, points: list, cfg: dict) -> list:
        valid = [
            p for p in points
            if isinstance(p.value, (int, float)) and not math.isnan(p.value)
        ]
        if len(valid) < len(points):
            logger.warning("Filtered out %d invalid (NaN/non-numeric) ind values", len(points) - len(valid))
        if cfg.get("quarterly_flow"):
            return collapse_flow_to_quarterly(valid)
        return valid

    async def _post_upsert(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
        points: list,
        records_added: int,
        records_updated: int,
    ) -> tuple[int, int]:
        if not cfg.get("quarterly_flow"):
            return 0, 0
        pruned = await prune_indicator_dates_not_in(db, indicator.id, points)
        if pruned:
            logger.info(
                "Pruned %d stale monthly date(s) for '%s' (quarterly_flow)",
                pruned, indicator.code,
            )
        return pruned, 0

    async def _after_storage(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
        pruned: int,
        records_added: int,
        records_updated: int,
    ) -> None:
        if not cfg.get("quarterly_flow"):
            return
        if records_added <= 0 and records_updated <= 0:
            return
        derived = await calculation_engine.run_for_direct_dependents(db, [indicator.code])
        if derived:
            logger.info(
                "CalculationEngine (quarterly_flow): %s → %s",
                indicator.code, derived,
            )
