"""ETL: Еженедельный ИПЦ → IndicatorData.

Источники (по приоритету):

1. HTML-бюллетени «Об оценке индекса потребительских цен с N по M месяца YYYY г»
   на `rosstat.gov.ru/storage/mediabank/<num>_<DD-MM-YYYY>.html`.
   Публикуются каждую неделю (обычно в среду) и содержат официальный
   агрегированный недельный ИПЦ. Это основной источник.

   **Важно**: Росстат начал публиковать **отдельные** еженедельные bulletin'ы
   только в 2023 году. До 2023 публиковались monthly «Об индексе
   потребительских цен в <месяц>» — других еженедельных документов нет
   (verified 2026-05-12 deep dive Wayback Machine CDX: 174 candidate URLs
   за 2021-2022-2023 → 0 weekly bulletin'ов за 2021, 0 за 2022, 1 за 2023-02-08).
   См. `docs/audits/weekly_inflation_research_2026-05.md`.

   Discovery (union стратегий, `_find_bulletin_urls`):
   - `_find_bulletin_urls_central_news` — пагинированный crawl
     `rosstat.gov.ru/central-news?page=1..N`. На сайте Росстата архив
     ограничен **с 2023-05-04** до today (раньше — Росстат удалил из ленты).
   - `ROSSTAT_SEARCH_URL` — поиск по слову «оценке индекса потребительских цен
     <месяц> <год>» (fallback / current-week edge cases / 2023-01..04
     не покрытые central-news).

2. Фоллбэк — `Nedel_ipc.xlsx` (~110 товаров, листы по годам) + `ipc_spr_MM-YYYY.xlsx`
   (веса). Взвешенное среднее по продовольственной корзине — приближение, поэтому
   используется только для дат **≥ `weekly_cutoff_date`** (config). Текущий
   cutoff = 2023-01-09 (первая неделя где есть данные в XLSX за 2023).
   До этой даты у Росстата вообще не было отдельных weekly публикаций;
   XLSX-приближение расходилось с bulletin до 0.1pp на 2023 (good enough)
   и до 3 pp на 2022 (заметные расхождения — bulletinов не было, не показываем).

HTML-источник имеет приоритет: если одна и та же дата есть в обеих коллекциях —
берётся значение из бюллетеня (оно совпадает с официальным Росстатом).
"""

from __future__ import annotations

import asyncio
import logging
import re
from dataclasses import dataclass
from datetime import date, timedelta
from difflib import get_close_matches
from io import BytesIO
from typing import ClassVar

import openpyxl
import requests
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.cache import cache_invalidate_indicator
from app.models import FetchLog, Indicator, IndicatorData
from app.services.base_parser import BaseParser, _utcnow_naive
from app.services.http_client import create_session
from app.services.upsert import bulk_upsert

logger = logging.getLogger(__name__)

NEDEL_IPC_URL = "https://rosstat.gov.ru/storage/mediabank/nedel_Ipc.xlsx"
NEDEL_IPC_URL_FALLBACK = "https://rosstat.gov.ru/storage/mediabank/Nedel_ipc.xlsx"
IPC_SPR_URL = "https://rosstat.gov.ru/storage/mediabank/ipc_spr_{mm}-{yyyy}.xlsx"

ROSSTAT_SEARCH_URL = "https://rosstat.gov.ru/search"
BULLETIN_URL_RE = re.compile(r"/storage/mediabank/\d+_\d{2}-\d{2}-\d{4}\.html?")

_MONTH_MAP = {
    "января": 1, "февраля": 2, "марта": 3, "апреля": 4,
    "мая": 5, "июня": 6, "июля": 7, "августа": 8,
    "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12,
}

# Предлог начала периода — «с» либо «со» (Росстат пишет «со 2 по 8 июня»:
# «со второго» по правилам русского, перед числами, читающимися с «вт-»).
# `со?` без него bulletin за такую неделю не парсился и ETL возвращал 0 точек.
_BULLETIN_RANGE_RE = re.compile(
    r"\bсо?\s+\d{1,2}\s+(?:[а-яё]+\s+)?по\s+(\d{1,2})\s+([а-яё]+)\s+(\d{4})\s*г",
    re.IGNORECASE,
)
_BULLETIN_VALUE_RE = re.compile(r"составил\s+(\d+[,.]?\d*)\s*%")

_HEADER_RE = re.compile(
    r"на\s+(\d{1,2})\s+("
    + "|".join(_MONTH_MAP.keys())
    + r")",
    re.IGNORECASE,
)


@dataclass
class WeeklyPoint:
    date: date
    value: float


WEEKLY_SEGMENT_CODES = {
    "food": "inflation-weekly-food",
    "nonfood": "inflation-weekly-nonfood",
    "services": "inflation-weekly-services",
}


def _classify_local_code(code) -> str | None:
    """Классификация позиции справочника ipc_spr → food | nonfood | services."""
    if code is None:
        return None
    if isinstance(code, str):
        raw = code.strip().replace(",", ".")
        try:
            num = float(raw)
        except ValueError:
            return "services"
    else:
        try:
            num = float(code)
        except (TypeError, ValueError):
            return None
    if num >= 9000:
        return "services"
    if num >= 4100:
        return "nonfood"
    if num >= 10:
        return "food"
    return None


def _parse_column_date(header: str, year: int) -> date | None:
    """Parse 'на 10 января **' → date(year, 1, 10)."""
    m = _HEADER_RE.search(header)
    if not m:
        return None
    day = int(m.group(1))
    month = _MONTH_MAP.get(m.group(2).lower())
    if month is None:
        return None
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _load_product_weights(session: requests.Session) -> dict[str, tuple[float, str]]:
    """Download ipc_spr and extract per-product weights with CPI segment.

    Tries recent months (descending) to find the latest available file.
    Returns {product_name: (weight, segment)} where segment ∈ food|nonfood|services.
    """
    today = date.today()
    for month_offset in range(0, 6):
        m = today.month - month_offset
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        url = IPC_SPR_URL.format(mm=f"{m:02d}", yyyy=y)
        try:
            r = session.get(url, timeout=30, verify=False)
            if r.status_code == 200 and len(r.content) > 5000:
                break
        except requests.RequestException:
            continue
    else:
        logger.warning("Could not download ipc_spr for weights")
        return {}

    wb = openpyxl.load_workbook(BytesIO(r.content), data_only=True)
    sheet_name = None
    for sn in reversed(wb.sheetnames):
        if sn != "Содержание":
            sheet_name = sn
            break
    if not sheet_name:
        return {}

    ws = wb[sheet_name]
    weights: dict[str, tuple[float, str]] = {}
    for ri in range(7, ws.max_row + 1):
        name = str(ws.cell(ri, 1).value or "").strip()
        w_raw = ws.cell(ri, 3).value
        local_code = ws.cell(ri, 2).value
        if not name or w_raw is None:
            continue
        try:
            w = float(w_raw)
        except (ValueError, TypeError):
            continue
        segment = _classify_local_code(local_code)
        if w > 0 and segment:
            weights[name] = (w, segment)
    return weights


def _match_product(
    name: str,
    weights: dict[str, tuple[float, str]],
    cache: dict[str, tuple[float, str] | None],
) -> tuple[float, str] | None:
    """Find (weight, segment) for a weekly product name, caching fuzzy results."""
    if name in cache:
        return cache[name]
    hit = weights.get(name)
    if hit is not None:
        cache[name] = hit
        return hit
    close = get_close_matches(name, weights.keys(), n=1, cutoff=0.75)
    if close:
        hit = weights[close[0]]
        cache[name] = hit
        return hit
    cache[name] = None
    return None


def _parse_weekly_xlsx_multi(
    weekly_content: bytes,
    weights: dict[str, tuple[float, str]],
    *,
    years: set[int] | None = None,
) -> dict[str, list[WeeklyPoint]]:
    """Parse Nedel_ipc.xlsx → weighted weekly CPI for all + per segment.

    `years`: если задано — парсим только листы с этими годами (steady-state).
    """
    wb = openpyxl.load_workbook(BytesIO(weekly_content), data_only=True)
    buckets: dict[str, list[WeeklyPoint]] = {
        "all": [],
        "food": [],
        "nonfood": [],
        "services": [],
    }
    match_cache: dict[str, tuple[float, str] | None] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == "Содержание":
            continue
        try:
            year = int(sheet_name)
        except ValueError:
            continue
        if years is not None and year not in years:
            continue

        ws = wb[sheet_name]
        header_row = 4

        col_dates: list[tuple[int, date]] = []
        for ci in range(2, ws.max_column + 1):
            hdr = str(ws.cell(header_row, ci).value or "")
            d = _parse_column_date(hdr, year)
            if d:
                col_dates.append((ci, d))

        if not col_dates:
            continue

        products: list[tuple[int, str, float, str]] = []
        for ri in range(5, ws.max_row + 1):
            name = str(ws.cell(ri, 1).value or "").strip()
            if not name or name.startswith("*") or name.startswith("…"):
                continue
            matched = _match_product(name, weights, match_cache)
            if matched is None:
                continue
            w, segment = matched
            products.append((ri, name, w, segment))

        for ci, d in col_dates:
            sums: dict[str, tuple[float, float]] = {
                key: (0.0, 0.0) for key in buckets
            }
            for ri, _name, w, segment in products:
                raw = ws.cell(ri, ci).value
                if raw is None or raw == "…" or raw == "":
                    continue
                try:
                    val = float(str(raw).replace(",", ".").replace("\u2212", "-"))
                except (ValueError, TypeError):
                    continue
                if not (95 < val < 110):
                    continue
                for key in ("all", segment):
                    ws_sum, w_sum = sums[key]
                    sums[key] = (ws_sum + w * val, w_sum + w)

            for key, (weighted_sum, weight_sum) in sums.items():
                if weight_sum > 0:
                    aggregate = weighted_sum / weight_sum
                    buckets[key].append(WeeklyPoint(date=d, value=round(aggregate, 2)))

    for key in buckets:
        buckets[key].sort(key=lambda p: p.date)
    return buckets


def _parse_weekly_xlsx(
    weekly_content: bytes,
    weights: dict[str, tuple[float, str]],
) -> list[WeeklyPoint]:
    """Backward-compatible: full-basket weekly CPI from XLSX."""
    return _parse_weekly_xlsx_multi(weekly_content, weights)["all"]


def _strip_html(html: str) -> str:
    text = re.sub(r"<[^>]+>", " ", html)
    return re.sub(r"\s+", " ", text)


def _parse_bulletin_html(html: str) -> WeeklyPoint | None:
    """Extract (end_date, value) from a weekly CPI bulletin HTML."""
    text = _strip_html(html)
    m_range = _BULLETIN_RANGE_RE.search(text)
    m_val = _BULLETIN_VALUE_RE.search(text)
    if not m_range or not m_val:
        return None
    end_day = int(m_range.group(1))
    month = _MONTH_MAP.get(m_range.group(2).lower())
    year = int(m_range.group(3))
    if month is None:
        return None
    try:
        d = date(year, month, end_day)
    except ValueError:
        return None
    try:
        v = float(m_val.group(1).replace(",", "."))
    except ValueError:
        return None
    if not (95 < v < 110):
        return None
    return WeeklyPoint(date=d, value=round(v, 2))


CENTRAL_NEWS_URL = "https://rosstat.gov.ru/central-news"
CENTRAL_NEWS_MAX_PAGES = 70
CENTRAL_NEWS_STEADY_MAX_PAGES = 12
STEADY_SEARCH_MONTHS_BACK = 4
# Окно «перечитываем недели заново»: последние N дней всегда повторно
# скачиваются и передаются в идемпотентный upsert (восстановление пропусков +
# ревизии Росстата). Старее окна известные week-end не трогаем.
WEEKLY_REFRESH_WINDOW_DAYS = 120
_CPI_BULLETIN_TITLE_RE = re.compile(
    r"оценке\s+индекса\s+потребительских\s+цен", re.IGNORECASE,
)


def _find_bulletin_urls_central_news(
    session: requests.Session, year: int, max_pages: int = CENTRAL_NEWS_MAX_PAGES,
) -> list[str]:
    """Discover weekly CPI bulletin URLs via central-news pagination.

    Альтернатива `_find_bulletin_urls` (через rosstat search) — central-news
    crawl надёжнее: пагинированный список новостей с заголовками и стабильным
    layout, нет лимитов поискового API. Архив доступен с **2023-05-04** до
    последней публикации (на page=1). За пределы 2023-05 уходить нельзя:
    page=66+ возвращают пустую ленту (Росстат очистил).

    Останавливаемся когда:
    1. На странице нашли bulletin с годом меньше `year` (мы прошли границу)
    2. Страница пустая (mediabank-links нет)
    3. Достигли `max_pages`.
    """
    found: set[str] = set()
    for page in range(1, max_pages + 1):
        try:
            r = session.get(
                CENTRAL_NEWS_URL,
                params={"page": page},
                timeout=20,
                verify=False,
            )
            if r.status_code != 200:
                logger.warning(
                    "central-news page=%d HTTP %d", page, r.status_code,
                )
                continue
            text = r.text
            items = re.findall(
                r'href="(/storage/mediabank/\d+_\d{2}-\d{2}-\d{4}\.html)"\s*[^>]*>\s*([^<]{15,300})',
                text,
            )
            if not items:
                logger.info("central-news page=%d empty — stopping", page)
                break
            page_year_min = 9999
            page_year_max = 0
            page_added = 0
            for href, title in items:
                ym = re.search(r"-(\d{4})\.html$", href)
                if not ym:
                    continue
                bul_year = int(ym.group(1))
                page_year_min = min(page_year_min, bul_year)
                page_year_max = max(page_year_max, bul_year)
                if bul_year != year:
                    continue
                if not _CPI_BULLETIN_TITLE_RE.search(title):
                    continue
                found.add("https://rosstat.gov.ru" + href)
                page_added += 1
            logger.debug(
                "central-news page=%d: years=%d-%d, added=%d (cumulative=%d)",
                page, page_year_min, page_year_max, page_added, len(found),
            )
            if page_year_max < year:
                logger.info(
                    "central-news page=%d max_year=%d < %d — stopping crawl",
                    page, page_year_max, year,
                )
                break
        except requests.RequestException as exc:
            logger.warning(
                "central-news page=%d request failed: %s", page, exc,
            )
    return sorted(found)


def _steady_xlsx_years(today: date) -> set[int]:
    """Годы листов Nedel_ipc для steady-state (январь — ещё недели прошлого года)."""
    years = {today.year}
    if today.month <= 2:
        years.add(today.year - 1)
    return years


def _find_bulletin_urls(
    session: requests.Session,
    year: int,
    *,
    steady_state: bool = False,
) -> list[str]:
    """Find weekly CPI bulletin URLs for `year`.

    Стратегия: central-news crawl (primary) + rosstat search (fallback).
    Объединение — set union. central-news вернёт всё что есть в архиве за
    год (2023-05-04 + → today), search закрывает edge cases когда новый
    bulletin ещё не на page=1 central-news, но уже индексирован поиском.

    В steady-state: central-news ограничен `CENTRAL_NEWS_STEADY_MAX_PAGES`,
    search — только последние `STEADY_SEARCH_MONTHS_BACK` месяца.
    """
    max_pages = CENTRAL_NEWS_STEADY_MAX_PAGES if steady_state else CENTRAL_NEWS_MAX_PAGES
    via_central = _find_bulletin_urls_central_news(session, year, max_pages=max_pages)
    logger.info(
        "Bulletin discovery for year=%d: central-news=%d urls (steady=%s, max_pages=%d)",
        year, len(via_central), steady_state, max_pages,
    )

    via_search: set[str] = set()
    today = date.today()
    if steady_state:
        month_range = range(
            max(1, today.month - STEADY_SEARCH_MONTHS_BACK + 1),
            today.month + 1,
        )
    elif year < today.year:
        month_range = range(1, 13)
    else:
        month_range = range(1, today.month + 1)
    for month in month_range:
        month_name = next(
            (name for name, num in _MONTH_MAP.items() if num == month), None,
        )
        if not month_name:
            continue
        q = f"оценке индекса потребительских цен {month_name} {year}"
        try:
            r = session.get(
                ROSSTAT_SEARCH_URL, params={"q": q}, timeout=30, verify=False,
            )
            if r.status_code != 200:
                continue
            for m in BULLETIN_URL_RE.finditer(r.text):
                path = m.group(0)
                if f"-{year}." in path:
                    via_search.add("https://rosstat.gov.ru" + path)
        except requests.RequestException as exc:
            logger.warning("Rosstat search failed for %s %d: %s", month_name, year, exc)

    logger.info(
        "Bulletin discovery for year=%d: search=%d urls, union total=%d",
        year, len(via_search), len(set(via_central) | via_search),
    )
    return sorted(set(via_central) | via_search)


_BULLETIN_PUB_DATE_RE = re.compile(r"/(\d+)_(\d{2})-(\d{2})-(\d{4})\.html$")


def _bulletin_pub_date(url: str) -> date | None:
    """Extract publication date from a bulletin URL.

    The publication date in the URL is **not** the same as the week-end date
    inside the bulletin (publication is typically Wed; week-end is Sun/Mon
    of the prior week). But it's close enough — week-end is within 7-10 days
    before publication. We use that for skip-heuristics in steady-state.
    """
    m = _BULLETIN_PUB_DATE_RE.search(url)
    if not m:
        return None
    try:
        return date(int(m.group(4)), int(m.group(3)), int(m.group(2)))
    except ValueError:
        return None


def fetch_bulletin_points(
    session: requests.Session,
    years: list[int],
    existing_dates: set[date] | None = None,
    *,
    steady_state: bool = False,
) -> list[WeeklyPoint]:
    """Fetch weekly CPI from Rosstat HTML bulletins for the given years.

    `existing_dates`: если задано, GET'ы для bulletin'ов с publication-датой
    `>= 14 days ранее max(existing_dates)` будут пропущены — их week-end
    точно уже в БД. Это снижает steady-state GET'ов от сотен до 1-3.
    """
    points: list[WeeklyPoint] = []
    seen_dates: set[date] = set()
    skip_pub_before: date | None = None
    if existing_dates:
        # Всегда перечитываем последние WEEKLY_REFRESH_WINDOW_DAYS заново
        # (восстанавливаем возможные пропуски недель и подхватываем ревизии
        # Росстата), даже если самая свежая известная точка отстала. Старее окна
        # пропускаем — те week-end уже стабильны в БД.
        refresh_floor = date.today() - timedelta(days=WEEKLY_REFRESH_WINDOW_DAYS)
        max_known = max(existing_dates)
        skip_pub_before = min(max_known - timedelta(days=14), refresh_floor)
    skipped = 0
    for year in years:
        urls = _find_bulletin_urls(session, year, steady_state=steady_state)
        logger.info("Weekly CPI bulletins for %d: %d URLs", year, len(urls))
        for url in urls:
            if skip_pub_before is not None:
                pub = _bulletin_pub_date(url)
                if pub is not None and pub < skip_pub_before:
                    skipped += 1
                    continue
            try:
                r = session.get(url, timeout=30, verify=False)
                if r.status_code != 200:
                    continue
                html = r.content.decode("utf-8", errors="replace")
                pt = _parse_bulletin_html(html)
                if pt and pt.date not in seen_dates:
                    seen_dates.add(pt.date)
                    points.append(pt)
            except requests.RequestException as exc:
                logger.warning("Failed to fetch bulletin %s: %s", url, exc)
    if skipped:
        logger.info(
            "Weekly CPI: skipped %d bulletin GETs (pub_date < %s, week-end already in DB)",
            skipped, skip_pub_before,
        )
    points.sort(key=lambda p: p.date)
    return points


def _filter_new_points(
    points: list[WeeklyPoint],
    known_dates: set[date] | None,
) -> list[WeeklyPoint]:
    """Оставить точки для записи: все новые + свежее окна обновления.

    Точки за последние `WEEKLY_REFRESH_WINDOW_DAYS` дней пропускаем к upsert,
    даже если дата уже есть в БД — идемпотентный upsert обновит только реально
    изменившиеся значения (ревизии Росстата). Это и есть «проходим последние
    недели заново, а не оставляем как есть». Старее окна известные даты не
    перезаписываем — они стабильны.
    """
    if not known_dates:
        return points
    refresh_floor = date.today() - timedelta(days=WEEKLY_REFRESH_WINDOW_DAYS)
    return [
        p for p in points
        if p.date not in known_dates or p.date >= refresh_floor
    ]


def fetch_weekly_cpi_multi(
    existing_dates: set[date] | None = None,
    cutoff_date: date | None = None,
    *,
    segment_existing: dict[str, set[date]] | None = None,
) -> dict[str, list[WeeklyPoint]]:
    """Fetch weekly CPI: all basket + food / nonfood / services segments.

    Ключ ``all`` — HTML-бюллетени (primary) + XLSX (fallback). Сегменты —
    только взвешенное среднее по Nedel_ipc с весами ipc_spr (без bulletin,
    т.к. Росстат публикует официальный недельный агрегат только по полной корзине).
    """
    def _apply_cutoff(points: list[WeeklyPoint]) -> list[WeeklyPoint]:
        if not cutoff_date:
            return points
        return [p for p in points if p.date >= cutoff_date]

    session = create_session()
    try:
        session.verify = False

        today = date.today()
        BULLETIN_FIRST_YEAR = 2023
        last_year = today.year - 1
        steady_state = bool(
            existing_dates and any(d.year <= last_year for d in existing_dates)
        )
        if steady_state:
            bulletin_years = [today.year]
            logger.info(
                "Weekly CPI steady-state: %d existing points incl. %d-%d, "
                "fetching only current year %d",
                len(existing_dates or ()), BULLETIN_FIRST_YEAR, last_year, today.year,
            )
        else:
            bulletin_years = list(range(BULLETIN_FIRST_YEAR, today.year + 1))
            logger.info(
                "Weekly CPI cold-start: fetching backfill for years %s", bulletin_years,
            )

        bulletin_points = fetch_bulletin_points(
            session, bulletin_years, existing_dates, steady_state=steady_state,
        )
        logger.info("Weekly CPI: parsed %d points from HTML bulletins", len(bulletin_points))

        xlsx_by_segment: dict[str, list[WeeklyPoint]] = {
            "all": [], "food": [], "nonfood": [], "services": [],
        }
        for xlsx_url in (NEDEL_IPC_URL, NEDEL_IPC_URL_FALLBACK):
            logger.info("Downloading weekly XLSX: %s", xlsx_url)
            try:
                r = session.get(xlsx_url, timeout=60)
                if r.status_code != 200:
                    logger.warning("HTTP %d for %s", r.status_code, xlsx_url)
                    continue
                product_weights = _load_product_weights(session)
                if not product_weights:
                    logger.warning("No weights available — XLSX fallback skipped")
                    break
                xlsx_years = _steady_xlsx_years(today) if steady_state else None
                parsed = _parse_weekly_xlsx_multi(
                    r.content, product_weights, years=xlsx_years,
                )
                if steady_state:
                    # В steady-state для полной корзины достаточно bulletin'ов;
                    # XLSX нужен каждую неделю для сегментов food/nonfood/services.
                    xlsx_by_segment["food"] = parsed["food"]
                    xlsx_by_segment["nonfood"] = parsed["nonfood"]
                    xlsx_by_segment["services"] = parsed["services"]
                    logger.info(
                        "Weekly CPI steady-state XLSX segments: food=%d nonfood=%d services=%d",
                        len(parsed["food"]), len(parsed["nonfood"]), len(parsed["services"]),
                    )
                else:
                    xlsx_by_segment = parsed
                    logger.info(
                        "Weekly CPI: parsed XLSX segments — all=%d food=%d nonfood=%d services=%d",
                        len(parsed["all"]),
                        len(parsed["food"]),
                        len(parsed["nonfood"]),
                        len(parsed["services"]),
                    )
                break
            except requests.RequestException as exc:
                logger.warning("XLSX fetch failed for %s: %s", xlsx_url, exc)

        merged_all: dict[date, float] = (
            {} if steady_state else {p.date: p.value for p in xlsx_by_segment["all"]}
        )
        for p in bulletin_points:
            merged_all[p.date] = p.value
        all_points = [
            WeeklyPoint(date=d, value=v) for d, v in sorted(merged_all.items())
        ]

        result: dict[str, list[WeeklyPoint]] = {
            "all": all_points,
            "food": list(xlsx_by_segment["food"]),
            "nonfood": list(xlsx_by_segment["nonfood"]),
            "services": list(xlsx_by_segment["services"]),
        }
        known_by_key: dict[str, set[date] | None] = {
            "all": existing_dates,
            "food": (segment_existing or {}).get("food"),
            "nonfood": (segment_existing or {}).get("nonfood"),
            "services": (segment_existing or {}).get("services"),
        }
        for key in result:
            before = len(result[key])
            result[key] = _apply_cutoff(result[key])
            result[key] = _filter_new_points(result[key], known_by_key.get(key))
            logger.info(
                "Weekly CPI segment %s: %d → %d points after cutoff/filter",
                key, before, len(result[key]),
            )
        return result
    finally:
        session.close()


def fetch_weekly_cpi(
    existing_dates: set[date] | None = None,
    cutoff_date: date | None = None,
) -> list[WeeklyPoint]:
    """Fetch weekly CPI for the full basket (HTML bulletins + XLSX fallback)."""
    return fetch_weekly_cpi_multi(existing_dates, cutoff_date)["all"]


class RosstatWeeklyCpiParser(BaseParser):
    parser_type: ClassVar[str] = "rosstat_weekly_cpi"

    def __init__(self) -> None:
        self._segment_points: dict[str, list[WeeklyPoint]] | None = None

    async def run(self, db: AsyncSession, indicator: Indicator, fetch_log: FetchLog) -> None:
        """Primary `inflation-weekly` must upsert segment siblings even when `all` has 0 new points."""
        seg_role = (indicator.model_config_json or {}).get("weekly_segment", "all")
        if seg_role not in (None, "all"):
            return await super().run(db, indicator, fetch_log)

        code = indicator.code
        try:
            cfg = indicator.model_config_json or {}
            points, source_url = await self._fetch_and_parse(db, indicator, cfg, fetch_log)
            if source_url:
                fetch_log.source_url = source_url[:500]

            segment_snapshot = self._segment_points
            self._segment_points = None

            points = self._validate(points, cfg)
            records_added = records_updated = 0
            if points:
                records_added, records_updated = await bulk_upsert(db, indicator.id, points)
                logger.info(
                    "Upserted %d new, %d updated for '%s'",
                    records_added, records_updated, code,
                )
                fetch_log.records_added = records_added
                fetch_log.records_updated = records_updated
            elif not any(segment_snapshot.get(s) for s in WEEKLY_SEGMENT_CODES):
                logger.warning("No data points parsed for %s", code)
                fetch_log.status = "no_new_data"
                if not fetch_log.error_message:
                    fetch_log.error_message = "Parser returned 0 data points"
                fetch_log.completed_at = _utcnow_naive()
                await db.commit()
                await self._alert_zero_parse_if_regression(db, indicator)
                return

            if segment_snapshot:
                self._segment_points = segment_snapshot
                extra_added, extra_updated = await self._post_upsert(
                    db, indicator, cfg, fetch_log, points, records_added, records_updated,
                )
                if extra_added or extra_updated:
                    records_added += extra_added
                    records_updated += extra_updated
                    fetch_log.records_added = records_added
                    fetch_log.records_updated = records_updated

            await self._handle_forecasts(db, indicator, cfg, records_added, records_updated)

            if records_added > 0 or records_updated > 0:
                await cache_invalidate_indicator(code)
                for seg_code in WEEKLY_SEGMENT_CODES.values():
                    await cache_invalidate_indicator(seg_code)

            fetch_log.status = (
                "success" if (records_added > 0 or records_updated > 0) else "no_new_data"
            )
            fetch_log.completed_at = _utcnow_naive()
            await db.commit()

        except Exception as exc:
            logger.exception("ETL failed for '%s'", code)
            await db.rollback()
            fetch_log.status = "failed"
            fetch_log.error_message = str(exc)[:500]
            fetch_log.completed_at = _utcnow_naive()
            db.add(fetch_log)
            await db.commit()

    def _zero_parse_expected(self, indicator: Indicator, cfg: dict) -> bool:
        # Сегментные ряды пусты by design — их наполняет primary-прогон;
        # zero-parse-алерт (Н-4) для них ложный.
        return cfg.get("weekly_segment") not in (None, "all")

    async def _fetch_and_parse(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
    ) -> tuple[list, str]:
        if cfg.get("weekly_segment") not in (None, "all"):
            # Сегментные ряды пишет primary `inflation-weekly` в _post_upsert.
            return [], NEDEL_IPC_URL

        cutoff_raw = cfg.get("weekly_cutoff_date")
        cutoff: date | None = None
        if cutoff_raw:
            try:
                cutoff = date.fromisoformat(str(cutoff_raw))
            except ValueError:
                logger.warning(
                    "Weekly CPI: invalid weekly_cutoff_date=%r, ignoring", cutoff_raw,
                )

        existing_dates, segment_existing = await self._load_existing_dates(db, indicator.id)
        logger.info(
            "Weekly CPI: %d existing primary points; segments in DB: %s",
            len(existing_dates),
            {k: len(v) for k, v in segment_existing.items()},
        )

        multi = await asyncio.to_thread(
            fetch_weekly_cpi_multi,
            existing_dates,
            cutoff,
            segment_existing=segment_existing,
        )
        self._segment_points = multi
        return multi.get("all", []), NEDEL_IPC_URL

    async def _load_existing_dates(
        self,
        db: AsyncSession,
        primary_indicator_id: int,
    ) -> tuple[set[date], dict[str, set[date]]]:
        primary_q = await db.execute(
            select(IndicatorData.date).where(IndicatorData.indicator_id == primary_indicator_id)
        )
        existing_dates = {row[0] for row in primary_q.all()}

        segment_existing: dict[str, set[date]] = {}
        for segment, seg_code in WEEKLY_SEGMENT_CODES.items():
            ind_q = await db.execute(
                select(Indicator.id).where(
                    Indicator.code == seg_code,
                    Indicator.is_active.is_(True),
                )
            )
            seg_id = ind_q.scalar_one_or_none()
            if seg_id is None:
                segment_existing[segment] = set()
                continue
            dates_q = await db.execute(
                select(IndicatorData.date).where(IndicatorData.indicator_id == seg_id)
            )
            segment_existing[segment] = {row[0] for row in dates_q.all()}
        return existing_dates, segment_existing

    async def _post_upsert(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
        points: list,
        records_added: int,
        records_updated: int,
    ) -> tuple[int, int]:
        if indicator.code != "inflation-weekly" or not self._segment_points:
            return 0, 0

        _, segment_existing = await self._load_existing_dates(db, indicator.id)

        extra_added = 0
        extra_updated = 0
        for segment, seg_code in WEEKLY_SEGMENT_CODES.items():
            seg_points = self._segment_points.get(segment) or []
            seg_points = _filter_new_points(seg_points, segment_existing.get(segment))
            if not seg_points:
                continue
            q = await db.execute(
                select(Indicator).where(
                    Indicator.code == seg_code,
                    Indicator.is_active.is_(True),
                )
            )
            target = q.scalar_one_or_none()
            if not target:
                logger.warning("Weekly CPI segment indicator %s not found", seg_code)
                continue
            added, updated = await bulk_upsert(db, target.id, seg_points)
            extra_added += added
            extra_updated += updated
            if added or updated:
                await cache_invalidate_indicator(seg_code)
            logger.info(
                "Weekly CPI segment %s → %s: %d new, %d updated",
                segment, seg_code, added, updated,
            )

        self._segment_points = None
        return extra_added, extra_updated
