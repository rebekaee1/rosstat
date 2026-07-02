"""ETL: Росстат еженедельные средние потребительские цены (абсолютные, руб.).

Источники (по приоритету свежести):

1. `rosstat.gov.ru/storage/mediabank/nedel_sred_cen.xlsx` — официальный файл
   «Еженедельные средние потребительские цены (на конец периода)». Листы по
   годам (2022, 2023, …, текущий). На каждом листе:

     row 0: «К содержанию»
     row 3: заголовки дат «на 12 января», «на 19 января», … (с col 1)
     row 4+: строки товаров, col 0 — наименование, далее значения по неделям.

   **Trap (2026-07-01):** Росстат обновляет xlsx с лагом — свежая неделя
   выходит сначала ТОЛЬКО в HTML-бюллетене (п. 2), а xlsx догоняет позже.

2. HTML-бюллетень «О потребительских ценах на нефтепродукты с N по M …»
   (`/storage/mediabank/<num>_<DD-MM-YYYY>.html`, публикуется по средам,
   discovery через central-news — переиспользуем механизм
   `rosstat_weekly_inflation_parser._find_bulletin_urls_central_news`).
   В бюллетене таблица «Средние потребительские цены на бензин автомобильный
   и дизельное топливо по Российской Федерации» (рублей за литр) с колонками
   двух дат регистрации. `_parse_fuel_bulletin_html` вытаскивает строки
   «марки АИ-92» / «марки АИ-95» / «Дизельное топливо». Только для
   fuel-меток; для прочих товаров источник один — xlsx.

Результат — union (бюллетень поверх xlsx на совпадающих датах: бюллетень
первичен, xlsx повторяет его же значения).

В отличие от `rosstat_weekly_inflation_parser` (недельный ИПЦ, % к пред. неделе),
здесь **абсолютная цена** в рублях за единицу (л, кг, шт.). Целевая строка
config-driven через `model_config_json.product_label` (точное совпадение col 0,
иначе подстрока).

Используется для топлива:
  fuel-ai92   «Бензин автомобильный марки АИ-92, л»
  fuel-ai95   «Бензин автомобильный марки АИ-95, л»
  fuel-diesel «Дизельное топливо, л»

Прогноз — месячный (avg-month → monthly_auto), см. view_model_families.
"""

from __future__ import annotations

import asyncio
import logging
import re
from dataclasses import dataclass
from datetime import date
from typing import ClassVar

import openpyxl
import requests
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import FetchLog, Indicator
from app.services.base_parser import BaseParser
from app.services.data_validator import validate_points
from app.services.http_client import create_session
from app.services.rosstat_weekly_inflation_parser import (
    _MONTH_MAP,
    _parse_column_date,
)

logger = logging.getLogger(__name__)

WEEKLY_SRED_CEN_URL = "https://rosstat.gov.ru/storage/mediabank/nedel_sred_cen.xlsx"
_HEADER_ROW = 3  # 0-based: строка «на 12 января …»

# --- HTML-бюллетень «О потребительских ценах на нефтепродукты» -------------
CENTRAL_NEWS_URL = "https://rosstat.gov.ru/central-news"
_FUEL_BULLETIN_TITLE_RE = re.compile(
    r"потребительских\s+ценах\s+на\s+нефтепродукты", re.IGNORECASE,
)
_BULLETIN_HREF_RE = re.compile(
    r'href="(/storage/mediabank/\d+_\d{2}-\d{2}-\d{4}\.html)"\s*[^>]*>\s*([^<]{15,300})',
)
_REG_DATE_RE = re.compile(
    r"(\d{1,2})\s*(" + "|".join(_MONTH_MAP.keys()) + r")\s*(\d{4})", re.IGNORECASE,
)

# product_label (как в model_config) -> подпись строки в таблице бюллетеня
_BULLETIN_ROW_BY_LABEL = {
    "бензин автомобильный марки аи-92, л": "аи-92",
    "бензин автомобильный марки аи-95, л": "аи-95",
    "дизельное топливо, л": "дизельное топливо",
}


@dataclass
class PricePoint:
    date: date
    value: float


def parse_weekly_price_xlsx(content: bytes, product_label: str) -> list[PricePoint]:
    """Parse nedel_sred_cen.xlsx → [(week_end_date, price)] for `product_label`.

    Совпадение строки: сначала точное (по col 0, без регистра/пробелов), затем
    подстрока. Берём первую подходящую строку на каждом листе-годе.
    """
    wb = openpyxl.load_workbook(io_bytes(content), data_only=True, read_only=True)
    target = product_label.strip().lower()
    out: dict[date, float] = {}
    try:
        for sheet_name in wb.sheetnames:
            try:
                year = int(sheet_name)
            except ValueError:
                continue
            ws = wb[sheet_name]
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            if len(rows) <= _HEADER_ROW:
                continue

            header = rows[_HEADER_ROW]
            col_dates: list[tuple[int, date]] = []
            for ci in range(1, len(header)):
                d = _parse_column_date(str(header[ci] or ""), year)
                if d:
                    col_dates.append((ci, d))
            if not col_dates:
                continue

            data_row = _find_product_row(rows, target)
            if data_row is None:
                continue

            for ci, d in col_dates:
                if ci >= len(data_row):
                    continue
                raw = data_row[ci]
                if raw is None or raw == "" or raw == "…":
                    continue
                try:
                    val = float(str(raw).replace(",", ".").replace("\u2212", "-"))
                except (ValueError, TypeError):
                    continue
                if val <= 0:
                    continue
                out[d] = round(val, 2)
    finally:
        wb.close()

    return [PricePoint(date=d, value=v) for d, v in sorted(out.items())]


def _find_product_row(rows: list[list], target: str) -> list | None:
    exact: list | None = None
    substr: list | None = None
    for row in rows:
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip().lower()
        if name == target:
            exact = row
            break
        if substr is None and target in name:
            substr = row
    return exact if exact is not None else substr


def io_bytes(content: bytes):
    from io import BytesIO

    return BytesIO(content)


def _parse_fuel_bulletin_html(html: str, row_label: str) -> list[PricePoint]:
    """Средние цены (руб/л) из бюллетеня «О потребительских ценах на нефтепродукты».

    Ищем таблицу под заголовком «Средние потребительские цены … по Российской
    Федерации»: шапка — две даты регистрации («22 июня 2026 г.», «29 июня
    2026 г.»), строки — «марки АИ-92» / «марки АИ-95» / «Дизельное топливо».
    """
    anchor = html.find("по Российской Федерации")
    if anchor < 0:
        return []
    frag = html[anchor:]
    tstart = frag.find("<table")
    tend = frag.find("</table>")
    if tstart < 0 or tend < 0:
        return []
    table = frag[tstart:tend]

    import html as html_mod

    rows: list[list[str]] = []
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", table, re.S):
        cells = [
            re.sub(r"\s+", " ", html_mod.unescape(re.sub(r"<[^>]+>", "", c)))
            .replace("\xa0", " ").strip()
            for c in re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", tr, re.S)
        ]
        cells = [c for c in cells if c]
        if cells:
            rows.append(cells)

    reg_dates: list[date] = []
    for cells in rows:
        found = []
        for c in cells:
            m = _REG_DATE_RE.search(c)
            if m:
                month = _MONTH_MAP.get(m.group(2).lower())
                if month:
                    try:
                        found.append(date(int(m.group(3)), month, int(m.group(1))))
                    except ValueError:
                        pass
        if len(found) >= 2:
            reg_dates = found
            break
    if not reg_dates:
        return []

    target = row_label.lower()
    for cells in rows:
        name = cells[0].lower()
        if target not in name:
            continue
        vals: list[float] = []
        for c in cells[1:]:
            try:
                vals.append(float(c.replace(",", ".").replace(" ", "")))
            except ValueError:
                continue
        if len(vals) < len(reg_dates):
            continue
        return [
            PricePoint(date=d, value=round(v, 2))
            for d, v in zip(reg_dates, vals)
            if 1 < v < 1000
        ]
    return []


def fetch_fuel_bulletin_points(
    session: requests.Session, product_label: str, max_pages: int = 3,
) -> list[PricePoint]:
    """Свежие цены топлива из HTML-бюллетеней ленты central-news.

    Смотрим только первые страницы ленты (бюллетень выходит еженедельно,
    steady-state — page=1); xlsx закрывает всю историю, бюллетень нужен
    ради 1-2 недель, которые в xlsx ещё не доехали.
    """
    row_label = _BULLETIN_ROW_BY_LABEL.get(product_label.strip().lower())
    if not row_label:
        return []
    out: dict[date, float] = {}
    for page in range(1, max_pages + 1):
        try:
            r = session.get(CENTRAL_NEWS_URL, params={"page": page}, timeout=20)
            if r.status_code != 200:
                continue
            for href, title in _BULLETIN_HREF_RE.findall(r.text):
                if not _FUEL_BULLETIN_TITLE_RE.search(title):
                    continue
                try:
                    br = session.get("https://rosstat.gov.ru" + href, timeout=30)
                    if br.status_code != 200:
                        continue
                    html = br.content.decode("utf-8", errors="replace")
                    for pt in _parse_fuel_bulletin_html(html, row_label):
                        out.setdefault(pt.date, pt.value)
                except requests.RequestException as exc:
                    logger.warning("fuel bulletin %s fetch failed: %s", href, exc)
        except requests.RequestException as exc:
            logger.warning("central-news page=%d failed: %s", page, exc)
    return [PricePoint(date=d, value=v) for d, v in sorted(out.items())]


def fetch_weekly_price(product_label: str) -> tuple[list[PricePoint], str]:
    session = create_session()
    try:
        session.verify = False
        r = session.get(WEEKLY_SRED_CEN_URL, timeout=60)
        r.raise_for_status()
        points = parse_weekly_price_xlsx(r.content, product_label)

        # Свежая неделя выходит в HTML-бюллетене раньше, чем догоняет xlsx.
        # Бюллетень первичен: на совпадающих датах его значение затирает xlsx.
        try:
            bulletin = fetch_fuel_bulletin_points(session, product_label)
        except Exception as exc:  # noqa: BLE001 — bulletin best-effort
            logger.warning("fuel bulletin discovery failed: %s", exc)
            bulletin = []
        if bulletin:
            merged = {p.date: p.value for p in points}
            merged.update({p.date: p.value for p in bulletin})
            points = [PricePoint(date=d, value=v) for d, v in sorted(merged.items())]
            logger.info(
                "weekly price %s: xlsx=%d, bulletin=%d, merged=%d (last=%s)",
                product_label, len(merged) - len(bulletin), len(bulletin),
                len(points), points[-1].date if points else None,
            )
        return points, WEEKLY_SRED_CEN_URL
    finally:
        session.close()


class RosstatWeeklyPriceParser(BaseParser):
    parser_type: ClassVar[str] = "rosstat_weekly_price"

    async def _fetch_and_parse(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
    ) -> tuple[list, str]:
        product_label = str(cfg.get("product_label") or "").strip()
        if not product_label:
            fetch_log.error_message = "rosstat_weekly_price: model_config.product_label не задан"
            return [], WEEKLY_SRED_CEN_URL
        try:
            points, url = await asyncio.to_thread(fetch_weekly_price, product_label)
        except requests.RequestException as e:
            fetch_log.error_message = f"nedel_sred_cen.xlsx fetch failed: {e}"[:500]
            return [], WEEKLY_SRED_CEN_URL
        return points, url

    def _validate(self, points: list, cfg: dict) -> list:
        return validate_points(points, cfg)
