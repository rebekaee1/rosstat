"""
Parsers for Rosstat demographic data.

Catalog pages (устойчивый discovery имён файлов):
  - https://rosstat.gov.ru/folder/12781 — демография
    * demo21_{YYYY}.xlsx|xls — рождаемость/смертность (годовая таблица)
    * demo14.xlsx — возрастные группы (трудоспособное население)
    * Edn_12-{YYYY}_t1.xlsx — оперативные итоги ЕДН за полный календарный год
      (дополняет demo21, пока годовая таблица не обновлена)
  - Pensioners: Sp_2.1_{YYYY}.xlsx (probe + fallback)

Internals: только секция «Все население» в demo21 (город/село отбрасываются
через first-seen year). Единицы: births/deaths — тыс. чел. (абсолют в файле
в чел. → /1000); rates — на 1000 чел. населения.
"""

from __future__ import annotations

import io
import logging
import math
import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import ClassVar

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.models import Indicator, FetchLog
from app.services.base_parser import BaseParser
from app.services.http_client import create_session
from app.services.rosstat_sdds_fetcher import (
    download_mediabank_bytes,
    list_mediabank_filenames_from_page,
    pick_mediabank_filename,
    resolve_mediabank_file,
)

logger = logging.getLogger(__name__)

DEMOGRAPHY_CATALOG_URL = "https://rosstat.gov.ru/folder/12781"
BASE_URL = "https://rosstat.gov.ru/storage/mediabank/"


@dataclass
class DataPoint:
    date: date
    value: float


def _extract_year(cell) -> int | None:
    if cell is None:
        return None
    s = str(cell).strip()
    m = re.match(r"(\d{4})", s)
    if m:
        y = int(m.group(1))
        if 1900 <= y <= 2100:
            return y
    return None


def _to_float(cell) -> float | None:
    if cell is None:
        return None
    s = str(cell).strip().replace("\u2212", "-").replace(",", ".").replace("\xa0", "").replace(" ", "")
    if s in ("", "…", "-", "..."):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_demo21_xlsx(content: bytes) -> dict[str, list[DataPoint]]:
    """Parse demo21_YYYY.xlsx → births, deaths, birth_rate, death_rate.

    Файл содержит три блока (всё / город / село) с повторяющимися годами.
    Берём только первое вхождение каждого года — это блок «Все население».
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb.worksheets[0]
        rows_data = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    births, deaths, birth_rate, death_rate = [], [], [], []
    seen_years: set[int] = set()

    for row in rows_data:
        if not row or len(row) < 7:
            continue
        year = _extract_year(row[0])
        if year is None or year < 1990 or year in seen_years:
            continue

        d = date(year, 1, 1)
        b = _to_float(row[1])
        de = _to_float(row[2])
        br = _to_float(row[4])
        dr = _to_float(row[5])

        # Пропуск строк-заголовков секций без чисел.
        if b is None and de is None and br is None and dr is None:
            continue

        seen_years.add(year)

        if b is not None:
            births.append(DataPoint(date=d, value=round(b / 1000, 1)))
        if de is not None:
            deaths.append(DataPoint(date=d, value=round(de / 1000, 1)))
        if br is not None:
            birth_rate.append(DataPoint(date=d, value=br))
        if dr is not None:
            death_rate.append(DataPoint(date=d, value=dr))

    return {
        "births": sorted(births, key=lambda p: p.date),
        "deaths": sorted(deaths, key=lambda p: p.date),
        "birth-rate": sorted(birth_rate, key=lambda p: p.date),
        "death-rate": sorted(death_rate, key=lambda p: p.date),
    }


def parse_edn_annual_t1_xlsx(content: bytes) -> dict[str, DataPoint] | None:
    """Parse Edn_12-{YYYY}_t1.xlsx — оперативные итоги РФ за полный календарный год.

    Births/deaths в файле уже в тысячах; коэффициенты — на 1000 населения.
    Отчётный год — первая метка «YYYY г.» в шапке (колонка текущего года).
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb.worksheets[0]
        rows_data = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    year: int | None = None
    for row in rows_data[:10]:
        years_in_row = []
        for cell in row:
            m = re.search(r"(20\d{2})\s*г", str(cell or ""), re.I)
            if m:
                years_in_row.append(int(m.group(1)))
        if years_in_row:
            year = years_in_row[0]
            break

    if year is None:
        return None

    births = deaths = birth_rate = death_rate = None
    for row in rows_data:
        if not row:
            continue
        label = str(row[0] or "").strip().lower().replace("\xa0", " ")
        if label.startswith("родившихся") and "из них" not in label:
            births = _to_float(row[1])
            birth_rate = _to_float(row[4]) if len(row) > 4 else None
        elif label.startswith("умерших") and "из них" not in label and "детей" not in label:
            deaths = _to_float(row[1])
            death_rate = _to_float(row[4]) if len(row) > 4 else None

    if births is None and deaths is None and birth_rate is None and death_rate is None:
        return None

    d = date(year, 1, 1)
    out: dict[str, DataPoint] = {}
    if births is not None:
        out["births"] = DataPoint(date=d, value=round(births, 1))
    if deaths is not None:
        out["deaths"] = DataPoint(date=d, value=round(deaths, 1))
    if birth_rate is not None:
        out["birth-rate"] = DataPoint(date=d, value=birth_rate)
    if death_rate is not None:
        out["death-rate"] = DataPoint(date=d, value=death_rate)
    return out


def parse_demo14_xlsx(content: bytes) -> dict[str, list[DataPoint]]:
    """Parse demo14.xlsx → 3 age groups (тыс. чел.): under-working, working, over-working."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb.worksheets[0]
        rows_data = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if len(rows_data) < 26:
        raise ValueError(f"demo14: expected >= 26 rows, got {len(rows_data)}")

    year_row = rows_data[5]

    age_groups: dict[str, list | None] = {
        "working-age-population": None,
        "pop-under-working-age": None,
        "pop-over-working-age": None,
    }

    for i in range(15, min(35, len(rows_data))):
        cell = str(rows_data[i][0] or "").lower().replace("\xa0", " ").strip()
        if "трудоспособном" in cell and "моложе" not in cell and "старше" not in cell:
            age_groups["working-age-population"] = rows_data[i]
        elif "моложе" in cell and "трудоспособ" in cell:
            age_groups["pop-under-working-age"] = rows_data[i]
        elif "старше" in cell and "трудоспособ" in cell:
            age_groups["pop-over-working-age"] = rows_data[i]

    result: dict[str, list[DataPoint]] = {}
    for code, data_row in age_groups.items():
        if data_row is None:
            logger.warning("demo14: row not found for %s", code)
            continue
        points = []
        for col_idx in range(1, min(len(year_row), len(data_row))):
            year = _extract_year(year_row[col_idx])
            if year is None or year < 1990:
                continue
            val = _to_float(data_row[col_idx])
            if val is not None:
                points.append(DataPoint(date=date(year, 1, 1), value=round(val / 1000, 2)))
        result[code] = sorted(points, key=lambda p: p.date)

    return result


def parse_pensioners_xlsx(content: bytes) -> list[DataPoint]:
    """Parse Sp_2.1_YYYY.xlsx → total pensioners (тыс. чел.)."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        sheets = wb.sheetnames
        ws = None
        for name in sheets:
            if "РФ" in name or "2014" in name:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.worksheets[-1]

        rows_data = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if len(rows_data) < 5:
        raise ValueError(f"Pensioners: expected >= 5 rows, got {len(rows_data)}")

    year_row = None
    data_row = None
    for i in range(min(10, len(rows_data))):
        row = rows_data[i]
        years_found = sum(1 for c in row[1:15] if _extract_year(c) is not None)
        if years_found >= 3:
            year_row = row
            if i + 1 < len(rows_data):
                data_row = rows_data[i + 1]
            break

    if year_row is None or data_row is None:
        raise ValueError("Pensioners: year/data row not found")

    points = []
    for col_idx in range(1, min(len(year_row), len(data_row))):
        year = _extract_year(year_row[col_idx])
        if year is None:
            continue
        val = _to_float(data_row[col_idx])
        if val is not None:
            points.append(DataPoint(date=date(year, 1, 1), value=round(val, 1)))

    return sorted(points, key=lambda p: p.date)


DEMO_FILES = {
    "births": ("demo21", "births"),
    "deaths": ("demo21", "deaths"),
    "birth-rate": ("demo21", "birth-rate"),
    "death-rate": ("demo21", "death-rate"),
    "working-age-population": ("demo14", "working-age-population"),
    "pop-under-working-age": ("demo14", "pop-under-working-age"),
    "pop-over-working-age": ("demo14", "pop-over-working-age"),
    "pensioners": ("pensioners", None),
}


def _merge_edn_full_year(
    series: dict[str, list[DataPoint]],
    session,
) -> str | None:
    """Дополнить demo21 оперативным полным годом из Edn_12-{YYYY}_t1.xlsx.

    Годовая таблица demo21 обновляется с лагом; ЕДН за декабрь уже даёт
    календарный итог РФ. Добавляем только годы строго новее max(demo21).
    """
    max_year = 0
    for pts in series.values():
        for p in pts:
            if p.date.year > max_year:
                max_year = p.date.year

    current_year = datetime.now().year
    # Полный год N публикуется как EDN_12-N; пробуем N=current-1..max_year+1.
    candidates_years = list(range(current_year - 1, max_year, -1))
    if not candidates_years:
        # всё же проверим, нет ли на каталоге более свежего декабря
        candidates_years = [current_year - 1]

    page_files = list_mediabank_filenames_from_page(DEMOGRAPHY_CATALOG_URL, session=session)
    used_url = None

    for year in candidates_years:
        if year <= max_year:
            continue
        patterns = [
            rf"(?i)Edn_12-{year}_t1\.xlsx",
            rf"(?i)EDN_12-{year}_t1\.xlsx",
            rf"(?i)edn_12-{year}_t1\.xlsx",
        ]
        name = pick_mediabank_filename(page_files, patterns)
        fallback = [f"Edn_12-{year}_t1.xlsx", f"EDN_12-{year}_t1.xlsx"]
        content = None
        url = None
        for cand in ([name] if name else []) + fallback:
            if not cand:
                continue
            got = download_mediabank_bytes(cand, session=session)
            if got:
                content, url = got
                break
        if not content:
            continue

        edn = parse_edn_annual_t1_xlsx(content)
        if not edn or next(iter(edn.values())).date.year <= max_year:
            continue

        for key, point in edn.items():
            if key not in series:
                continue
            existing_years = {p.date.year for p in series[key]}
            if point.date.year in existing_years:
                continue
            series[key].append(point)
            series[key] = sorted(series[key], key=lambda p: p.date)
        used_url = url
        logger.info("demo21: merged EDN full-year %s from %s", year, url)
        break

    return used_url


def _fetch_demo21(session) -> tuple[dict[str, list[DataPoint]], str]:
    current_year = datetime.now().year
    # openpyxl читает только .xlsx; .xls оставляем вне candidates.
    fallback = [f"demo21_{y}.xlsx" for y in range(current_year + 1, current_year - 10, -1)]

    content, url = resolve_mediabank_file(
        catalog_urls=[DEMOGRAPHY_CATALOG_URL],
        name_patterns=[r"(?i)demo21[_-](\d{4})\.xlsx"],
        fallback_filenames=fallback,
        session=session,
    )
    series = parse_demo21_xlsx(content)
    edn_url = _merge_edn_full_year(series, session)
    if edn_url:
        url = f"{url}; {edn_url}"
    return series, url


class RosstatDemoParser(BaseParser):
    parser_type: ClassVar[str] = "rosstat_demo"

    async def _fetch_and_parse(
        self,
        db: AsyncSession,
        indicator: Indicator,
        cfg: dict,
        fetch_log: FetchLog,
    ) -> tuple[list, str]:
        code = indicator.code
        file_type = cfg.get("demo_file", "demo21")

        session = create_session()
        try:
            session.verify = settings.rosstat_ca_cert

            if file_type == "demo21":
                result, used_url = _fetch_demo21(session)
                series_key = cfg.get("demo_series", code)
                return result.get(series_key, []), used_url

            if file_type == "demo14":
                content, used_url = resolve_mediabank_file(
                    catalog_urls=[DEMOGRAPHY_CATALOG_URL],
                    name_patterns=[r"(?i)demo14\.xlsx", r"(?i)demo_14\.xlsx"],
                    fallback_filenames=["demo14.xlsx"],
                    session=session,
                )
                result = parse_demo14_xlsx(content)
                series_key = cfg.get("demo_series", code)
                return result.get(series_key, []), used_url

            if file_type == "pensioners":
                current_year = datetime.now().year
                fallback = [
                    f"Sp_2.1_{y}.xlsx" for y in range(current_year + 1, current_year - 7, -1)
                ]
                content, used_url = resolve_mediabank_file(
                    catalog_urls=[DEMOGRAPHY_CATALOG_URL],
                    name_patterns=[r"(?i)Sp_2\.1_(\d{4})\.xlsx"],
                    fallback_filenames=fallback,
                    session=session,
                )
                return parse_pensioners_xlsx(content), used_url

            raise ValueError(f"Unknown demo_file type: {file_type}")
        finally:
            session.close()

    def _validate(self, points: list, cfg: dict) -> list:
        """demo-парсер исторически фильтровал NaN до bulk_upsert вместо validate_points()."""
        valid = [
            p for p in points
            if isinstance(p.value, (int, float)) and not math.isnan(p.value)
        ]
        if len(valid) < len(points):
            logger.warning(
                "Filtered out %d invalid (NaN/non-numeric) demographic values",
                len(points) - len(valid),
            )
        return valid
