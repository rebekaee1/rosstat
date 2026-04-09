"""ETL: ЦБ РФ — платёжный баланс (bal_of_payments_standart.xlsx) → IndicatorData.

Источник: https://www.cbr.ru/vfs/statistics/credit_statistics/bop/bal_of_payments_standart.xlsx
Лист «Кварталы»:
  Row 6: заголовки — "1 квартал 1994 г.", "2 квартал 1994 г.", …
  Row 14: Экспорт товаров (млн $)
  Row 15: Импорт товаров (млн $)
"""

from __future__ import annotations

import asyncio
import io
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from typing import ClassVar

import openpyxl
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import FetchLog, Indicator, IndicatorData
from app.services.base_parser import BaseParser
from app.services.http_client import create_session
from app.services.upsert import upsert_indicator_data
from app.services.forecast_pipeline import retrain_indicator_forecast
from app.core.cache import cache_invalidate_indicator

logger = logging.getLogger(__name__)

BOP_URL = "https://www.cbr.ru/vfs/statistics/credit_statistics/bop/bal_of_payments_standart.xlsx"

_Q_RE = re.compile(r"(\d)\s*квартал\s*(\d{4})", re.IGNORECASE)

QUARTER_MONTH = {1: 3, 2: 6, 3: 9, 4: 12}


@dataclass
class DataPoint:
    date: date
    value: float


def _parse_quarter_header(header: str | None) -> date | None:
    if not header:
        return None
    m = _Q_RE.search(str(header))
    if not m:
        return None
    q, year = int(m.group(1)), int(m.group(2))
    if q not in QUARTER_MONTH or year < 1990 or year > 2100:
        return None
    return date(year, QUARTER_MONTH[q], 1)


def _find_row_by_label(rows: list[list], label_contains: str, exact_strip: str | None = None) -> int | None:
    for i, row in enumerate(rows):
        cell = str(row[0] or "").strip()
        if exact_strip and cell == exact_strip:
            return i
        if label_contains and label_contains.lower() in cell.lower():
            return i
    return None


def fetch_bop_xlsx() -> tuple[bytes, str]:
    session = create_session()
    try:
        resp = session.get(BOP_URL, timeout=90)
        resp.raise_for_status()
        ct = resp.headers.get("content-type", "").lower()
        if "spreadsheet" not in ct and "openxml" not in ct and resp.status_code == 200:
            logger.warning("BOP unexpected content-type: %s", resp.headers.get("content-type"))
        if resp.content[:4] != b"PK\x03\x04":
            raise ValueError("BOP response is not XLSX")
        logger.info("Downloaded BOP XLSX: %d KB", len(resp.content) // 1024)
        return resp.content, BOP_URL
    finally:
        session.close()


def parse_bop_xlsx(content: bytes, target: str) -> list[DataPoint]:
    """Parse BOP standard XLSX.

    target: "exports" → row «Экспорт» under «Товары»
            "imports" → row «Импорт» under «Товары»
            "trade-balance" → row «Товары » (with trailing space in source)
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = None
        for name in wb.sheetnames:
            if "квартал" in name.lower():
                ws = wb[name]
                break
        if ws is None:
            ws = wb.worksheets[0]

        rows_data: list[list] = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append(list(row))
    finally:
        wb.close()

    dates: list[tuple[int, date]] = []
    header_row_idx = None
    for ri in range(4, min(10, len(rows_data))):
        for ci in range(1, min(20, len(rows_data[ri]))):
            d = _parse_quarter_header(rows_data[ri][ci])
            if d:
                header_row_idx = ri
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        raise ValueError("BOP XLSX: no quarter headers found")

    for ci in range(1, len(rows_data[header_row_idx])):
        d = _parse_quarter_header(rows_data[header_row_idx][ci])
        if d:
            dates.append((ci, d))

    if not dates:
        raise ValueError("BOP XLSX: no valid quarter dates")

    goods_section_start = None
    for i, row in enumerate(rows_data):
        cell = str(row[0] or "").strip()
        if cell.startswith("Товары") and "услуг" not in cell.lower():
            goods_section_start = i
            break

    if goods_section_start is None:
        raise ValueError("BOP XLSX: 'Товары' section not found")

    data_row_idx = None

    if target in ("exports", "imports", "trade-balance"):
        search_start = goods_section_start
        if target == "exports":
            for i in range(search_start, min(search_start + 5, len(rows_data))):
                cell = str(rows_data[i][0] or "").strip()
                if cell == "Экспорт":
                    data_row_idx = i
                    break
        elif target == "imports":
            for i in range(search_start, min(search_start + 5, len(rows_data))):
                cell = str(rows_data[i][0] or "").strip()
                if cell == "Импорт":
                    data_row_idx = i
                    break
        elif target == "trade-balance":
            data_row_idx = goods_section_start
    elif target in ("services-exports", "services-imports", "services-balance"):
        services_row = None
        for i, row in enumerate(rows_data):
            cell = str(row[0] or "").strip()
            if cell == "Услуги" and i > (goods_section_start or 0):
                services_row = i
                break
        if services_row is None:
            raise ValueError("BOP XLSX: 'Услуги' section not found")
        if target == "services-balance":
            data_row_idx = services_row
        elif target == "services-exports":
            for i in range(services_row, min(services_row + 3, len(rows_data))):
                cell = str(rows_data[i][0] or "").strip()
                if cell == "Экспорт":
                    data_row_idx = i
                    break
        elif target == "services-imports":
            for i in range(services_row, min(services_row + 3, len(rows_data))):
                cell = str(rows_data[i][0] or "").strip()
                if cell == "Импорт":
                    data_row_idx = i
                    break
    elif target == "fdi-net":
        for i, row in enumerate(rows_data):
            cell = str(row[0] or "").strip()
            if cell == "Прямые инвестиции" and i > 250:
                data_row_idx = i
                break
    elif target == "current-account-bop":
        for i, row in enumerate(rows_data):
            cell = str(row[0] or "").strip()
            if cell.startswith("Счет текущих операций"):
                data_row_idx = i
                break
    else:
        raise ValueError(f"Unknown BOP target: {target}")

    if data_row_idx is None:
        raise ValueError(f"BOP XLSX: row for '{target}' not found")

    row = rows_data[data_row_idx]
    points: list[DataPoint] = []
    for ci, d in dates:
        val = row[ci] if ci < len(row) else None
        if val is not None:
            try:
                points.append(DataPoint(date=d, value=round(float(val), 2)))
            except (ValueError, TypeError):
                pass

    points.sort(key=lambda p: p.date)
    return points


class CbrBopParser(BaseParser):
    parser_type: ClassVar[str] = "cbr_bop_xlsx"

    async def run(self, db: AsyncSession, indicator: Indicator, fetch_log: FetchLog) -> None:
        code = indicator.code
        try:
            content, final_url = await asyncio.to_thread(fetch_bop_xlsx)
            fetch_log.source_url = final_url[:500]

            cfg = indicator.model_config_json or {}
            target = cfg.get("bop_target", code)

            points = await asyncio.to_thread(parse_bop_xlsx, content, target)

            if not points:
                logger.warning("No data points parsed for %s", code)
                fetch_log.status = "no_new_data"
                fetch_log.error_message = "BOP parser returned 0 data points"
                fetch_log.completed_at = datetime.now(timezone.utc)
                await db.commit()
                return

            count_before = (await db.execute(
                select(func.count(IndicatorData.id))
                .where(IndicatorData.indicator_id == indicator.id)
            )).scalar() or 0

            for p in points:
                await db.execute(upsert_indicator_data(indicator.id, p.date, p.value))

            await db.flush()
            count_after = (await db.execute(
                select(func.count(IndicatorData.id))
                .where(IndicatorData.indicator_id == indicator.id)
            )).scalar() or 0

            records_added = count_after - count_before
            fetch_log.records_added = records_added
            logger.info("BOP '%s' (%s): +%d rows (total %d)", code, target, records_added, count_after)

            steps = int(cfg.get("forecast_steps", 0) or 0)
            if steps > 0 and records_added > 0:
                await retrain_indicator_forecast(db, indicator)

            if records_added > 0:
                await cache_invalidate_indicator(code)

            fetch_log.status = "success" if records_added > 0 else "no_new_data"
            fetch_log.completed_at = datetime.now(timezone.utc)
            await db.commit()

        except Exception as e:
            logger.exception("ETL failed for '%s'", code)
            await db.rollback()
            fetch_log.status = "failed"
            fetch_log.error_message = str(e)[:500]
            fetch_log.completed_at = datetime.now(timezone.utc)
            db.add(fetch_log)
            await db.commit()
