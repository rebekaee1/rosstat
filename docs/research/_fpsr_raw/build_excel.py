#!/usr/bin/env python3
"""Сборка единого Excel из agent01..agent10.json (ФПСР субъекты)."""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

RAW = Path(__file__).resolve().parent
OUT = RAW.parent / f"russia-fpsr-official-sources-{date.today().isoformat()}.xlsx"

HDR_SRC = [
    "Субъект",
    "Полное название",
    "Роль",
    "Официальный сайт",
    "Портал статистики",
    "Открытый доступ",
    "API",
    "Форматы",
    "Оценка объёма",
    "Число datasets",
    "Частоты (сводно)",
    "ETL 1-5",
    "Заметки",
    "Агент",
]

HDR_DS = [
    "Субъект",
    "Семейство / индикатор",
    "Частота (D/W/M/Q/Y)",
    "История с",
    "История по",
    "Единица",
    "Декомпозиция",
    "Формат",
    "URL",
    "Заметки",
    "Агент",
]


def load_agents() -> list[dict]:
    items = []
    for i in range(1, 11):
        p = RAW / f"agent{i:02d}.json"
        if not p.exists():
            print(f"MISSING {p.name}")
            continue
        data = json.loads(p.read_text(encoding="utf-8"))
        items.append(data)
        print(f"OK {p.name}: subjects={len(data.get('subjects', []))}")
    return items


def style_header(ws, n_cols: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, n_cols + 1):
        cell = ws.cell(1, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws, max_width: int = 56) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for row in range(1, min(ws.max_row + 1, 200)):
            val = ws.cell(row, col).value
            if val is None:
                continue
            width = max(width, min(max_width, len(str(val)) + 2))
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    agents = load_agents()
    if not agents:
        raise SystemExit("No agent JSON files found")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Источники"
    ws2 = wb.create_sheet("Индикаторы")

    for i, h in enumerate(HDR_SRC, 1):
        ws1.cell(1, i, h)
    for i, h in enumerate(HDR_DS, 1):
        ws2.cell(1, i, h)

    r1 = 2
    r2 = 2
    n_subj = 0
    n_ds = 0

    for data in agents:
        agent_id = str(data.get("agent", "?"))
        for s in data.get("subjects", []):
            n_subj += 1
            datasets = s.get("datasets") or []
            freqs = sorted(
                {
                    (d.get("frequency") or "").strip()
                    for d in datasets
                    if (d.get("frequency") or "").strip()
                }
            )
            formats = s.get("formats") or []
            if isinstance(formats, list):
                formats_s = ", ".join(str(x) for x in formats)
            else:
                formats_s = str(formats)

            row = [
                s.get("code") or "",
                s.get("full_name") or "",
                s.get("role") or "",
                s.get("official_site") or "",
                s.get("stats_portal") or "",
                s.get("open_access") or "",
                s.get("api") or "",
                formats_s,
                s.get("volume_estimate") or "",
                len(datasets),
                " / ".join(freqs),
                s.get("etl_difficulty_1_5") or "",
                s.get("notes") or "",
                agent_id,
            ]
            for c, v in enumerate(row, 1):
                ws1.cell(r1, c, v)
            r1 += 1

            for d in datasets:
                n_ds += 1
                drow = [
                    s.get("code") or "",
                    d.get("family") or "",
                    d.get("frequency") or "",
                    d.get("history_from") or "",
                    d.get("history_to") or "",
                    d.get("unit") or "",
                    d.get("decomposition") or "",
                    d.get("format") or "",
                    d.get("url") or "",
                    d.get("notes") or "",
                    agent_id,
                ]
                for c, v in enumerate(drow, 1):
                    ws2.cell(r2, c, v)
                r2 += 1

    style_header(ws1, len(HDR_SRC))
    style_header(ws2, len(HDR_DS))
    autosize(ws1)
    autosize(ws2)

    wb.save(OUT)
    print(f"WROTE {OUT}")
    print(f"subjects={n_subj} datasets={n_ds}")


if __name__ == "__main__":
    main()
